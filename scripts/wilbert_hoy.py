"""ETIQUETAS WILBERT HOY — solo cuenta Wilbert, 60 días de ventana, formato 4x6.
Captura TODO lo que esté ready_to_ship/handling, no enviado, no entregado.
Loggea substatus de cada shipment para diagnóstico.
"""
import os, io, time, requests
from datetime import datetime, timedelta, timezone
from collections import defaultdict, Counter
from pypdf import PdfReader, PdfWriter, PageObject, Transformation
from reportlab.pdfgen import canvas
from reportlab.lib.colors import Color
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pdf2image import convert_from_bytes

APP_ID="5211907102822632"
APP_SECRET=os.environ["MELI_APP_SECRET"]
RT = os.environ["MELI_REFRESH_TOKEN_WILBERT"]
TZ = timezone(timedelta(hours=-6))
PAGE_W = 4*72
PAGE_H = 6*72
SUB_HARD_EXCLUDE = {"shipped","delivered","not_delivered","returned"}


def tok(rt):
    r=requests.post("https://api.mercadolibre.com/oauth/token",data={
        "grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":rt}).json()
    return r.get("access_token")


def clean_title(title):
    t=(title or "").strip()
    for w in ["Bocina ","bocina ","Parlante ","parlante ","Altavoz ","altavoz ","Speaker ","speaker ",
              "JBL ","jbl ","Jbl ","Sony ","SONY ","Bose ","BOSE "]:
        t=t.replace(w,"")
    tl=t.lower()
    if "go 4" in tl or "go4" in tl: model="Go 4"
    elif "go 3" in tl or "go3" in tl: model="Go 3"
    elif "clip 5" in tl or "clip5" in tl: model="Clip 5"
    elif "charge 6" in tl or "charge6" in tl: model="Charge 6"
    elif "flip 7" in tl or "flip7" in tl: model="Flip 7"
    elif "grip" in tl: model="Grip"
    elif "xb100" in tl: model="Sony XB100"
    elif "soundlink" in tl: model="Bose SoundLink"
    else: model=t[:30]
    col_map=[("camuflaj","Camuflaje"),("camo","Camuflaje"),("azul marino","Azul Marino"),
             ("aqua","Aqua"),("celeste","Celeste"),("negr","Negro"),("black","Negro"),
             ("roj","Rojo"),(" red","Rojo"),("rosa","Rosa"),("pink","Rosa"),
             ("morad","Morado"),("violeta","Morado"),("purple","Morado"),
             (" azul","Azul"),(" blue","Azul")]
    color=None
    for k,v in col_map:
        if k in (" "+tl):
            color=v; break
    return f"{model} {color}" if color else model


def detect_content_bbox(pdf_bytes, page_idx=0):
    try:
        imgs = convert_from_bytes(pdf_bytes, dpi=72,
                                  first_page=page_idx+1, last_page=page_idx+1)
        if not imgs: return None
        img = imgs[0].convert("L")
        bw = img.point(lambda p: 0 if p < 245 else 255, mode="L")
        from PIL import ImageOps
        inv = ImageOps.invert(bw)
        bbox = inv.getbbox()
        if not bbox: return None
        x0, y0_top, x1, y1_top = bbox
        img_w, img_h = img.size
        pdf_y0 = img_h - y1_top
        pdf_y1 = img_h - y0_top
        m = 2
        return (max(0, x0-m), max(0, pdf_y0-m),
                min(img_w, x1-x0+2*m), min(img_h, pdf_y1-pdf_y0+2*m))
    except Exception as e:
        print(f"   bbox warn: {type(e).__name__}: {str(e)[:80]}")
        return None


# === FASE 1: shipments Wilbert (60 días) ===
NOW = datetime.now(timezone.utc)
START = NOW - timedelta(days=60)
print(f"=== Wilbert: ventana de orders {START.date()} a {NOW.date()} ===")
at = tok(RT)
if not at:
    raise SystemExit("No pude obtener access_token de Wilbert")
H = {"Authorization": f"Bearer {at}"}
me = requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=15).json()
uid = me.get("id")
print(f"  uid={uid}")
orders=[]; offset=0
while True:
    r=requests.get("https://api.mercadolibre.com/orders/search",headers=H,timeout=20,
        params={"seller":uid,"order.status":"paid",
                "order.date_created.from":START.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                "order.date_created.to":NOW.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                "limit":50,"offset":offset}).json()
    res=r.get("results",[])
    if not res: break
    orders.extend(res); offset+=len(res)
    if offset>=r.get("paging",{}).get("total",0): break
print(f"  orders 'paid' totales: {len(orders)}")
obs={}
for o in orders:
    sid=(o.get("shipping") or {}).get("id")
    if sid: obs[sid]=o
print(f"  shipping ids únicos: {len(obs)}")

shipments=[]
status_count=Counter()
sub_count=Counter()
for sid, ord_o in obs.items():
    try:
        sh=requests.get(f"https://api.mercadolibre.com/shipments/{sid}",headers=H,timeout=10).json()
        st=sh.get("status"); sub=sh.get("substatus") or "(none)"
        status_count[st]+=1
        sub_count[(st,sub)]+=1
        # Filtro: status de "no enviado"
        if st in ("delivered","cancelled"): continue
        if st == "shipped" and sub in ("shipped","delivered","not_delivered"): continue
        if sub in SUB_HARD_EXCLUDE: continue
        items=ord_o.get("order_items",[])
        comp_lines=[f"{clean_title((it.get('item') or {}).get('title',''))} x{it.get('quantity',1)}" for it in items]
        buyer=(ord_o.get("buyer") or {}).get("nickname","?")
        deadline=None
        try:
            sla=requests.get(f"https://api.mercadolibre.com/shipments/{sid}/sla",headers=H,timeout=8).json()
            ed=sla.get("expected_date")
            if ed: deadline=datetime.fromisoformat(ed.replace("Z","+00:00")).astimezone(TZ)
        except: pass
        shipments.append({
            "sid":sid,"comp_lines":comp_lines,"buyer":buyer,
            "status":st,"substatus":sub,
            "deadline": deadline.strftime("%a %d %H:%M") if deadline else "s/d",
            "tracking": sh.get("tracking_number","")
        })
        time.sleep(0.04)
    except Exception as e:
        print(f"  err {sid}: {str(e)[:80]}")

print(f"\n=== Distribución status / substatus en {len(obs)} shipments ===")
for (st,sub),n in sub_count.most_common():
    print(f"  {n:4}  status={st!r:20}  substatus={sub!r}")
print(f"\nCandidatos a etiqueta (no enviados): {len(shipments)}")
shipments.sort(key=lambda s: ("/".join(s["comp_lines"]), s["sid"]))


# === FASE 2: PDF 4x6 con header integrado ===
def render_header(s, header_h):
    buf=io.BytesIO()
    c=canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    c.setFillColor(Color(1, 0.96, 0.74))
    c.rect(0, PAGE_H-header_h, PAGE_W, header_h, fill=1, stroke=0)
    c.setFillColorRGB(0,0,0)
    cx = PAGE_W/2.0
    c.setFont("Helvetica-Bold", 7.5)
    c.drawCentredString(cx, PAGE_H-11, f"[WILBERT] {s['buyer'][:30]} | Ship:{s['sid']}")
    big = s["comp_lines"][:4]; n=len(big); lh=16
    block_top=PAGE_H-18; block_bot=PAGE_H-header_h+4
    block_h=block_top-block_bot; text_h=n*lh
    first_y = block_top - (block_h - text_h)/2.0 - 12
    c.setFont("Helvetica-Bold", 14)
    y=first_y
    for line in big:
        c.drawCentredString(cx, y, line[:30]); y-=lh
    c.showPage(); c.save()
    buf.seek(0)
    return PdfReader(buf).pages[0]

print("\n=== Generando PDF 4x6 Wilbert ===")
writer = PdfWriter()
fail = []
for s in shipments:
    try:
        r=requests.get("https://api.mercadolibre.com/shipment_labels",
                      headers=H, params={"shipment_ids":s["sid"],"response_type":"pdf"},
                      timeout=30)
        if r.status_code!=200 or not r.headers.get("content-type","").lower().startswith("application/pdf"):
            fail.append({"sid":s["sid"],"sub":s["substatus"],"reason":r.text[:200]})
            print(f"  err {s['sid']} sub={s['substatus']}: HTTP {r.status_code} {r.text[:150]}")
            continue
        raw = r.content
        lbl_pdf=PdfReader(io.BytesIO(raw))
        for pidx, label_page in enumerate(lbl_pdf.pages):
            box = label_page.cropbox if label_page.cropbox else label_page.mediabox
            lbl_x0=float(box.left); lbl_y0=float(box.bottom)
            lbl_w=float(box.width); lbl_h=float(box.height)
            cb = detect_content_bbox(raw, pidx)
            if cb:
                cx0,cy0,cw,ch = cb
                lbl_x0=float(box.left)+float(cx0)
                lbl_y0=float(box.bottom)+float(cy0)
                lbl_w=float(cw); lbl_h=float(ch)
            n_lines=min(len(s["comp_lines"]),4)
            header_h = 22 + n_lines*16
            label_area_h = PAGE_H - header_h
            new_page = PageObject.create_blank_page(width=PAGE_W, height=PAGE_H)
            sx = PAGE_W/lbl_w; sy = label_area_h/lbl_h
            op = (Transformation()
                  .translate(-lbl_x0, -lbl_y0)
                  .scale(sx, sy))
            new_page.merge_transformed_page(label_page, op)
            new_page.merge_page(render_header(s, header_h))
            writer.add_page(new_page)
    except Exception as e:
        fail.append({"sid":s["sid"],"sub":s.get("substatus",""),"reason":f"{type(e).__name__}: {str(e)[:200]}"})
        print(f"  err {s['sid']}: {type(e).__name__}: {str(e)[:140]}")
    time.sleep(0.08)

with open("ETIQUETAS_WILBERT_HOY_4x6.pdf","wb") as f: writer.write(f)
print(f"\n✅ PDF Wilbert: {len(writer.pages)} págs | Fallidas: {len(fail)}")
for f_ in fail:
    print(f"   FALLIDA sid={f_['sid']} sub={f_['sub']} → {f_['reason'][:140]}")


# === FASE 3: XLSX ===
wb=Workbook(); ws=wb.active; ws.title="Wilbert Hoy"
hf=PatternFill("solid", fgColor="2C3E50")
hF=Font(bold=True, color="FFFFFF", size=11)
center=Alignment(horizontal="center", vertical="center", wrap_text=True)
border=Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
headers=["#","Shipment ID","Comprador","PRODUCTO","Status","Substatus","Tracking","Deadline"]
for col,h in enumerate(headers,1):
    c=ws.cell(row=1,column=col,value=h)
    c.fill=hf; c.font=hF; c.alignment=center; c.border=border
fill = PatternFill("solid", fgColor="FAD7A0")
for i,s in enumerate(shipments,1):
    row=[i, s["sid"], s["buyer"], "\n".join(s["comp_lines"]),
         s["status"], s["substatus"], s["tracking"], s["deadline"]]
    for col,val in enumerate(row,1):
        c=ws.cell(row=i+1,column=col,value=val)
        c.fill=fill; c.border=border
        c.alignment=Alignment(vertical="center", wrap_text=True)
        if col==4: c.font=Font(bold=True, size=10)
widths={1:5,2:14,3:24,4:50,5:14,6:18,7:24,8:14}
for col,w in widths.items(): ws.column_dimensions[chr(64+col)].width=w
for r in range(2, len(shipments)+2): ws.row_dimensions[r].height=28
ws.freeze_panes="A2"

ws2=wb.create_sheet("Resumen producto")
ws2["A1"]="PRODUCTO"; ws2["B1"]="CANT"
for c in ["A1","B1"]: ws2[c].fill=hf; ws2[c].font=hF
prod=Counter("/".join(s["comp_lines"]) for s in shipments)
for i,(p,n) in enumerate(sorted(prod.items(), key=lambda x:-x[1]),2):
    ws2.cell(row=i,column=1,value=p).font=Font(bold=True)
    ws2.cell(row=i,column=2,value=n).alignment=Alignment(horizontal="center")
ws2.column_dimensions["A"].width=55; ws2.column_dimensions["B"].width=10

wb.save("DESPACHO_WILBERT.xlsx")
print(f"✅ XLSX: DESPACHO_WILBERT.xlsx")
