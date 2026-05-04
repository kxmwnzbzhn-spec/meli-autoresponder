"""Genera PDF y XLSX de etiquetas LUNES con producto + cantidad VISIBLE EN CADA PAGINA.
Cada etiqueta lleva un header impreso encima: 'PRODUCTO: <composicion> | Cuenta: <X> | Comprador: <Y>'
Asi el equipo puede armar los paquetes correctamente.
"""
import os, io, time, requests
from datetime import datetime, timedelta, timezone
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm

APP_ID = "5211907102822632"
APP_SECRET = os.environ["MELI_APP_SECRET"]

ACCS = {
    "Juan":     os.environ.get("MELI_REFRESH_TOKEN_JUAN") or os.environ.get("MELI_REFRESH_TOKEN"),
    "Raymundo": os.environ.get("MELI_REFRESH_TOKEN_RAYMUNDO"),
    "Claribel": os.environ.get("MELI_REFRESH_TOKEN_CLARIBEL"),
    "Asva":     os.environ.get("MELI_REFRESH_TOKEN_ASVA"),
    "Mildred":  os.environ.get("MELI_REFRESH_TOKEN_MILDRED"),
    "Dilcie":   os.environ.get("MELI_REFRESH_TOKEN_DILCIE"),
    "Bren":     os.environ.get("MELI_REFRESH_TOKEN_BREN"),
}
TZ_CDMX = timezone(timedelta(hours=-6))
LIMIT_DAY = datetime.fromisoformat("2026-05-04").replace(hour=23,minute=59,tzinfo=TZ_CDMX)


def tok(rt):
    r = requests.post("https://api.mercadolibre.com/oauth/token", data={
        "grant_type":"refresh_token","client_id":APP_ID,
        "client_secret":APP_SECRET,"refresh_token":rt}).json()
    return r.get("access_token")


# === FASE 1: Recolectar shipments LUNES con productos + tokens ===
NOW = datetime.now(timezone.utc)
START = NOW - timedelta(days=10)

shipments_data = []  # list of {sid, account, token, composition, buyer, deadline}

for acc, rt in ACCS.items():
    if not rt: continue
    print(f"\n=== {acc} ===")
    at = tok(rt)
    if not at: continue
    H = {"Authorization": f"Bearer {at}"}
    me = requests.get("https://api.mercadolibre.com/users/me", headers=H, timeout=15).json()
    uid = me.get("id")
    if not uid: continue

    orders = []
    offset = 0
    while True:
        r = requests.get("https://api.mercadolibre.com/orders/search",
            headers=H, timeout=20,
            params={"seller":uid, "order.status":"paid",
                    "order.date_created.from":START.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "order.date_created.to":NOW.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "limit":50,"offset":offset}).json()
        res = r.get("results",[])
        if not res: break
        orders.extend(res)
        offset += len(res)
        if offset >= r.get("paging",{}).get("total",0): break

    order_by_ship = {}
    for o in orders:
        sid = (o.get("shipping") or {}).get("id")
        if sid: order_by_ship[sid] = o

    matches = 0
    for sid, ord_o in order_by_ship.items():
        try:
            sh = requests.get(f"https://api.mercadolibre.com/shipments/{sid}",
                              headers=H, timeout=10).json()
            status = sh.get("status")
            substatus = sh.get("substatus")
            if status not in ("ready_to_ship","handling"): continue
            if substatus in ("shipped","ready_to_pickup"): continue

            deadline = None
            try:
                sla = requests.get(f"https://api.mercadolibre.com/shipments/{sid}/sla",
                                   headers=H, timeout=8).json()
                ed = sla.get("expected_date")
                if ed:
                    deadline = datetime.fromisoformat(ed.replace("Z","+00:00")).astimezone(TZ_CDMX)
            except: pass
            if not deadline:
                hist = sh.get("status_history") or {}
                dh = hist.get("date_handling")
                if dh:
                    dh_dt = datetime.fromisoformat(dh.replace("Z","+00:00"))
                    deadline = (dh_dt + timedelta(hours=48)).astimezone(TZ_CDMX)
            if not deadline or deadline > LIMIT_DAY: continue

            items = ord_o.get("order_items",[])
            comp_parts = [f"{(it.get('item') or {}).get('title','')[:40]} x{it.get('quantity',1)}"
                          for it in items]
            comp = " + ".join(comp_parts)
            buyer = (ord_o.get("buyer") or {}).get("nickname","?")

            shipments_data.append({
                "sid": sid, "account": acc, "token": at,
                "composition": comp, "buyer": buyer,
                "deadline": deadline.strftime("%m-%d %H:%M"),
                "tracking": sh.get("tracking_number",""),
            })
            matches += 1
            time.sleep(0.05)
        except Exception as e:
            print(f"  err {sid}: {str(e)[:60]}")
    print(f"  {acc}: {matches} matches")

print(f"\nTotal shipments LUNES: {len(shipments_data)}")

# === FASE 2: Ordenar por composicion para picking eficiente ===
shipments_data.sort(key=lambda s: (s["composition"], s["account"]))

# === FASE 3: Generar XLSX ===
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb = Workbook()
ws = wb.active
ws.title = "Etiquetas Lunes"

headers = ["#","Cuenta","Shipment ID","Comprador","PRODUCTO A EMPACAR","Tracking","Deadline"]
hf = PatternFill("solid", fgColor="2C3E50")
hF = Font(bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = hf; c.font = hF; c.alignment = center; c.border = border

# Color por cuenta
acc_colors = {"Juan":"D5E8D4","Raymundo":"DAE8FC","Claribel":"FFE6CC","Asva":"E1D5E7"}

for i, s in enumerate(shipments_data, 1):
    fill = PatternFill("solid", fgColor=acc_colors.get(s["account"],"FFFFFF"))
    row_vals = [i, s["account"], s["sid"], s["buyer"], s["composition"], s["tracking"], s["deadline"]]
    for col, val in enumerate(row_vals, 1):
        c = ws.cell(row=i+1, column=col, value=val)
        c.fill = fill; c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if col == 5:  # PRODUCTO
            c.font = Font(bold=True, size=10)

widths = {1:5, 2:11, 3:14, 4:24, 5:55, 6:24, 7:13}
for col, w in widths.items():
    ws.column_dimensions[chr(64+col)].width = w
for r in range(2, len(shipments_data)+2):
    ws.row_dimensions[r].height = 28
ws.freeze_panes = "A2"

# Resumen por composicion
ws2 = wb.create_sheet("Resumen por producto")
ws2.cell(row=1,column=1,value="PRODUCTO").font = hF
ws2.cell(row=1,column=2,value="CANTIDAD").font = hF
ws2.cell(row=1,column=3,value="CUENTAS").font = hF
ws2.cell(row=1,column=1).fill = hf
ws2.cell(row=1,column=2).fill = hf
ws2.cell(row=1,column=3).fill = hf
from collections import defaultdict
comp_count = defaultdict(lambda: {"qty":0, "accounts":set()})
for s in shipments_data:
    comp_count[s["composition"]]["qty"] += 1
    comp_count[s["composition"]]["accounts"].add(s["account"])
for r, (comp, info) in enumerate(sorted(comp_count.items(), key=lambda x:-x[1]["qty"]), 2):
    ws2.cell(row=r,column=1,value=comp).font = Font(bold=True)
    ws2.cell(row=r,column=2,value=info["qty"]).alignment = Alignment(horizontal="center")
    ws2.cell(row=r,column=3,value=", ".join(sorted(info["accounts"])))
ws2.column_dimensions["A"].width = 60
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 25

xlsx_out = "etiquetas_lunes_con_productos.xlsx"
wb.save(xlsx_out)
print(f"\n✅ XLSX: {xlsx_out}")

# === FASE 4: Generar PDF combinando etiquetas con header de producto ===
print(f"\n=== Generando PDF con headers ===")

# Agrupar por cuenta para fetch eficiente de labels
from collections import defaultdict as dd
by_acc = dd(list)
for s in shipments_data:
    by_acc[s["account"]].append(s)

writer = PdfWriter()
fail = []
for acc, ships in by_acc.items():
    if not ships: continue
    token = ships[0]["token"]
    H = {"Authorization": f"Bearer {token}"}
    print(f"  {acc}: {len(ships)} etiquetas")
    BATCH = 1  # 1 a la vez para poder mappear shipment->paginas
    for s in ships:
        try:
            r = requests.get("https://api.mercadolibre.com/shipment_labels",
                            headers=H, params={"shipment_ids":s["sid"],"response_type":"pdf"},
                            timeout=30)
            if r.status_code == 200 and r.headers.get("content-type","").lower().startswith("application/pdf"):
                # Crear page con header
                buf = io.BytesIO()
                c = canvas.Canvas(buf, pagesize=letter)
                c.setFillColorRGB(0,0,0)
                c.setFont("Helvetica-Bold", 18)
                c.drawString(20*mm, 250*mm, f"📦 {s['composition'][:60]}")
                c.setFont("Helvetica", 12)
                c.drawString(20*mm, 240*mm, f"Cuenta: {s['account']}  |  Shipment: {s['sid']}")
                c.drawString(20*mm, 232*mm, f"Comprador: {s['buyer'][:50]}")
                c.drawString(20*mm, 224*mm, f"Deadline: {s['deadline']}  |  Tracking: {s['tracking']}")
                c.line(20*mm, 218*mm, 195*mm, 218*mm)
                c.setFont("Helvetica-Bold", 14)
                c.drawString(20*mm, 210*mm, "↓ ETIQUETA DE ENVIO ↓")
                c.showPage()
                c.save()
                buf.seek(0)

                # Append: primero el header, despues la etiqueta original
                hdr_pdf = PdfReader(buf)
                writer.add_page(hdr_pdf.pages[0])
                lbl_pdf = PdfReader(io.BytesIO(r.content))
                for p in lbl_pdf.pages:
                    writer.add_page(p)
            else:
                fail.append(s["sid"])
        except Exception as e:
            fail.append(s["sid"])
            print(f"    err {s['sid']}: {e}")
        time.sleep(0.1)

pdf_out = "ETIQUETAS_LUNES_CON_PRODUCTOS.pdf"
with open(pdf_out,"wb") as f:
    writer.write(f)
print(f"\n✅ PDF: {pdf_out} ({len(writer.pages)} páginas)")
print(f"  Etiquetas en orden por producto + header con composición/comprador antes de cada etiqueta")
print(f"  Fallidas: {len(fail)}")
