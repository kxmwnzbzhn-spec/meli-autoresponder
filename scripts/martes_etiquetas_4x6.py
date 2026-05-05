"""ETIQUETAS TODAS PENDIENTES — formato 4x6 in (288x432 pt).
Hoy + mañana en adelante. Status ready_to_ship/handling que aún no se enviaron.
Header amarillo integrado arriba + etiqueta MELI escalada para llenar 4x6.
"""
import os, io, time, requests
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from pypdf import PdfReader, PdfWriter, PageObject, Transformation
from reportlab.pdfgen import canvas
from reportlab.lib.colors import Color
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pdf2image import convert_from_bytes
from PIL import Image

APP_ID="5211907102822632"
APP_SECRET=os.environ["MELI_APP_SECRET"]
ACCS={
    "Juan":     os.environ.get("MELI_REFRESH_TOKEN_JUAN") or os.environ.get("MELI_REFRESH_TOKEN"),
    "Raymundo": os.environ.get("MELI_REFRESH_TOKEN_RAYMUNDO"),
    "Wilbert":  os.environ.get("MELI_REFRESH_TOKEN_WILBERT"),
    "Claribel": os.environ.get("MELI_REFRESH_TOKEN_CLARIBEL"),
    "Asva":     os.environ.get("MELI_REFRESH_TOKEN_ASVA"),
    "Mildred":  os.environ.get("MELI_REFRESH_TOKEN_MILDRED"),
    "Dilcie":   os.environ.get("MELI_REFRESH_TOKEN_DILCIE"),
    "Bren":     os.environ.get("MELI_REFRESH_TOKEN_BREN"),
    "Yc_New":   os.environ.get("MELI_REFRESH_TOKEN_YC_NEW"),
}
TZ=timezone(timedelta(hours=-6))
# Hoy CDMX
TODAY_CDMX = datetime.now(TZ).replace(hour=0, minute=0, second=0, microsecond=0)
END_TODAY  = TODAY_CDMX.replace(hour=23, minute=59)
# Substatuses excluidos (ya no se pueden imprimir o ya están en transito)
SUB_EXCLUDE = {"shipped","ready_to_pickup","delivered","not_delivered","returned","picked_up"}

# Tamaño impresora 4x6 in
PAGE_W = 4 * 72   # 288 pt
PAGE_H = 6 * 72   # 432 pt

# 0 = todas. >0 = limitar (modo preview)
LIMIT = int(os.environ.get("LIMIT") or "0")


def detect_content_bbox(pdf_bytes, page_idx=0):
    """Rasteriza la página y devuelve (x0, y0, w, h) en puntos PDF
    correspondientes al recuadro de pixeles no-blancos. Si falla, None."""
    try:
        imgs = convert_from_bytes(pdf_bytes, dpi=72,
                                  first_page=page_idx+1, last_page=page_idx+1)
        if not imgs: return None
        img = imgs[0].convert("L")
        # threshold: <245 => content
        bw = img.point(lambda p: 0 if p < 245 else 255, mode="L")
        # bbox of black pixels (content)
        # PIL.getbbox returns bbox of non-zero, but we want non-255.
        # Invert: now content=255.
        from PIL import ImageOps
        inv = ImageOps.invert(bw)
        bbox = inv.getbbox()
        if not bbox: return None
        x0, y0_top, x1, y1_top = bbox  # PIL coords: y=0 top
        img_w, img_h = img.size
        # PDF coords: y=0 bottom
        pdf_y0 = img_h - y1_top
        pdf_y1 = img_h - y0_top
        # margen pequeño
        m = 2
        return (max(0, x0-m), max(0, pdf_y0-m),
                min(img_w, x1-x0+2*m), min(img_h, pdf_y1-pdf_y0+2*m))
    except Exception as e:
        print(f"   detect_bbox warn: {type(e).__name__}: {str(e)[:80]}")
        return None


def tok(rt):
    r=requests.post("https://api.mercadolibre.com/oauth/token",data={
        "grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":rt}).json()
    return r.get("access_token")


def clean_title(title):
    t=(title or "").strip()
    for w in ["Bocina ","bocina ","Parlante ","parlante ","Altavoz ","altavoz ","Speaker ","speaker ",
              "JBL ","jbl ","Jbl ","Sony ","SONY ","Bose ","BOSE "]:
        t=t.replace(w,"")
    tl=t.lower()
    if "go 4" in tl or "go4" in tl: model="Go 4"
    elif "go 3" in tl or "go3" in tl: model="Go 3"
    elif "clip 5" in tl or "clip5" in tl: model="Clip 5"
    elif "charge 6" in tl or "charge6" in tl: model="Charge 6"
    elif "flip 7" in tl or "flip7" in tl: model="Flip 7"
    elif "grip" in tl: model="Grip"
    elif "xb100" in tl: model="Sony XB100"
    elif "soundlink" in tl: model="Bose SoundLink"
    else: model=t[:30]
    col_map=[("camuflaj","Camuflaje"),("camo","Camuflaje"),("azul marino","Azul Marino"),
             ("aqua","Aqua"),("celeste","Celeste"),("negr","Negro"),("black","Negro"),
             ("roj","Rojo"),(" red","Rojo"),("rosa","Rosa"),("pink","Rosa"),
             ("morad","Morado"),("violeta","Morado"),("purple","Morado"),
             (" azul","Azul"),(" blue","Azul")]
    color=None
    for k,v in col_map:
        if k in (" "+tl):
            color=v; break
    return f"{model} {color}" if color else model


# === FASE 1: shipments MARTES ===
NOW=datetime.now(timezone.utc); START=NOW-timedelta(days=10)
shipments=[]
for acc, rt in ACCS.items():
    if not rt: continue
    print(f"=== {acc} ===")
    at=tok(rt)
    if not at: continue
    H={"Authorization":f"Bearer {at}"}
    me=requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=15).json()
    uid=me.get("id")
    if not uid: continue
    orders=[]; offset=0
    while True:
        r=requests.get("https://api.mercadolibre.com/orders/search",headers=H,timeout=20,
            params={"seller":uid,"order.status":"paid",
                    "order.date_created.from":START.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "order.date_created.to":NOW.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "limit":50,"offset":offset}).json()
        res=r.get("results",[])
        if not res: break
        orders.extend(res); offset+=len(res)
        if offset>=r.get("paging",{}).get("total",0): break
    obs={}
    for o in orders:
        sid=(o.get("shipping") or {}).get("id")
        if sid: obs[sid]=o
    matches=0
    for sid, ord_o in obs.items():
        try:
            sh=requests.get(f"https://api.mercadolibre.com/shipments/{sid}",headers=H,timeout=10).json()
            st=sh.get("status"); sub=sh.get("substatus")
            if st not in ("ready_to_ship","handling"): continue
            if sub in SUB_EXCLUDE: continue
            deadline=None
            try:
                sla=requests.get(f"https://api.mercadolibre.com/shipments/{sid}/sla",headers=H,timeout=8).json()
                ed=sla.get("expected_date")
                if ed: deadline=datetime.fromisoformat(ed.replace("Z","+00:00")).astimezone(TZ)
            except: pass
            if not deadline:
                hist=sh.get("status_history") or {}
                dh=hist.get("date_handling")
                if dh: deadline=(datetime.fromisoformat(dh.replace("Z","+00:00"))+timedelta(hours=48)).astimezone(TZ)
            # Política operativa:
            # - Wilbert: SIEMPRE incluir, scope=HOY (acumulado por pausa, despacho urgente)
            # - Otras cuentas: SOLO deadline > hoy (las de hoy ya se imprimieron ayer)
            if acc == "Wilbert":
                scope = "HOY"
            else:
                if not deadline or deadline <= END_TODAY:
                    continue
                scope = "FUTURO"
            items=ord_o.get("order_items",[])
            comp_lines=[f"{clean_title((it.get('item') or {}).get('title',''))} x{it.get('quantity',1)}" for it in items]
            buyer=(ord_o.get("buyer") or {}).get("nickname","?")
            dl_str = deadline.strftime("%a %d %H:%M") if deadline else "s/d"
            shipments.append({"sid":sid,"account":acc,"token":at,
                              "comp_lines":comp_lines,"buyer":buyer,
                              "deadline":dl_str,"scope":scope,
                              "substatus":sub or "",
                              "tracking":sh.get("tracking_number","")})
            matches+=1
            time.sleep(0.04)
        except Exception as e:
            print(f"  err {sid}: {str(e)[:60]}")
    print(f"  matches: {matches}")

print(f"\nTotal pendientes: {len(shipments)}")
hoy_n = sum(1 for s in shipments if s["scope"]=="HOY")
fut_n = len(shipments) - hoy_n
print(f"  HOY: {hoy_n}  FUTURO: {fut_n}")
# Orden: HOY primero, luego FUTURO; dentro de cada grupo por producto+cuenta
shipments.sort(key=lambda s: (0 if s["scope"]=="HOY" else 1,
                              "/".join(s["comp_lines"]), s["account"]))
if LIMIT > 0:
    print(f"   PREVIEW MODE: detener tras {LIMIT} páginas OK")


# === FASE 2: 2 PDFs separados (HOY y FUTURO) ===
print("\n=== Generando PDFs 4x6 (HOY + FUTURO) ===")
writer_hoy=PdfWriter()
writer_fut=PdfWriter()
fail=[]
ok_pages=0

def render_header(s, header_h):
    """Devuelve un PdfPage 4x6 con la franja amarilla en el header. Texto centrado."""
    buf=io.BytesIO()
    c=canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    c.setFillColor(Color(1, 0.96, 0.74))
    c.rect(0, PAGE_H-header_h, PAGE_W, header_h, fill=1, stroke=0)
    c.setFillColorRGB(0,0,0)
    cx = PAGE_W / 2.0
    # Linea 1 pequeña - centrada
    c.setFont("Helvetica-Bold", 7.5)
    line1=f"[{s['account'].upper()}] {s['buyer'][:30]} | Ship:{s['sid']}"
    c.drawCentredString(cx, PAGE_H-11, line1)
    # Lineas grandes con productos - centradas y centradas verticalmente en el bloque restante
    big_lines = s["comp_lines"][:4]
    n = len(big_lines)
    line_h = 16
    block_top = PAGE_H - 18                # espacio reservado para line1
    block_bot = PAGE_H - header_h + 4
    block_h = block_top - block_bot
    text_h = n * line_h
    # baseline de la primera línea
    first_y = block_top - (block_h - text_h)/2.0 - 12
    c.setFont("Helvetica-Bold", 14)
    y = first_y
    for line in big_lines:
        c.drawCentredString(cx, y, line[:30])
        y -= line_h
    c.showPage(); c.save()
    buf.seek(0)
    return PdfReader(buf).pages[0]

stop_render=False
for s in shipments:
    if stop_render: break
    H={"Authorization":f"Bearer {s['token']}"}
    try:
        r=requests.get("https://api.mercadolibre.com/shipment_labels",
                      headers=H, params={"shipment_ids":s["sid"],"response_type":"pdf"},
                      timeout=30)
        if r.status_code!=200 or not r.headers.get("content-type","").lower().startswith("application/pdf"):
            fail.append({"sid":s["sid"],"account":s["account"],"sub":s["substatus"],
                         "scope":s["scope"],"reason":r.text[:300]})
            print(f"    err {s['account']}/{s['sid']} sub={s['substatus']}: HTTP {r.status_code} {r.text[:200]}")
            continue
        raw_pdf_bytes = r.content
        lbl_pdf=PdfReader(io.BytesIO(raw_pdf_bytes))
        for pidx, label_page in enumerate(lbl_pdf.pages):
            box = label_page.cropbox if label_page.cropbox else label_page.mediabox
            lbl_x0=float(box.left); lbl_y0=float(box.bottom)
            lbl_w=float(box.width); lbl_h=float(box.height)
            # Auto-detectar contenido para eliminar espacio en blanco intrínseco del PDF
            cb = detect_content_bbox(raw_pdf_bytes, pidx)
            if cb:
                cx0, cy0, cw, ch = cb
                # ajustar por offset original del cropbox
                lbl_x0 = float(box.left) + float(cx0)
                lbl_y0 = float(box.bottom) + float(cy0)
                lbl_w = float(cw)
                lbl_h = float(ch)
                print(f"   bbox {s['sid']}: ({lbl_x0:.0f},{lbl_y0:.0f}) {lbl_w:.0f}x{lbl_h:.0f} pt")
            n_lines=min(len(s["comp_lines"]), 4)
            header_h = 22 + n_lines*16
            label_area_h = PAGE_H - header_h

            new_page=PageObject.create_blank_page(width=PAGE_W, height=PAGE_H)
            sx = PAGE_W / lbl_w
            sy = label_area_h / lbl_h
            op = (Transformation()
                  .translate(-lbl_x0, -lbl_y0)
                  .scale(sx, sy))
            new_page.merge_transformed_page(label_page, op)

            hdr_page=render_header(s, header_h)
            new_page.merge_page(hdr_page)

            target = writer_hoy if s["scope"]=="HOY" else writer_fut
            target.add_page(new_page)
            ok_pages+=1
            if LIMIT>0 and ok_pages>=LIMIT:
                stop_render=True
                break
    except Exception as e:
        fail.append({"sid":s["sid"],"account":s["account"],"sub":s.get("substatus",""),
                     "scope":s.get("scope",""),"reason":f"{type(e).__name__}: {str(e)[:200]}"})
        print(f"    err {s['account']}/{s['sid']}: {type(e).__name__}: {str(e)[:140]}")
    time.sleep(0.08)

pdf_hoy="ETIQUETAS_HOY_4x6.pdf"
pdf_fut="ETIQUETAS_FUTURO_4x6.pdf"
with open(pdf_hoy,"wb") as f: writer_hoy.write(f)
with open(pdf_fut,"wb") as f: writer_fut.write(f)
print(f"\n✅ PDF HOY:    {pdf_hoy} ({len(writer_hoy.pages)} págs)")
print(f"✅ PDF FUTURO: {pdf_fut} ({len(writer_fut.pages)} págs)")
print(f"   Fallidas: {len(fail)}")
# Resumen de fallas por motivo
if fail:
    from collections import Counter
    reasons=Counter()
    for f_ in fail:
        # extrae mensaje del JSON si existe
        import re as _re
        m=_re.search(r'"message":"([^"]+)', f_.get("reason",""))
        key = m.group(1)[:80] if m else f_.get("reason","?")[:80]
        reasons[(f_.get("sub",""),key)] += 1
    print("\n=== Fallidas por motivo ===")
    for (sub,msg),n in reasons.most_common(15):
        print(f"  {n:4} sub={sub!r}  {msg}")


# === FASE 3: XLSX organizado por producto ===
wb=Workbook()
ws=wb.active
ws.title="Pendientes hoy + futuro"
hf=PatternFill("solid", fgColor="2C3E50")
hF=Font(bold=True, color="FFFFFF", size=11)
center=Alignment(horizontal="center", vertical="center", wrap_text=True)
border=Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
headers=["#","Scope","Cuenta","Shipment ID","Comprador","PRODUCTO","Tracking","Deadline"]
for col,h in enumerate(headers,1):
    cell=ws.cell(row=1,column=col,value=h)
    cell.fill=hf; cell.font=hF; cell.alignment=center; cell.border=border

acc_colors={"Juan":"D5E8D4","Raymundo":"DAE8FC","Wilbert":"FAD7A0","Claribel":"FFE6CC",
            "Asva":"E1D5E7","Mildred":"FFF2CC","Dilcie":"F5CBA7","Bren":"D4E6F1","Yc_New":"D5DBDB"}
for i,s in enumerate(shipments,1):
    fill=PatternFill("solid", fgColor=acc_colors.get(s["account"],"FFFFFF"))
    prod_text="\n".join(s["comp_lines"])
    row=[i, s["scope"], s["account"], s["sid"], s["buyer"], prod_text, s["tracking"], s["deadline"]]
    for col,val in enumerate(row,1):
        c=ws.cell(row=i+1, column=col, value=val)
        c.fill=fill; c.border=border
        c.alignment=Alignment(vertical="center", wrap_text=True)
        if col==6: c.font=Font(bold=True, size=10)
        if col==2 and val=="HOY":
            c.fill=PatternFill("solid", fgColor="FFD966"); c.font=Font(bold=True)

widths={1:5, 2:8, 3:11, 4:14, 5:24, 6:55, 7:24, 8:14}
for col,w in widths.items(): ws.column_dimensions[chr(64+col)].width=w
for r in range(2, len(shipments)+2): ws.row_dimensions[r].height=30
ws.freeze_panes="A2"

# Hoja 2: Resumen por producto
ws2=wb.create_sheet("Resumen por producto")
for col,h in enumerate(["PRODUCTO","CANTIDAD","CUENTAS"],1):
    cell=ws2.cell(row=1,column=col,value=h)
    cell.fill=hf; cell.font=hF; cell.alignment=center
prod_count=defaultdict(lambda: {"qty":0, "accounts":set()})
for s in shipments:
    key="/".join(s["comp_lines"])
    prod_count[key]["qty"]+=1
    prod_count[key]["accounts"].add(s["account"])
for r,(prod,info) in enumerate(sorted(prod_count.items(),key=lambda x:-x[1]["qty"]),2):
    ws2.cell(row=r,column=1,value=prod).font=Font(bold=True)
    ws2.cell(row=r,column=2,value=info["qty"]).alignment=Alignment(horizontal="center")
    ws2.cell(row=r,column=3,value=", ".join(sorted(info["accounts"])))
ws2.column_dimensions["A"].width=60
ws2.column_dimensions["B"].width=12
ws2.column_dimensions["C"].width=25

xlsx_out="DESPACHO_TODAS_PENDIENTES.xlsx"
wb.save(xlsx_out)
print(f"✅ XLSX: {xlsx_out}")

print(f"\n=== RESUMEN TODAS PENDIENTES (4x6) ===")
print(f"Total shipments: {len(shipments)}")
acc_count=defaultdict(int)
for s in shipments: acc_count[s["account"]]+=1
for acc,n in sorted(acc_count.items(),key=lambda x:-x[1]):
    print(f"  {acc}: {n}")
print(f"\nProductos top:")
for prod,info in sorted(prod_count.items(),key=lambda x:-x[1]["qty"])[:10]:
    print(f"  {info['qty']:3}u  {prod}")
