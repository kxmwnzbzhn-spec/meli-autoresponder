#!/usr/bin/env python3
"""Libro contable auditable para la cuenta MELI ROCIOANGEL."""
import json, os, sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

APP_ID = "2008666770714005"
APP_SECRET = os.environ["MELI_APP_SECRET"]
REFRESH_TOKEN = os.environ["MELI_REFRESH_TOKEN_ROCIOANGEL"]
START_DATE = os.environ.get("ACCOUNTING_START_DATE", "2026-01-01")
OUTPUT = os.environ.get("ACCOUNTING_OUTPUT", "CONTABILIDAD_ROCIOANGEL.xlsx")
DRIVE_ROOT = os.environ.get("DRIVE_FOLDER_ID", "1aIDN3iq6zwCacL57iamptvQCPoSDyRbL")
TZ = timezone(timedelta(hours=-6))


def api(method, url, headers=None, **kwargs):
    r = requests.request(method, url, headers=headers, timeout=30, **kwargs)
    r.raise_for_status()
    return r.json() if r.content else {}


tok = api("POST", "https://api.mercadolibre.com/oauth/token", data={
    "grant_type": "refresh_token", "client_id": APP_ID,
    "client_secret": APP_SECRET, "refresh_token": REFRESH_TOKEN,
})
H = {"Authorization": f"Bearer {tok['access_token']}"}
me = api("GET", "https://api.mercadolibre.com/users/me", headers=H)
seller_id = me["id"]

start = datetime.fromisoformat(START_DATE).replace(tzinfo=TZ)
end = (datetime.now(TZ) + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
date_from = start.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
date_to = end.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")

orders = []
offset = 0
while True:
    data = api("GET", "https://api.mercadolibre.com/orders/search", headers=H, params={
        "seller": seller_id, "order.date_created.from": date_from,
        "order.date_created.to": date_to, "sort": "date_asc", "limit": 50, "offset": offset,
    })
    batch = data.get("results", [])
    orders.extend(batch)
    offset += len(batch)
    if not batch or offset >= data.get("paging", {}).get("total", 0):
        break

rows = []
for summary in orders:
    oid = summary["id"]
    order = api("GET", f"https://api.mercadolibre.com/orders/{oid}", headers=H)
    items = order.get("order_items", [])
    gross_order = sum((i.get("unit_price") or 0) * (i.get("quantity") or 0) for i in items)
    payments = order.get("payments", [])
    fee_order = sum((p.get("marketplace_fee") or 0) for p in payments if p.get("status") in ("approved", "refunded", "charged_back"))
    refund_order = sum((p.get("transaction_amount_refunded") or 0) for p in payments)
    if not refund_order:
        refund_order = sum((p.get("transaction_amount") or 0) for p in payments if p.get("status") in ("refunded", "charged_back"))
    refund_date = ""
    for p in payments:
        if (p.get("transaction_amount_refunded") or 0) or p.get("status") in ("refunded", "charged_back"):
            refund_date = (p.get("date_last_modified") or "")[:10]
    ship_cost = 0
    sid = (order.get("shipping") or {}).get("id")
    if sid:
        try:
            shipment = api("GET", f"https://api.mercadolibre.com/shipments/{sid}", headers=H)
            option = shipment.get("shipping_option") or {}
            ship_cost = max(0, (option.get("list_cost") or 0) - (option.get("cost") or 0))
        except Exception as exc:
            print(f"WARN shipment {sid}: {exc}")
    reason = ""
    try:
        claims = api("GET", "https://api.mercadolibre.com/post-purchase/v1/claims/search", headers=H,
                     params={"resource_id": oid, "limit": 50}).get("data", [])
        reason = "; ".join(sorted({c.get("reason_id") or c.get("type") or "" for c in claims if c}))
    except Exception as exc:
        print(f"WARN claims {oid}: {exc}")
    created = datetime.fromisoformat(order["date_created"].replace("Z", "+00:00")).astimezone(TZ)
    for item in items:
        qty = item.get("quantity") or 0
        price = item.get("unit_price") or 0
        line_gross = qty * price
        share = line_gross / gross_order if gross_order else 0
        attrs = {a.get("id"): a.get("value_name") for a in (item.get("item", {}).get("variation_attributes") or [])}
        rows.append({
            "date": created.date(), "month": created.strftime("%Y-%m"), "order": str(oid),
            "shipment": str(sid or ""), "status": order.get("status", ""),
            "item_id": item.get("item", {}).get("id", ""), "variation_id": item.get("item", {}).get("variation_id", ""),
            "title": item.get("item", {}).get("title", ""), "color": attrs.get("COLOR", ""),
            "qty": qty, "unit_price": price, "gross": line_gross,
            "fee": round(fee_order * share, 2), "shipping": round(ship_cost * share, 2),
            "refund": round(refund_order * share, 2), "refund_date": refund_date,
            "refund_reason": reason, "allocation": "Exacto" if len(items) == 1 else "Prorrateado por venta bruta",
        })

wb = Workbook()
ws = wb.active
ws.title = "Ventas"
headers = ["Fecha venta", "Mes", "Order ID", "Shipment ID", "Estado", "Item ID", "Variation ID", "Producto", "Color",
           "Cantidad", "Precio unitario", "Venta bruta", "Comisión MELI", "Envío seller", "Devolución", "Fecha devolución",
           "Motivo devolución", "Asignación", "Ingreso neto", "Costo unitario", "Costo total", "Utilidad"]
ws.append(headers)
for d in rows:
    ws.append([d["date"], d["month"], d["order"], d["shipment"], d["status"], d["item_id"], d["variation_id"],
               d["title"], d["color"], d["qty"], d["unit_price"], d["gross"], d["fee"], d["shipping"], d["refund"],
               d["refund_date"], d["refund_reason"], d["allocation"]])
    r = ws.max_row
    ws.cell(r, 19, f"=L{r}-M{r}-N{r}-O{r}")
    ws.cell(r, 20, f'=IFERROR(VLOOKUP(F{r},Costos!A:D,4,FALSE),0)')
    ws.cell(r, 21, f"=J{r}*T{r}")
    ws.cell(r, 22, f"=S{r}-U{r}")

cost = wb.create_sheet("Costos")
cost.append(["Item ID", "Costo producto", "Gastos operativos", "Costo unitario total", "Notas"])
for iid, title in sorted({(d["item_id"], d["title"]) for d in rows}):
    normalized = (title or "").lower().replace("-", " ")
    if "go 5" in normalized or "go5" in normalized:
        product_cost, operating_cost = 260, 10
    elif "go 4" in normalized or "go4" in normalized:
        product_cost, operating_cost = 213, 10
    else:
        product_cost, operating_cost = 0, 0
    cost.append([iid, product_cost, operating_cost, product_cost + operating_cost, title])

daily = wb.create_sheet("Resumen diario")
daily.append(["Fecha", "Órdenes", "Unidades", "Venta bruta", "Comisiones", "Envíos", "Devoluciones", "Ingreso neto", "Utilidad"])
dates = sorted({d["date"] for d in rows})
for dt in dates:
    daily.append([dt])
    r = daily.max_row
    daily.cell(r, 2, f'=COUNTIFS(Ventas!A:A,A{r})')
    daily.cell(r, 3, f'=SUMIFS(Ventas!J:J,Ventas!A:A,A{r})')
    for col, source in zip(range(4, 10), "LMNOSV"):
        daily.cell(r, col, f'=SUMIFS(Ventas!{source}:{source},Ventas!A:A,A{r})')

monthly = wb.create_sheet("Cierre mensual")
monthly.append(["Mes", "Órdenes", "Unidades", "Venta bruta", "Comisiones", "Envíos", "Devoluciones", "Ingreso neto", "Utilidad", "% devolución"])
months = sorted({d["month"] for d in rows})
for month in months:
    monthly.append([month])
    r = monthly.max_row
    monthly.cell(r, 2, f'=COUNTIFS(Ventas!B:B,A{r})')
    monthly.cell(r, 3, f'=SUMIFS(Ventas!J:J,Ventas!B:B,A{r})')
    for col, source in zip(range(4, 10), "LMNOSV"):
        monthly.cell(r, col, f'=SUMIFS(Ventas!{source}:{source},Ventas!B:B,A{r})')
    monthly.cell(r, 10, f'=IFERROR(G{r}/D{r},0)')

returns = wb.create_sheet("Devoluciones")
returns.append(headers[:18])
for d in rows:
    if d["refund"] > 0:
        returns.append([d["date"], d["month"], d["order"], d["shipment"], d["status"], d["item_id"], d["variation_id"],
                        d["title"], d["color"], d["qty"], d["unit_price"], d["gross"], d["fee"], d["shipping"], d["refund"],
                        d["refund_date"], d["refund_reason"], d["allocation"]])

for sheet in wb.worksheets:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in range(1, sheet.max_column + 1):
        width = min(45, max(12, max(len(str(sheet.cell(r, col).value or "")) for r in range(1, min(sheet.max_row, 200) + 1)) + 2))
        sheet.column_dimensions[get_column_letter(col)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.column in (11,12,13,14,15,19,20,21,22) or sheet.title in ("Resumen diario", "Cierre mensual") and cell.column >= 4:
                cell.number_format = '"$"#,##0.00'
    if sheet.title == "Cierre mensual":
        for cell in sheet["J"][1:]: cell.number_format = "0.0%"

wb.save(OUTPUT)


def drive_service():
    from googleapiclient.discovery import build
    from google.oauth2.credentials import Credentials
    creds = Credentials(None, refresh_token=os.environ["GOOGLE_OAUTH_REFRESH_TOKEN"],
                        token_uri="https://oauth2.googleapis.com/token",
                        client_id=os.environ["GOOGLE_OAUTH_CLIENT_ID"],
                        client_secret=os.environ["GOOGLE_OAUTH_CLIENT_SECRET"],
                        scopes=["https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def folder(svc, name, parent):
    safe = name.replace("'", "\\'")
    q = f"name='{safe}' and mimeType='application/vnd.google-apps.folder' and '{parent}' in parents and trashed=false"
    found = svc.files().list(q=q, fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute().get("files", [])
    if found: return found[0]["id"]
    return svc.files().create(body={"name": name, "mimeType": "application/vnd.google-apps.folder", "parents": [parent]},
                              fields="id", supportsAllDrives=True).execute()["id"]


svc = drive_service()
accounting_id = folder(svc, "Contabilidad", DRIVE_ROOT)
account_id = folder(svc, "ROCIOANGEL", accounting_id)
from googleapiclient.http import MediaFileUpload
q = f"name='{OUTPUT}' and '{account_id}' in parents and trashed=false"
found = svc.files().list(q=q, fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute().get("files", [])
media = MediaFileUpload(OUTPUT, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resumable=True)
if found:
    fid = svc.files().update(fileId=found[0]["id"], media_body=media, fields="id", supportsAllDrives=True).execute()["id"]
else:
    fid = svc.files().create(body={"name": OUTPUT, "parents": [account_id]}, media_body=media,
                             fields="id", supportsAllDrives=True).execute()["id"]
print(f"ACCOUNT=ROCIOANGEL SELLER={seller_id} ROWS={len(rows)} FILE=https://drive.google.com/file/d/{fid}/view FOLDER=https://drive.google.com/drive/folders/{account_id}")
