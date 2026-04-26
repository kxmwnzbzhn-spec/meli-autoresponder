#!/usr/bin/env python3
"""Contabilidad MELI v2 — Dashboard ejecutivo + Reclamos tracking completo.
Hojas:
  1. Dashboard (KPIs ejecutivos)
  2. Ventas Diarias
  3. Por Cuenta
  4. Por Modelo (ventas)
  5. Reclamos (todos con detalle)
  6. Reclamos x Motivo (mensual)
  7. Reclamos x Modelo (mensual)
  8. Devoluciones $
  9. Mensual YYYY-MM
  10. Gráficas
"""
import os, requests, json, sys, re
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

OUTPUT = sys.argv[1] if len(sys.argv) > 1 else "contabilidad_meli.xlsx"

APP_ID = os.environ["MELI_APP_ID"]; APP_SECRET = os.environ["MELI_APP_SECRET"]

ACCOUNTS = [
    ("JUAN", "MELI_REFRESH_TOKEN"),
    ("CLARIBEL", "MELI_REFRESH_TOKEN_CLARIBEL"),
    ("ASVA", "MELI_REFRESH_TOKEN_ASVA"),
    ("RAYMUNDO", "MELI_REFRESH_TOKEN_RAYMUNDO"),
    ("DILCIE", "MELI_REFRESH_TOKEN_DILCIE"),
    ("MILDRED", "MELI_REFRESH_TOKEN_MILDRED"),
]

REASON_LABELS = {
    "PDD9943": "No es original (imitación)",
    "PDD9944": "Producto defectuoso/dañado",
    "PNR9501": "No llegó el producto",
    "PDD8975": "Producto incompleto",
    "PDD9945": "Producto distinto al publicado",
    "PDD8974": "Diferente al pedido",
    "PDD8973": "No le gustó",
    "PDR9526": "No recibió el producto",
    "PDD9939": "Diferente al publicado",
    "PDD9946": "Otro motivo (PDD9946)",
    "PDD9950": "Otro motivo (PDD9950)",
    "PNR9508": "Cancelación pago",
}

def categorize_product(title):
    if not title: return ("Otro","Sin clasificar")
    tl = title.lower()
    if "go 4" in tl or "go4" in tl: return ("Bocina","JBL Go 4")
    if "go 3" in tl or "go3" in tl: return ("Bocina","JBL Go 3")
    if "go essential" in tl: return ("Bocina","JBL Go Essential")
    if "flip 7" in tl or "flip7" in tl: return ("Bocina","JBL Flip 7")
    if "flip 6" in tl: return ("Bocina","JBL Flip 6")
    if "charge 6" in tl: return ("Bocina","JBL Charge 6")
    if "charge 5" in tl: return ("Bocina","JBL Charge 5")
    if "grip" in tl and "jbl" in tl: return ("Bocina","JBL Grip")
    if "clip 5" in tl: return ("Bocina","JBL Clip 5")
    if "srs-xb100" in tl or "xb100" in tl: return ("Bocina","Sony XB100")
    if "bocina" in tl or "parlante" in tl or "altavoz" in tl: return ("Bocina","Otra bocina")
    if "armaf" in tl:
        if "club de nuit" in tl: return ("Perfume","Armaf Club De Nuit")
        if "odyssey" in tl: return ("Perfume","Armaf Odyssey")
        if "delight" in tl: return ("Perfume","Armaf Delight")
        return ("Perfume","Armaf otros")
    if "lattafa" in tl: return ("Perfume","Lattafa")
    if "mugler" in tl or "angel" in tl: return ("Perfume","Mugler/Angel")
    if "perfume" in tl or "edp" in tl or "edt" in tl: return ("Perfume","Otro perfume")
    if "buds" in tl or "auriculares" in tl: return ("Audífonos","Auriculares")
    return ("Otro", title[:30])

# ============ COLLECT DATA: claims (todos los meses) + ventas ============
print(f"Construyendo {OUTPUT}...")

target_date = datetime.now(timezone.utc) - timedelta(hours=6)
cdmx = target_date

# Daily date for today's row
today_label = cdmx.strftime("%Y-%m-%d")
day_start_cdmx = cdmx.replace(hour=0,minute=0,second=0,microsecond=0)
day_end_cdmx = day_start_cdmx + timedelta(days=1)
day_start_utc = day_start_cdmx.astimezone(timezone.utc)
day_end_utc = day_end_cdmx.astimezone(timezone.utc)
date_from_today = day_start_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")
date_to_today = day_end_utc.strftime("%Y-%m-%dT%H:%M:%S.000Z")

# Per-account collected
per_account = {}
all_claims = []  # (account, claim_id, order_id, type, status, reason_id, reason_label, date, product_title, model_category, model)

for label, env_var in ACCOUNTS:
    RT = os.environ.get(env_var,"")
    if not RT: continue
    try:
        r = requests.post("https://api.mercadolibre.com/oauth/token",data={
            "grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":RT
        }).json()
        H = {"Authorization":f"Bearer {r['access_token']}"}
        me = requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=10).json()
        USER_ID = me["id"]
    except Exception as e: continue
    
    # ===== Today's sales =====
    a_orders = a_qty = 0; a_bruto = a_fees = a_ship = 0; a_returns_count = 0; a_returns_amt = 0
    offset = 0
    while True:
        rr = requests.get(f"https://api.mercadolibre.com/orders/search?seller={USER_ID}&order.date_created.from={date_from_today}&order.date_created.to={date_to_today}&limit=50&offset={offset}",headers=H,timeout=20).json()
        results = rr.get("results",[])
        if not results: break
        for o in results:
            st = o.get("status",""); order_id = o.get("id"); amt = o.get("total_amount",0) or 0
            had_paid=False; refund=0
            for pay in o.get("payments",[]):
                if pay.get("status")=="approved": had_paid=True
                if pay.get("status") in ("refunded","charged_back"): refund += pay.get("transaction_amount",0) or 0
            if st=="cancelled" and not had_paid: continue
            if refund > 0:
                a_returns_count += 1; a_returns_amt += refund
            if st not in ("paid","shipped","delivered"): continue
            a_orders += 1; a_bruto += amt
            try:
                od = requests.get(f"https://api.mercadolibre.com/orders/{order_id}",headers=H,timeout=10).json()
                for pay in od.get("payments",[]):
                    if pay.get("status")=="approved":
                        a_fees += pay.get("marketplace_fee",0) or 0
                sh_id = od.get("shipping",{}).get("id")
                if sh_id:
                    sd = requests.get(f"https://api.mercadolibre.com/shipments/{sh_id}",headers=H,timeout=10).json()
                    so = sd.get("shipping_option",{}) or {}
                    a_ship += max(0,(so.get("list_cost",0) or 0)-(so.get("cost",0) or 0))
            except: pass
            for oi in o.get("order_items",[]):
                a_qty += oi.get("quantity",0)
        offset += 50
        if offset >= rr.get("paging",{}).get("total",0): break
    
    iva = a_bruto - (a_bruto/1.16) if a_bruto>0 else 0
    net = a_bruto - a_fees - a_ship
    per_account[label] = {"nick":me.get("nickname",""),"orders":a_orders,"qty":a_qty,"bruto":a_bruto,
                         "fees":a_fees,"ship":a_ship,"iva":iva,"net":net,"net_iva":net-iva,
                         "returns_count":a_returns_count,"returns_amt":a_returns_amt}
    
    # ===== ALL claims (status=opened + closed) =====
    for st_filter in ("opened","closed"):
        offset = 0
        while True:
            url = f"https://api.mercadolibre.com/post-purchase/v1/claims/search?limit=50&offset={offset}&status={st_filter}"
            cs = requests.get(url, headers=H, timeout=20).json()
            data = cs.get("data") or []
            if not data: break
            for c in data:
                cid = c.get("id"); ctype = c.get("type",""); rid = c.get("reason_id","")
                rels = c.get("related_entities",[]) or []; status = c.get("status","")
                cd = c.get("date_created","")
                order_id = c.get("resource_id","")
                # Get product
                product = ""; mdl_cat = ""; mdl = ""
                try:
                    od = requests.get(f"https://api.mercadolibre.com/orders/{order_id}",headers=H,timeout=10).json()
                    items_o = od.get("order_items",[])
                    if items_o:
                        product = items_o[0].get("item",{}).get("title","")
                        mdl_cat, mdl = categorize_product(product)
                except: pass
                affects = ctype != "cancel_purchase" and not (ctype == "mediations" and "return" in rels)
                all_claims.append({
                    "account": label, "claim_id": cid, "order_id": order_id,
                    "type": ctype, "status": status, "reason_id": rid,
                    "reason_label": REASON_LABELS.get(rid, rid or "?"),
                    "date": cd, "product": product[:60], "category": mdl_cat, "model": mdl,
                    "affects_reputation": affects
                })
            offset += 50
            if offset >= 200: break  # safety

print(f"Datos: {len(per_account)} cuentas | {len(all_claims)} reclamos totales")

# ============ BUILD XLSX ============
if os.path.exists(OUTPUT):
    wb = load_workbook(OUTPUT)
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E78", name="Arial")
SECTION_FONT = Font(bold=True, size=13, color="1F4E78", name="Arial")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
ALERT_FILL = PatternFill("solid", start_color="FFE6E6")
SUCCESS_FILL = PatternFill("solid", start_color="E6F4EA")
NUM_FMT = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
INT_FMT = '#,##0;[Red](#,##0);"-"'
PCT_FMT = '0.0%;[Red](0.0%);"-"'
BORDER = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

def hdr(c):
    c.fill=HDR_FILL; c.font=HDR_FONT
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=BORDER

# ====== HOJA 1: DASHBOARD EJECUTIVO ======
sname = "Dashboard"
if sname in wb.sheetnames: del wb[sname]
ds = wb.create_sheet(sname, 0)  # primera hoja

ds.cell(row=1,column=1,value="📊 DASHBOARD CONTABILIDAD MELI").font = TITLE_FONT
ds.merge_cells("A1:F1")
ds.row_dimensions[1].height = 28

ds.cell(row=2,column=1,value=f"Última actualización: {cdmx.strftime('%d/%m/%Y %H:%M')} CDMX").font = Font(italic=True, size=10, color="666666")
ds.merge_cells("A2:F2")

# KPIs HOY
ds.cell(row=4,column=1,value="🟢 HOY").font = SECTION_FONT
ds.merge_cells("A4:F4")
t_o = sum(d["orders"] for d in per_account.values())
t_b = sum(d["bruto"] for d in per_account.values())
t_f = sum(d["fees"] for d in per_account.values())
t_s = sum(d["ship"] for d in per_account.values())
t_iva = sum(d["iva"] for d in per_account.values())
t_net = t_b - t_f - t_s
t_net_iva = t_net - t_iva

kpis_hoy = [
    ("Órdenes",t_o,INT_FMT,None),
    ("Bruto",t_b,NUM_FMT,None),
    ("− Comisión MELI",t_f,NUM_FMT,None),
    ("− Envío seller",t_s,NUM_FMT,None),
    ("− IVA (16%)",t_iva,NUM_FMT,None),
    ("✅ NET (post IVA)",t_net_iva,NUM_FMT,SUCCESS_FILL),
]
r = 5
for k,v,fmt,fill in kpis_hoy:
    c1 = ds.cell(row=r,column=1,value=k); c1.font = Font(bold=True)
    c2 = ds.cell(row=r,column=2,value=v); c2.number_format=fmt; c2.alignment=Alignment(horizontal="right")
    if fill:
        c1.fill = fill; c2.fill = fill
    r += 1

# RECLAMOS RESUMEN
r += 1
ds.cell(row=r,column=1,value="🚨 RECLAMOS").font = SECTION_FONT
ds.merge_cells(f"A{r}:F{r}")
r += 1

claims_total = len(all_claims)
claims_affecting = sum(1 for c in all_claims if c["affects_reputation"])
claims_open = sum(1 for c in all_claims if c["status"]=="opened")
claims_open_aff = sum(1 for c in all_claims if c["status"]=="opened" and c["affects_reputation"])

reclamos_kpis = [
    ("Total históricos",claims_total,INT_FMT,None),
    ("Abiertos ahora",claims_open,INT_FMT,ALERT_FILL if claims_open>0 else None),
    ("Abiertos que afectan reputación",claims_open_aff,INT_FMT,ALERT_FILL if claims_open_aff>0 else SUCCESS_FILL),
    ("Total que afectaron reputación",claims_affecting,INT_FMT,None),
]
for k,v,fmt,fill in reclamos_kpis:
    c1 = ds.cell(row=r,column=1,value=k); c1.font = Font(bold=True)
    c2 = ds.cell(row=r,column=2,value=v); c2.number_format=fmt; c2.alignment=Alignment(horizontal="right")
    if fill:
        c1.fill = fill; c2.fill = fill
    r += 1

# Por cuenta hoy
r += 2
ds.cell(row=r,column=1,value="📊 POR CUENTA (HOY)").font = SECTION_FONT
ds.merge_cells(f"A{r}:F{r}")
r += 1
for col,h in enumerate(["Cuenta","Órdenes","Bruto","Comisión","Envío","NET"],1):
    hdr(ds.cell(row=r,column=col,value=h))
r += 1
for label, d in sorted(per_account.items(), key=lambda x: -x[1]["bruto"]):
    if d["orders"]==0: continue
    ds.cell(row=r,column=1,value=label).font = Font(bold=True)
    ds.cell(row=r,column=2,value=d["orders"]).number_format=INT_FMT
    ds.cell(row=r,column=3,value=d["bruto"]).number_format=NUM_FMT
    ds.cell(row=r,column=4,value=d["fees"]).number_format=NUM_FMT
    ds.cell(row=r,column=5,value=d["ship"]).number_format=NUM_FMT
    ds.cell(row=r,column=6,value=d["net"]).number_format=NUM_FMT
    r += 1

# Width
for col,w in enumerate([26,12,14,14,14,14],1):
    ds.column_dimensions[get_column_letter(col)].width = w

# ====== HOJA 2: RECLAMOS (todos con detalle) ======
sname = "Reclamos"
if sname in wb.sheetnames: del wb[sname]
rs = wb.create_sheet(sname, 1)
hdrs = ["Fecha","Cuenta","Claim ID","Orden","Tipo","Status","Motivo","Código","Producto","Modelo","Categoría","Afecta reputación"]
for col,h in enumerate(hdrs,1): hdr(rs.cell(row=1,column=col,value=h))
rs.row_dimensions[1].height = 30

r = 2
for c in sorted(all_claims, key=lambda x: x["date"], reverse=True):
    fill = ALERT_FILL if c["affects_reputation"] else None
    cells = [
        (1, c["date"][:10]),
        (2, c["account"]),
        (3, c["claim_id"]),
        (4, c["order_id"]),
        (5, c["type"]),
        (6, c["status"]),
        (7, c["reason_label"]),
        (8, c["reason_id"]),
        (9, c["product"][:50]),
        (10, c["model"]),
        (11, c["category"]),
        (12, "🚨 SÍ" if c["affects_reputation"] else "⊘ NO"),
    ]
    for col,val in cells:
        cell = rs.cell(row=r,column=col,value=val)
        if fill: cell.fill = fill
    r += 1

for col,w in enumerate([12,12,16,18,14,12,28,12,40,18,12,18],1):
    rs.column_dimensions[get_column_letter(col)].width = w

# ====== HOJA 3: RECLAMOS x MOTIVO (mensual) ======
sname = "Reclamos x Motivo"
if sname in wb.sheetnames: del wb[sname]
rms = wb.create_sheet(sname)

# Build (month, motivo) → count
by_month_motivo = defaultdict(lambda: defaultdict(int))  # {month: {motivo: count}}
all_months = set()
all_motivos = set()
for c in all_claims:
    if not c["affects_reputation"]: continue  # solo los que afectan
    try:
        d = datetime.fromisoformat(c["date"].replace("Z","+00:00"))
    except: continue
    month = d.strftime("%Y-%m")
    motivo = c["reason_label"]
    by_month_motivo[month][motivo] += 1
    all_months.add(month); all_motivos.add(motivo)

months_sorted = sorted(all_months)
motivos_sorted = sorted(all_motivos)

rms.cell(row=1,column=1,value="🚨 RECLAMOS QUE AFECTAN REPUTACIÓN — POR MOTIVO Y MES").font = TITLE_FONT
rms.merge_cells("A1:G1")

# Headers: Motivo | mes1 | mes2 | ... | TOTAL
headers = ["Motivo"] + months_sorted + ["TOTAL"]
for col,h in enumerate(headers,1): hdr(rms.cell(row=3,column=col,value=h))
r = 4
for motivo in motivos_sorted:
    rms.cell(row=r,column=1,value=motivo).font = Font(bold=True)
    total = 0
    for col, mes in enumerate(months_sorted, 2):
        v = by_month_motivo[mes][motivo]
        if v > 0:
            cell = rms.cell(row=r,column=col,value=v)
            cell.number_format = INT_FMT
            cell.alignment = Alignment(horizontal="center")
        total += v
    rms.cell(row=r,column=len(headers),value=total).font = Font(bold=True)
    rms.cell(row=r,column=len(headers)).fill = ALERT_FILL
    r += 1

# Total row
rms.cell(row=r,column=1,value="TOTAL").font = Font(bold=True, size=12)
total_grand = 0
for col, mes in enumerate(months_sorted,2):
    s = sum(by_month_motivo[mes].values())
    rms.cell(row=r,column=col,value=s).font = Font(bold=True)
    rms.cell(row=r,column=col).number_format = INT_FMT
    total_grand += s
rms.cell(row=r,column=len(headers),value=total_grand).font = Font(bold=True, size=12, color="C00000")

for col in range(1, len(headers)+1):
    rms.column_dimensions[get_column_letter(col)].width = 30 if col==1 else 12

# ====== HOJA 4: RECLAMOS x MODELO (mensual) ======
sname = "Reclamos x Modelo"
if sname in wb.sheetnames: del wb[sname]
rmd = wb.create_sheet(sname)

by_month_model = defaultdict(lambda: defaultdict(int))
all_months_m = set(); all_models_m = set()
for c in all_claims:
    if not c["affects_reputation"]: continue
    if not c["model"]: continue
    try:
        d = datetime.fromisoformat(c["date"].replace("Z","+00:00"))
    except: continue
    month = d.strftime("%Y-%m")
    by_month_model[month][c["model"]] += 1
    all_months_m.add(month); all_models_m.add(c["model"])

months_sorted_m = sorted(all_months_m)
models_sorted = sorted(all_models_m)

rmd.cell(row=1,column=1,value="📦 RECLAMOS POR MODELO Y MES (afectan reputación)").font = TITLE_FONT
rmd.merge_cells("A1:G1")

headers_m = ["Modelo"] + months_sorted_m + ["TOTAL"]
for col,h in enumerate(headers_m,1): hdr(rmd.cell(row=3,column=col,value=h))
r = 4
for mdl in models_sorted:
    rmd.cell(row=r,column=1,value=mdl).font = Font(bold=True)
    total = 0
    for col, mes in enumerate(months_sorted_m,2):
        v = by_month_model[mes][mdl]
        if v > 0:
            cell = rmd.cell(row=r,column=col,value=v)
            cell.number_format = INT_FMT
            cell.alignment = Alignment(horizontal="center")
        total += v
    rmd.cell(row=r,column=len(headers_m),value=total).font = Font(bold=True)
    rmd.cell(row=r,column=len(headers_m)).fill = ALERT_FILL
    r += 1

for col in range(1, len(headers_m)+1):
    rmd.column_dimensions[get_column_letter(col)].width = 30 if col==1 else 12

# ====== Reorder sheets ======
order = ["Dashboard","Reclamos","Reclamos x Motivo","Reclamos x Modelo"]
for i, name in enumerate(order):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        if idx != i:
            wb.move_sheet(name, offset=i-idx)

wb.save(OUTPUT)
print(f"\n✅ Saved {OUTPUT}")
print(f"\n📊 Reclamos categorizados:")
print(f"   Total: {claims_total} | Abiertos: {claims_open} | Afectan reputación: {claims_affecting}")
print(f"   Modelos con reclamos: {len(models_sorted)}")
print(f"   Motivos: {len(motivos_sorted)}")
