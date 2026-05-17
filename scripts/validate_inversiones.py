"""Validador semanal: confirma que docs/INVERSIONES_enriched.xlsx existe en repo y tiene los emails esperados."""
import os,requests,base64,sys
from openpyxl import load_workbook
import io

GHT=os.environ.get("GH_TOKEN","")
REPO="kxmwnzbzhn-spec/meli-autoresponder"
TG=os.environ.get("TELEGRAM_BOT_TOKEN")
TC=os.environ.get("TELEGRAM_CHAT_ID")

def tg(msg):
    if TG and TC:
        requests.post(f"https://api.telegram.org/bot{TG}/sendMessage",data={"chat_id":TC,"text":msg,"parse_mode":"Markdown"},timeout=10)

r=requests.get(f"https://api.github.com/repos/{REPO}/contents/docs/INVERSIONES_enriched.xlsx",
    headers={"Authorization":f"Bearer {GHT}","Accept":"application/vnd.github+json"})
if r.status_code!=200:
    msg=f"⚠️ INVERSIONES_enriched.xlsx no encontrado en repo (http={r.status_code})"
    print(msg); tg(msg); sys.exit(1)

# Download raw content (it's > 1MB needs raw download via download_url)
meta=r.json()
dl=requests.get(meta["download_url"]).content
print(f"file size: {len(dl)} bytes")

# Parse and validate
wb=load_workbook(io.BytesIO(dl),data_only=True)
ws=wb["PORTAFOLIO"]
expected={
  "Wilbert":"ventas.meli017@gmail.com",
  "Juan":"ventas.meli001@gmail.com",
  "Mildred":"ventas.meli015@gmail.com",
  "Dilcie":"ventas.meli014@gmail.com",
  "Bren":"ventas.meli013@gmail.com",
  "YC_NEW":"ventas.meli016@gmail.com",
  "Claribel":"ventas.meli011@gmail.com",
  "Raymundo":"enviamesantacruz1@gmail.com",
  "Asva":"asvaelectronics@gmail.com",
}
issues=[]
found={}
for r_idx in range(11,30):
    cuenta=ws.cell(row=r_idx,column=2).value
    email=ws.cell(row=r_idx,column=4).value
    if not cuenta: continue
    found[str(cuenta).strip()]=email
for k,exp in expected.items():
    actual=None
    for fc in found:
        if k.lower()==str(fc).lower() or k.lower() in str(fc).lower():
            actual=found[fc]; break
    if actual!=exp:
        issues.append(f"  {k}: expected `{exp}` got `{actual}`")
if issues:
    msg="⚠️ *INVERSIONES_enriched validation FAILED*:\n"+"\n".join(issues)
    print(msg); tg(msg); sys.exit(2)
print(f"✓ Validación OK ({len(expected)} cuentas verificadas)")
tg(f"✓ INVERSIONES_enriched OK ({len(expected)} emails verificados)")
