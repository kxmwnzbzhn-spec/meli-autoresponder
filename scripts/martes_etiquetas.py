"""ETIQUETAS MARTES 5/MAY: shipments con deadline > lunes 23:59 y <= martes 23:59.
NO incluye los de hoy lunes (ya impresos). Mismo formato V4: header integrado en la pagina.
"""
import os, io, time, requests
from datetime import datetime, timedelta, timezone
from collections import defaultdict
from pypdf import PdfReader, PdfWriter, PageObject
from reportlab.pdfgen import canvas
from reportlab.lib.colors import Color
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP_ID="5211907102822632"
APP_SECRET=os.environ["MELI_APP_SECRET"]
ACCS={
    "Juan":     os.environ.get("MELI_REFRESH_TOKEN_JUAN") or os.environ.get("MELI_REFRESH_TOKEN"),
    "Raymundo": os.environ.get("MELI_REFRESH_TOKEN_RAYMUNDO"),
    "Claribel": os.environ.get("MELI_REFRESH_TOKEN_CLARIBEL"),
    "Asva":     os.environ.get("MELI_REFRESH_TOKEN_ASVA"),
    "Mildred":  os.environ.get("MELI_REFRESH_TOKEN_MILDRED"),
    "Dilcie":   os.environ.get("MELI_REFRESH_TOKEN_DILCIE"),
    "Bren":     os.environ.get("MELI_REFRESH_TOKEN_BREN"),
}
TZ=timezone(timedelta(hours=-6))
# RANGO MARTES: > lunes 23:59 (osea martes 00:00) hasta martes 23:59
START_MARTES = datetime.fromisoformat("2026-05-05").replace(hour=0,  minute=0,  tzinfo=TZ)
END_MARTES   = datetime.fromisoformat("2026-05-05").replace(hour=23, minute=59, tzinfo=TZ)


def tok(rt):
    r=requests.post("https://api.mercadolibre.com/oauth/token",data={
        "grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":rt}).json()
    return r.get("access_token")


def clean_title(title):
    t=(title or "").strip()
    for w in ["Bocina ","bocina ","Parlante ","parlante ","Altavoz ","altavoz ","Speaker ","speaker ",
              "JBL ","jbl ","Jbl ","Sony ","SONY ","Bose ","BOSE "]:
        t=t.replace(w,"")
    tl=t.lower()
    if "go 4" in tl or "go4" in tl: model="Go 4"
    elif "go 3" in tl or "go3" in tl: model="Go 3"
    elif "clip 5" in tl or "clip5" in tl: model="Clip 5"
    elif "charge 6" in tl or "charge6" in tl: model="Charge 6"
    elif "flip 7" in tl or "flip7" in tl: model="Flip 7"
    elif "grip" in tl: model="Grip"
    elif "xb100" in tl: model="Sony XB100"
    elif "soundlink" in tl: model="Bose SoundLink"
    else: model=t[:30]
    col_map=[("camuflaj","Camuflaje"),("camo","Camuflaje"),("azul marino","Azul Marino"),
             ("aqua","Aqua"),("celeste","Celeste"),("negr","Negro"),("black","Negro"),
             ("roj","Rojo"),(" red","Rojo"),("rosa","Rosa"),("pink","Rosa"),
             ("morad","Morado"),("violeta","Morado"),("purple","Morado"),
             (" azul","Azul"),(" blue","Azul")]
    color=None
    for k,v in col_map:
        if k in (" "+tl):
            color=v; break
    return f"{model} {color}" if color else model


# === FASE 1: shipments MARTES ===
NOW=datetime.now(timezone.utc); START=NOW-timedelta(days=10)
shipments=[]
for acc, rt in ACCS.items():
    if not rt: continue
    print(f"=== {acc} ===")
    at=tok(rt)
    if not at: continue
    H={"Authorization":f"Bearer {at}"}
    me=requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=15).json()
    uid=me.get("id")
    if not uid: continue
    orders=[]; offset=0
    while True:
        r=requests.get("https://api.mercadolibre.com/orders/search",headers=H,timeout=20,
            params={"seller":uid,"order.status":"paid",
                    "order.date_created.from":START.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "order.date_created.to":NOW.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "limit":50,"offset":offset}).json()
        res=r.get("results",[])
        if not res: break
        orders.extend(res); offset+=len(res)
        if offset>=r.get("paging",{}).get("total",0): break
    obs={}
    for o in orders:
        sid=(o.get("shipping") or {}).get("id")
        if sid: obs[sid]=o
    matches=0
    for sid, ord_o in obs.items():
        try:
            sh=requests.get(f"https://api.mercadolibre.com/shipments/{sid}",headers=H,timeout=10).json()
            st=sh.get("status"); sub=sh.get("substatus")
            if st not in ("ready_to_ship","handling"): continue
            if sub in ("shipped","ready_to_pickup"): continue
            deadline=None
            try:
                sla=requests.get(f"https://api.mercadolibre.com/shipments/{sid}/sla",headers=H,timeout=8).json()
                ed=sla.get("expected_date")
                if ed: deadline=datetime.fromisoformat(ed.replace("Z","+00:00")).astimezone(TZ)
            except: pass
            if not deadline:
                hist=sh.get("status_history") or {}
                dh=hist.get("date_handling")
                if dh: deadline=(datetime.fromisoformat(dh.replace("Z","+00:00"))+timedelta(hours=48)).astimezone(TZ)
            if not deadline: continue
            # SOLO los del MARTES: deadline >= Mar 00:00 y <= Mar 23:59
            if deadline < START_MARTES or deadline > END_MARTES: continue
            items=ord_o.get("order_items",[])
            comp_lines=[f"{clean_title((it.get('item') or {}).get('title',''))} x{it.get('quantity',1)}" for it in items]
            buyer=(ord_o.get("buyer") or {}).get("nickname","?")
            shipments.append({"sid":sid,"account":acc,"token":at,
                              "comp_lines":comp_lines,"buyer":buyer,
                              "deadline":deadline.strftime("%a %H:%M"),
                              "tracking":sh.get("tracking_number","")})
            matches+=1
            time.sleep(0.04)
        except Exception as e:
            print(f"  err {sid}: {str(e)[:60]}")
    print(f"  matches: {matches}")

print(f"\nTotal MARTES: {len(shipments)}")
# Sort por composición para picking eficiente
shipments.sort(key=lambda s: ("/".join(s["comp_lines"]), s["account"]))


# === FASE 2: PDF con header integrado ===
print("\n=== Generando PDF MARTES ===")
by_acc=defaultdict(list)
for s in shipments:
    by_acc[s["account"]].append(s)

writer=PdfWriter()
fail=[]
for acc, ships in by_acc.items():
    if not ships: continue
    token=ships[0]["token"]
    H={"Authorization":f"Bearer {token}"}
    print(f"  {acc}: {len(ships)}")
    for s in ships:
        try:
            r=requests.get("https://api.mercadolibre.com/shipment_labels",
                          headers=H, params={"shipment_ids":s["sid"],"response_type":"pdf"},
                          timeout=30)
            if r.status_code!=200 or not r.headers.get("content-type","").lower().startswith("application/pdf"):
                fail.append(s["sid"]); continue
            lbl_pdf=PdfReader(io.BytesIO(r.content))
            for label_page in lbl_pdf.pages:
                lbl_w=float(label_page.mediabox.width); lbl_h=float(label_page.mediabox.height)
                n_lines=len(s["comp_lines"])
                hdr_h=60+n_lines*20
                new_h=lbl_h+hdr_h
                new_page=PageObject.create_blank_page(width=lbl_w, height=new_h)
                buf=io.BytesIO()
                c=canvas.Canvas(buf, pagesize=(lbl_w, new_h))
                c.setFillColor(Color(1, 0.96, 0.74))
                c.rect(0, lbl_h, lbl_w, hdr_h, fill=1, stroke=0)
                c.setFillColorRGB(0,0,0)
                c.setFont("Helvetica-Bold", 11)
                c.drawString(8, lbl_h+hdr_h-18, f"[{s['account'].upper()}] {s['buyer'][:25]} | Ship:{s['sid']}")
                c.setFont("Helvetica-Bold", 18)
                y=lbl_h+hdr_h-42
                for line in s["comp_lines"][:5]:
                    c.drawString(8, y, line); y-=20
                c.showPage(); c.save()
                buf.seek(0)
                hdr_page=PdfReader(buf).pages[0]
                new_page.merge_page(hdr_page)
                new_page.merge_page(label_page)
                writer.add_page(new_page)
        except Exception as e:
            fail.append(s["sid"])
            print(f"    err {s['sid']}: {str(e)[:80]}")
        time.sleep(0.08)

pdf_out="ETIQUETAS_MARTES_5_MAYO.pdf"
with open(pdf_out,"wb") as f: writer.write(f)
print(f"\n✅ PDF: {pdf_out} ({len(writer.pages)} páginas) | Fallidas: {len(fail)}")


# === FASE 3: XLSX organizado por producto ===
wb=Workbook()
ws=wb.active
ws.title="Despacho Martes 5 Mayo"
hf=PatternFill("solid", fgColor="2C3E50")
hF=Font(bold=True, color="FFFFFF", size=11)
center=Alignment(horizontal="center", vertical="center", wrap_text=True)
border=Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
headers=["#","Cuenta","Shipment ID","Comprador","PRODUCTO","Tracking","Deadline"]
for col,h in enumerate(headers,1):
    cell=ws.cell(row=1,column=col,value=h)
    cell.fill=hf; cell.font=hF; cell.alignment=center; cell.border=border

acc_colors={"Juan":"D5E8D4","Raymundo":"DAE8FC","Claribel":"FFE6CC","Asva":"E1D5E7"}
for i,s in enumerate(shipments,1):
    fill=PatternFill("solid", fgColor=acc_colors.get(s["account"],"FFFFFF"))
    prod_text="\n".join(s["comp_lines"])
    row=[i, s["account"], s["sid"], s["buyer"], prod_text, s["tracking"], s["deadline"]]
    for col,val in enumerate(row,1):
        c=ws.cell(row=i+1, column=col, value=val)
        c.fill=fill; c.border=border
        c.alignment=Alignment(vertical="center", wrap_text=True)
        if col==5: c.font=Font(bold=True, size=10)

widths={1:5, 2:11, 3:14, 4:24, 5:55, 6:24, 7:13}
for col,w in widths.items(): ws.column_dimensions[chr(64+col)].width=w
for r in range(2, len(shipments)+2): ws.row_dimensions[r].height=30
ws.freeze_panes="A2"

# Hoja 2: Resumen por producto
ws2=wb.create_sheet("Resumen por producto")
for col,h in enumerate(["PRODUCTO","CANTIDAD","CUENTAS"],1):
    cell=ws2.cell(row=1,column=col,value=h)
    cell.fill=hf; cell.font=hF; cell.alignment=center
prod_count=defaultdict(lambda: {"qty":0, "accounts":set()})
for s in shipments:
    key="/".join(s["comp_lines"])
    prod_count[key]["qty"]+=1
    prod_count[key]["accounts"].add(s["account"])
for r,(prod,info) in enumerate(sorted(prod_count.items(),key=lambda x:-x[1]["qty"]),2):
    ws2.cell(row=r,column=1,value=prod).font=Font(bold=True)
    ws2.cell(row=r,column=2,value=info["qty"]).alignment=Alignment(horizontal="center")
    ws2.cell(row=r,column=3,value=", ".join(sorted(info["accounts"])))
ws2.column_dimensions["A"].width=60
ws2.column_dimensions["B"].width=12
ws2.column_dimensions["C"].width=25

xlsx_out="DESPACHO_MARTES_5_MAYO.xlsx"
wb.save(xlsx_out)
print(f"✅ XLSX: {xlsx_out}")

# Resumen para terminal
print(f"\n=== RESUMEN MARTES ===")
print(f"Total shipments: {len(shipments)}")
acc_count=defaultdict(int)
for s in shipments: acc_count[s["account"]]+=1
for acc,n in sorted(acc_count.items(),key=lambda x:-x[1]):
    print(f"  {acc}: {n}")
print(f"\nProductos top:")
for prod,info in sorted(prod_count.items(),key=lambda x:-x[1]["qty"])[:10]:
    print(f"  {info['qty']:3}u  {prod}")
