#!/usr/bin/env python3
"""
Contabilidad MELI FULL — Backfill desde 2026-01-01 hasta hoy.
Itera todas las órdenes y reclamos, agrupa por día, mes, cuenta, modelo.
Genera XLSX profesional con:
  1. Dashboard ejecutivo
  2. Ventas Diarias (todas las fechas con venta)
  3. Mensual (resumen por mes)
  4. Por Cuenta (resumen total por cuenta)
  5. Por Modelo (ventas por producto/modelo)
  6. Reclamos (todos detallados)
  7. Reclamos x Motivo (mensual)
  8. Reclamos x Modelo (mensual)
  9. Devoluciones $
"""
import os, requests, json, sys, time
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

OUTPUT = sys.argv[1] if len(sys.argv) > 1 else "contabilidad_meli.xlsx"
START_DATE = "2026-01-01"

APP_ID = os.environ["MELI_APP_ID"]
APP_SECRET = os.environ["MELI_APP_SECRET"]

ACCOUNTS = [
    ("JUAN", "MELI_REFRESH_TOKEN"),
    ("CLARIBEL", "MELI_REFRESH_TOKEN_CLARIBEL"),
    ("ASVA", "MELI_REFRESH_TOKEN_ASVA"),
    ("RAYMUNDO", "MELI_REFRESH_TOKEN_RAYMUNDO"),
    ("DILCIE", "MELI_REFRESH_TOKEN_DILCIE"),
    ("MILDRED", "MELI_REFRESH_TOKEN_MILDRED"),
    ("BREN", "MELI_REFRESH_TOKEN_BREN"),
]

REASON_LABELS = {
    "PDD9943": "No es original (imitación)",
    "PDD9944": "Producto defectuoso/dañado",
    "PNR9501": "No llegó el producto",
    "PDD8975": "Producto incompleto",
    "PDD9945": "Producto distinto al publicado",
    "PDD8974": "Diferente al pedido",
    "PDD8973": "No le gustó",
    "PDR9526": "No recibió el producto",
    "PDD9939": "Diferente al publicado",
}

def categorize_product(title):
    if not title: return ("Otro", "Sin clasificar")
    tl = title.lower()
    if "go 4" in tl or "go4" in tl: return ("Bocina", "JBL Go 4")
    if "go 3" in tl or "go3" in tl: return ("Bocina", "JBL Go 3")
    if "go essential" in tl: return ("Bocina", "JBL Go Essential")
    if "flip 7" in tl or "flip7" in tl: return ("Bocina", "JBL Flip 7")
    if "flip 6" in tl: return ("Bocina", "JBL Flip 6")
    if "charge 6" in tl: return ("Bocina", "JBL Charge 6")
    if "charge 5" in tl: return ("Bocina", "JBL Charge 5")
    if "grip" in tl and "jbl" in tl: return ("Bocina", "JBL Grip")
    if "clip 5" in tl: return ("Bocina", "JBL Clip 5")
    if "srs-xb100" in tl or "xb100" in tl: return ("Bocina", "Sony XB100")
    if "bocina" in tl or "parlante" in tl or "altavoz" in tl: return ("Bocina", "Otra bocina")
    if "armaf" in tl:
        if "club de nuit" in tl: return ("Perfume", "Armaf Club De Nuit")
        if "odyssey" in tl: return ("Perfume", "Armaf Odyssey")
        return ("Perfume", "Armaf otros")
    if "lattafa" in tl: return ("Perfume", "Lattafa")
    if "perfume" in tl or "edp" in tl or "edt" in tl: return ("Perfume", "Otro perfume")
    if "buds" in tl or "auriculares" in tl: return ("Audífonos", "Auriculares")
    return ("Otro", title[:30])

def to_cdmx(iso_str):
    if not iso_str: return None
    try:
        s = iso_str.replace("Z", "+00:00")
        d = datetime.fromisoformat(s).astimezone(timezone(timedelta(hours=-6)))
        return d
    except Exception:
        return None

now_cdmx = datetime.now(timezone(timedelta(hours=-6)))
today_label = now_cdmx.strftime("%Y-%m-%d")
start_dt = datetime.fromisoformat(START_DATE).replace(tzinfo=timezone(timedelta(hours=-6)))
end_dt = (now_cdmx + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
date_from_utc = start_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
date_to_utc = end_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")

print(f"📅 Rango: {START_DATE} → {today_label} ({(end_dt - start_dt).days} días)")
print(f"📊 Construyendo: {OUTPUT}\n")

# Per-day by account: { (date, account): {orders, qty, bruto, fees, ship, returns_count, returns_amt} }
daily = defaultdict(lambda: {"orders":0,"qty":0,"bruto":0,"fees":0,"ship":0,"returns_count":0,"returns_amt":0})
# Per-account totals
per_account_totals = defaultdict(lambda: {"orders":0,"qty":0,"bruto":0,"fees":0,"ship":0,"returns_count":0,"returns_amt":0,"nick":""})
# By model
by_model = defaultdict(lambda: {"orders":0,"qty":0,"bruto":0})
# By month
by_month = defaultdict(lambda: {"orders":0,"qty":0,"bruto":0,"fees":0,"ship":0,"returns_count":0,"returns_amt":0})
all_claims = []
# Reclamos x motivo x mes
reclamos_motivo_mes = defaultdict(lambda: defaultdict(int))
# Reclamos x modelo x mes
reclamos_modelo_mes = defaultdict(lambda: defaultdict(int))

for label, env_var in ACCOUNTS:
    RT = os.environ.get(env_var, "")
    if not RT:
        print(f"[{label}] sin token — skip")
        continue
    try:
        r = requests.post("https://api.mercadolibre.com/oauth/token",
            data={"grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":RT},
            timeout=20).json()
        H = {"Authorization":f"Bearer {r['access_token']}"}
        me = requests.get("https://api.mercadolibre.com/users/me", headers=H, timeout=10).json()
        USER_ID = me["id"]
    except Exception as e:
        print(f"[{label}] auth error: {e}")
        continue
    
    per_account_totals[label]["nick"] = me.get("nickname","")
    print(f"[{label}] {me.get('nickname','')} ({USER_ID}) — descargando órdenes...")
    
    orders_count = 0
    offset = 0
    PAGE = 50
    while True:
        url = f"https://api.mercadolibre.com/orders/search?seller={USER_ID}&order.date_created.from={date_from_utc}&order.date_created.to={date_to_utc}&sort=date_asc&limit={PAGE}&offset={offset}"
        try:
            rr = requests.get(url, headers=H, timeout=25).json()
        except Exception as e:
            print(f"  retry offset={offset}"); time.sleep(3); continue
        results = rr.get("results", [])
        if not results: break
        
        for o in results:
            st = o.get("status","")
            amt = o.get("total_amount", 0) or 0
            d = to_cdmx(o.get("date_created"))
            if not d: continue
            day_key = d.strftime("%Y-%m-%d")
            month_key = d.strftime("%Y-%m")
            
            had_paid = False; refund = 0
            for pay in o.get("payments", []):
                if pay.get("status") == "approved": had_paid = True
                if pay.get("status") in ("refunded","charged_back"):
                    refund += pay.get("transaction_amount", 0) or 0
            
            if st == "cancelled" and not had_paid:
                continue  # cancelado sin pago = no es venta
            
            if refund > 0:
                daily[(day_key, label)]["returns_count"] += 1
                daily[(day_key, label)]["returns_amt"] += refund
                per_account_totals[label]["returns_count"] += 1
                per_account_totals[label]["returns_amt"] += refund
                by_month[month_key]["returns_count"] += 1
                by_month[month_key]["returns_amt"] += refund
            
            if st not in ("paid","shipped","delivered"):
                continue
            
            qty_total = sum(oi.get("quantity",0) for oi in o.get("order_items",[]))
            
            # Fee + envío via order detail (1 request per order, but only para órdenes pagadas)
            fee = 0; ship_cost = 0
            try:
                od_url = f"https://api.mercadolibre.com/orders/{o.get('id')}"
                od = requests.get(od_url, headers=H, timeout=10).json()
                for pay in od.get("payments",[]):
                    if pay.get("status") == "approved":
                        fee += pay.get("marketplace_fee",0) or 0
                sh_id = od.get("shipping",{}).get("id")
                if sh_id:
                    sd = requests.get(f"https://api.mercadolibre.com/shipments/{sh_id}", headers=H, timeout=10).json()
                    so = sd.get("shipping_option",{}) or {}
                    ship_cost = max(0, (so.get("list_cost",0) or 0) - (so.get("cost",0) or 0))
            except Exception:
                pass
            
            daily[(day_key, label)]["orders"] += 1
            daily[(day_key, label)]["qty"] += qty_total
            daily[(day_key, label)]["bruto"] += amt
            daily[(day_key, label)]["fees"] += fee
            daily[(day_key, label)]["ship"] += ship_cost
            
            per_account_totals[label]["orders"] += 1
            per_account_totals[label]["qty"] += qty_total
            per_account_totals[label]["bruto"] += amt
            per_account_totals[label]["fees"] += fee
            per_account_totals[label]["ship"] += ship_cost
            
            by_month[month_key]["orders"] += 1
            by_month[month_key]["qty"] += qty_total
            by_month[month_key]["bruto"] += amt
            by_month[month_key]["fees"] += fee
            by_month[month_key]["ship"] += ship_cost
            
            for oi in o.get("order_items",[]):
                title = oi.get("item",{}).get("title","")
                cat, model = categorize_product(title)
                model_key = f"{cat}|{model}"
                by_model[model_key]["orders"] += 1
                by_model[model_key]["qty"] += oi.get("quantity",0)
                by_model[model_key]["bruto"] += (oi.get("unit_price",0) or 0) * (oi.get("quantity",0) or 0)
            
            orders_count += 1
        
        offset += PAGE
        total = rr.get("paging",{}).get("total", 0)
        if offset >= total: break
        if offset % 500 == 0: print(f"  ... {offset}/{total} órdenes")
    
    print(f"[{label}] ✅ {orders_count} órdenes válidas | bruto ${per_account_totals[label]['bruto']:,.0f}")
    
    # ===== Reclamos =====
    print(f"[{label}] descargando reclamos...")
    for st_filter in ("opened","closed"):
        offset = 0
        while True:
            try:
                cu = f"https://api.mercadolibre.com/post-purchase/v1/claims/search?status={st_filter}&limit=50&offset={offset}"
                rc = requests.get(cu, headers=H, timeout=20).json()
            except Exception:
                break
            data = rc.get("data", []) if isinstance(rc, dict) else []
            if not data: break
            for cl in data:
                d = to_cdmx(cl.get("date_created"))
                if not d or d < start_dt: continue
                month_key = d.strftime("%Y-%m")
                reason_id = (cl.get("reason_id") or "").strip()
                reason_label = REASON_LABELS.get(reason_id, reason_id or "Sin motivo")
                # try get item title
                product_title = ""
                resource = cl.get("resource", "")
                if resource and resource.startswith("/orders/"):
                    try:
                        oid = resource.split("/")[-1]
                        odd = requests.get(f"https://api.mercadolibre.com/orders/{oid}", headers=H, timeout=8).json()
                        for it in odd.get("order_items",[]):
                            product_title = it.get("item",{}).get("title","")
                            if product_title: break
                    except Exception: pass
                cat, model = categorize_product(product_title)
                all_claims.append({
                    "account": label, "claim_id": cl.get("id"),
                    "order_id": cl.get("resource_id") or "",
                    "type": cl.get("type",""), "status": cl.get("status",""),
                    "reason_id": reason_id, "reason_label": reason_label,
                    "date": d.strftime("%Y-%m-%d %H:%M"),
                    "month": month_key, "model": model, "category": cat,
                    "title": product_title[:60],
                })
                reclamos_motivo_mes[month_key][reason_label] += 1
                reclamos_modelo_mes[month_key][model] += 1
            offset += 50
            if offset >= rc.get("paging",{}).get("total",0): break
    print(f"[{label}] reclamos cargados.")

# ============ BUILD XLSX ============
print(f"\n📝 Generando XLSX...")
wb = Workbook()
ws = wb.active; wb.remove(ws)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
TOTAL_FONT = Font(bold=True, size=11)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fmt_header(ws, row, cols):
    for col_idx, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

def fmt_money(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.number_format = '"$"#,##0.00'
    c.border = BORDER

def fmt_int(ws, row, col):
    c = ws.cell(row=row, column=col)
    c.number_format = '#,##0'; c.border = BORDER

# === HOJA 1: DASHBOARD ===
ws = wb.create_sheet("Dashboard")
ws["A1"] = f"📊 CONTABILIDAD MELI — {START_DATE} a {today_label}"
ws["A1"].font = Font(bold=True, size=16, color="1F4E78"); ws.merge_cells("A1:F1")

total_orders = sum(p["orders"] for p in per_account_totals.values())
total_bruto = sum(p["bruto"] for p in per_account_totals.values())
total_fees = sum(p["fees"] for p in per_account_totals.values())
total_ship = sum(p["ship"] for p in per_account_totals.values())
total_returns_amt = sum(p["returns_amt"] for p in per_account_totals.values())
total_iva = total_bruto - (total_bruto/1.16) if total_bruto > 0 else 0
total_net_pre_iva = total_bruto - total_fees - total_ship
total_net_post_iva = total_net_pre_iva - total_iva

kpis = [
    ("Período", f"{START_DATE} a {today_label}"),
    ("Órdenes totales", total_orders),
    ("Unidades vendidas", sum(p["qty"] for p in per_account_totals.values())),
    ("BRUTO total", total_bruto),
    ("- Comisión MELI", total_fees),
    ("- Envío seller", total_ship),
    ("- IVA (16%)", total_iva),
    ("= NET pre-IVA", total_net_pre_iva),
    ("= NET post-IVA", total_net_post_iva),
    ("Devoluciones $", total_returns_amt),
    ("Reclamos totales", len(all_claims)),
]
r = 3
for k, v in kpis:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True)
    cell = ws.cell(row=r, column=2, value=v)
    if isinstance(v, (int, float)) and ("$" in k or "BRUTO" in k or "NET" in k or "IVA" in k or "Devoluciones" in k or "MELI" in k or "envío" in k.lower() or "Envío" in k):
        cell.number_format = '"$"#,##0.00'
        if "NET" in k: cell.fill = GREEN_FILL; cell.font = Font(bold=True)
        if "Devoluciones" in k or "Comisión" in k: cell.fill = RED_FILL
    elif isinstance(v, int):
        cell.number_format = '#,##0'
    r += 1

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 30

# Por cuenta en dashboard
r += 2
ws.cell(row=r, column=1, value="💼 NET por cuenta").font = Font(bold=True, size=12)
r += 1
fmt_header(ws, r, ["Cuenta","Órdenes","Bruto","Fees","Envío","Devoluciones","NET pre-IVA","NET post-IVA"])
r += 1
for label, p in sorted(per_account_totals.items(), key=lambda x: -x[1]["bruto"]):
    if p["orders"] == 0: continue
    iva = p["bruto"] - (p["bruto"]/1.16) if p["bruto"] > 0 else 0
    net_pre = p["bruto"] - p["fees"] - p["ship"]
    net_post = net_pre - iva
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=p["orders"]); fmt_int(ws, r, 2)
    ws.cell(row=r, column=3, value=p["bruto"]); fmt_money(ws, r, 3)
    ws.cell(row=r, column=4, value=p["fees"]); fmt_money(ws, r, 4)
    ws.cell(row=r, column=5, value=p["ship"]); fmt_money(ws, r, 5)
    ws.cell(row=r, column=6, value=p["returns_amt"]); fmt_money(ws, r, 6)
    ws.cell(row=r, column=7, value=net_pre); fmt_money(ws, r, 7)
    ws.cell(row=r, column=8, value=net_post); fmt_money(ws, r, 8)
    ws.cell(row=r, column=8).fill = GREEN_FILL
    r += 1

# === HOJA 2: VENTAS DIARIAS ===
ws = wb.create_sheet("Ventas Diarias")
ws["A1"] = "📅 Ventas Diarias por Cuenta"
ws["A1"].font = Font(bold=True, size=14, color="1F4E78"); ws.merge_cells("A1:H1")
fmt_header(ws, 2, ["Fecha","Cuenta","Órdenes","Unidades","Bruto","Fees","Envío","NET pre-IVA"])
r = 3
sorted_keys = sorted(daily.keys())
for (day_key, acct) in sorted_keys:
    d = daily[(day_key, acct)]
    if d["orders"] == 0: continue
    ws.cell(row=r, column=1, value=day_key)
    ws.cell(row=r, column=2, value=acct)
    ws.cell(row=r, column=3, value=d["orders"]); fmt_int(ws, r, 3)
    ws.cell(row=r, column=4, value=d["qty"]); fmt_int(ws, r, 4)
    ws.cell(row=r, column=5, value=d["bruto"]); fmt_money(ws, r, 5)
    ws.cell(row=r, column=6, value=d["fees"]); fmt_money(ws, r, 6)
    ws.cell(row=r, column=7, value=d["ship"]); fmt_money(ws, r, 7)
    ws.cell(row=r, column=8, value=d["bruto"]-d["fees"]-d["ship"]); fmt_money(ws, r, 8)
    r += 1
for c, w in zip("ABCDEFGH", [12, 12, 10, 10, 12, 12, 12, 14]):
    ws.column_dimensions[c].width = w

# === HOJA 3: MENSUAL ===
ws = wb.create_sheet("Mensual")
ws["A1"] = "📆 Resumen Mensual"
ws["A1"].font = Font(bold=True, size=14, color="1F4E78"); ws.merge_cells("A1:I1")
fmt_header(ws, 2, ["Mes","Órdenes","Unidades","Bruto","Fees","Envío","IVA","NET post-IVA","Devoluciones $"])
r = 3
for m in sorted(by_month.keys()):
    p = by_month[m]
    iva = p["bruto"] - (p["bruto"]/1.16) if p["bruto"] > 0 else 0
    net_pre = p["bruto"] - p["fees"] - p["ship"]
    net_post = net_pre - iva
    ws.cell(row=r, column=1, value=m)
    ws.cell(row=r, column=2, value=p["orders"]); fmt_int(ws, r, 2)
    ws.cell(row=r, column=3, value=p["qty"]); fmt_int(ws, r, 3)
    ws.cell(row=r, column=4, value=p["bruto"]); fmt_money(ws, r, 4)
    ws.cell(row=r, column=5, value=p["fees"]); fmt_money(ws, r, 5)
    ws.cell(row=r, column=6, value=p["ship"]); fmt_money(ws, r, 6)
    ws.cell(row=r, column=7, value=iva); fmt_money(ws, r, 7)
    ws.cell(row=r, column=8, value=net_post); fmt_money(ws, r, 8); ws.cell(row=r, column=8).fill = GREEN_FILL
    ws.cell(row=r, column=9, value=p["returns_amt"]); fmt_money(ws, r, 9)
    r += 1
for c, w in zip("ABCDEFGHI", [10, 10, 10, 14, 14, 14, 14, 16, 14]):
    ws.column_dimensions[c].width = w

# === HOJA 4: POR CUENTA ===
ws = wb.create_sheet("Por Cuenta")
ws["A1"] = "💼 Totales por Cuenta"
ws["A1"].font = Font(bold=True, size=14, color="1F4E78"); ws.merge_cells("A1:I1")
fmt_header(ws, 2, ["Cuenta","Nick","Órdenes","Unidades","Bruto","Fees","Envío","NET pre-IVA","NET post-IVA"])
r = 3
for label in [a[0] for a in ACCOUNTS]:
    p = per_account_totals.get(label, {"orders":0})
    if p.get("orders", 0) == 0: continue
    iva = p["bruto"] - (p["bruto"]/1.16) if p["bruto"] > 0 else 0
    net_pre = p["bruto"] - p["fees"] - p["ship"]
    net_post = net_pre - iva
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=p.get("nick",""))
    ws.cell(row=r, column=3, value=p["orders"]); fmt_int(ws, r, 3)
    ws.cell(row=r, column=4, value=p["qty"]); fmt_int(ws, r, 4)
    ws.cell(row=r, column=5, value=p["bruto"]); fmt_money(ws, r, 5)
    ws.cell(row=r, column=6, value=p["fees"]); fmt_money(ws, r, 6)
    ws.cell(row=r, column=7, value=p["ship"]); fmt_money(ws, r, 7)
    ws.cell(row=r, column=8, value=net_pre); fmt_money(ws, r, 8)
    ws.cell(row=r, column=9, value=net_post); fmt_money(ws, r, 9); ws.cell(row=r, column=9).fill = GREEN_FILL
    r += 1
for c, w in zip("ABCDEFGHI", [12, 22, 10, 10, 14, 14, 14, 14, 14]):
    ws.column_dimensions[c].width = w

# === HOJA 5: POR MODELO ===
ws = wb.create_sheet("Por Modelo")
ws["A1"] = "📦 Ventas por Modelo"
ws["A1"].font = Font(bold=True, size=14, color="1F4E78"); ws.merge_cells("A1:E1")
fmt_header(ws, 2, ["Categoría","Modelo","Órdenes","Unidades","Bruto"])
r = 3
for k in sorted(by_model.keys(), key=lambda x: -by_model[x]["bruto"]):
    cat, model = k.split("|", 1)
    p = by_model[k]
    if p["orders"] == 0: continue
    ws.cell(row=r, column=1, value=cat)
    ws.cell(row=r, column=2, value=model)
    ws.cell(row=r, column=3, value=p["orders"]); fmt_int(ws, r, 3)
    ws.cell(row=r, column=4, value=p["qty"]); fmt_int(ws, r, 4)
    ws.cell(row=r, column=5, value=p["bruto"]); fmt_money(ws, r, 5)
    r += 1
for c, w in zip("ABCDE", [12, 28, 10, 10, 14]):
    ws.column_dimensions[c].width = w

# === HOJA 6: RECLAMOS ===
ws = wb.create_sheet("Reclamos")
ws["A1"] = "⚠️ Reclamos detallados"
ws["A1"].font = Font(bold=True, size=14, color="C00000"); ws.merge_cells("A1:I1")
fmt_header(ws, 2, ["Fecha","Cuenta","Claim ID","Order ID","Tipo","Status","Motivo","Modelo","Producto"])
r = 3
for cl in sorted(all_claims, key=lambda x: x["date"], reverse=True):
    ws.cell(row=r, column=1, value=cl["date"])
    ws.cell(row=r, column=2, value=cl["account"])
    ws.cell(row=r, column=3, value=cl["claim_id"])
    ws.cell(row=r, column=4, value=cl["order_id"])
    ws.cell(row=r, column=5, value=cl["type"])
    ws.cell(row=r, column=6, value=cl["status"])
    ws.cell(row=r, column=7, value=cl["reason_label"])
    ws.cell(row=r, column=8, value=cl["model"])
    ws.cell(row=r, column=9, value=cl["title"])
    if cl["type"] == "mediations": ws.cell(row=r, column=5).fill = RED_FILL
    r += 1
for c, w in zip("ABCDEFGHI", [16, 12, 12, 14, 14, 12, 28, 18, 40]):
    ws.column_dimensions[c].width = w

# === HOJA 7: RECLAMOS x MOTIVO x MES ===
ws = wb.create_sheet("Reclamos x Motivo")
ws["A1"] = "⚠️ Reclamos por Motivo (mensual)"
ws["A1"].font = Font(bold=True, size=14, color="C00000")
all_motivos = sorted({m for monthd in reclamos_motivo_mes.values() for m in monthd.keys()})
fmt_header(ws, 2, ["Mes"] + all_motivos + ["TOTAL"])
r = 3
for mes in sorted(reclamos_motivo_mes.keys()):
    ws.cell(row=r, column=1, value=mes)
    total = 0
    for i, mt in enumerate(all_motivos, 2):
        c = reclamos_motivo_mes[mes].get(mt, 0)
        ws.cell(row=r, column=i, value=c if c else "")
        total += c
    ws.cell(row=r, column=len(all_motivos)+2, value=total).font = Font(bold=True)
    r += 1
ws.column_dimensions["A"].width = 10
for i in range(2, len(all_motivos)+3):
    ws.column_dimensions[get_column_letter(i)].width = 22

# === HOJA 8: RECLAMOS x MODELO x MES ===
ws = wb.create_sheet("Reclamos x Modelo")
ws["A1"] = "⚠️ Reclamos por Modelo (mensual)"
ws["A1"].font = Font(bold=True, size=14, color="C00000")
all_modelos = sorted({m for monthd in reclamos_modelo_mes.values() for m in monthd.keys()})
fmt_header(ws, 2, ["Mes"] + all_modelos + ["TOTAL"])
r = 3
for mes in sorted(reclamos_modelo_mes.keys()):
    ws.cell(row=r, column=1, value=mes)
    total = 0
    for i, md in enumerate(all_modelos, 2):
        c = reclamos_modelo_mes[mes].get(md, 0)
        ws.cell(row=r, column=i, value=c if c else "")
        total += c
    ws.cell(row=r, column=len(all_modelos)+2, value=total).font = Font(bold=True)
    r += 1
ws.column_dimensions["A"].width = 10
for i in range(2, len(all_modelos)+3):
    ws.column_dimensions[get_column_letter(i)].width = 18

wb.save(OUTPUT)
print(f"\n✅ {OUTPUT} generado.")
print(f"   Total órdenes: {total_orders} | Bruto: ${total_bruto:,.2f} | NET post-IVA: ${total_net_post_iva:,.2f}")
print(f"   Reclamos: {len(all_claims)}")
