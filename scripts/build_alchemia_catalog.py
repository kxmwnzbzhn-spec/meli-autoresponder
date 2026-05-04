"""Genera catalogo XLSX de los 22 perfumes The Alchemia Lab con EAN-13 valido + sube a Drive."""
import os, io, hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRODUCTS = [
    # (name, gender, family, image_url, product_url)
    ("Xibalba Royal", "Unisex", "Oud Amaderado Oscuro",
     "https://thealchemialab.com/wp-content/uploads/2026/05/Xibalba-Royal-Eau-de-Parfum.png",
     "https://thealchemialab.com/product/xibalba-royal-eau-de-parfum-100ml-2/"),
    ("Tlaloc Intenso", "Unisex", "Acuatico Amaderado Intenso",
     "https://thealchemialab.com/wp-content/uploads/2026/05/img_7575.png",
     "https://thealchemialab.com/product/tlaloc-intenso-eau-de-parfum-100ml-2/"),
    ("Cenote Azul", "Unisex", "Acuatico Marino Fresco",
     "https://thealchemialab.com/wp-content/uploads/2026/05/Frasco-de-perfume-Cenote-Azul.png",
     "https://thealchemialab.com/product/cenote-azul-eau-de-parfum-100ml-2/"),
    ("Xochicopal", "Unisex", "Ambar Especiado Ceremonial",
     "https://thealchemialab.com/wp-content/uploads/2026/03/1B2912E3-6D32-4651-B923-AFED16F7BB26.png",
     "https://thealchemialab.com/product/xochicopal-eau-de-parfum-100-ml-perfume-unisex-ambar-especiado-ceremonial/"),
    ("Horizonte de Agua Azul", "Unisex", "Acuatico Premium",
     "https://thealchemialab.com/wp-content/uploads/2026/03/ChatGPT-Image-22-mar-2026-08_47_55-a.m.png",
     "https://thealchemialab.com/product/horizonte-de-agua-azul-eau-de-parfum-100-ml-perfume-unisex-acuatico-premium/"),
    ("Tlalocan", "Unisex", "Acuatico Fresco Premium",
     "https://thealchemialab.com/wp-content/uploads/2026/03/2E34A560-4DCA-419E-AA9D-DD5EA6C34111.png",
     "https://thealchemialab.com/product/tlalocan-eau-de-parfum-100-ml-perfume-unisex-acuatico-fresco-premium/"),
    ("Guerrero Sol", "Hombre", "Calido Especiado",
     "https://thealchemialab.com/wp-content/uploads/2026/03/DEB42C0A-F420-4299-9DC8-9CAE89D5D7B2.png",
     "https://thealchemialab.com/product/guerrero-sol-eau-de-parfum-100-ml-perfume-masculino-calido-especiado/"),
    ("Flor de la Noche", "Mujer", "Floral Nocturno",
     "https://thealchemialab.com/wp-content/uploads/2026/03/B636BC80-125D-400B-B083-7C140D2663A9.png",
     "https://thealchemialab.com/product/flor-de-la-noche-eau-de-parfum-100-ml-perfume-femenino-floral-nocturno/"),
    ("Dark Oud Cacao", "Unisex", "Oscuro Gourmand",
     "https://thealchemialab.com/wp-content/uploads/2026/03/95C6C303-B96C-4D7D-A741-A676819432A3.png",
     "https://thealchemialab.com/product/dark-oud-cacao-eau-de-parfum-100-ml-perfume-unisex-oscuro-gourmand/"),
    ("Manantial del Valle Real", "Unisex", "Acuatico Fresco",
     "https://thealchemialab.com/wp-content/uploads/2026/03/7FB9C94C-30B3-4B21-8E7F-2AE26B3517B5.png",
     "https://thealchemialab.com/product/manantial-del-valle-real-eau-de-parfum-100-ml-perfume-unisex-fresco-acuatico/"),
    ("Ixtli Nocturno", "Unisex", "Gourmand Oscuro",
     "https://thealchemialab.com/wp-content/uploads/2026/03/4BF2597E-EF13-4D1D-9AF1-169383972031.png",
     "https://thealchemialab.com/product/ixtli-nocturno-eau-de-parfum-100-ml-perfume-unisex-gourmand-oscuro/"),
    ("Dominio del Fuego", "Hombre", "Especiado Intenso",
     "https://thealchemialab.com/wp-content/uploads/2026/03/1FD6EB59-4119-4B78-BC6D-5230CDE24A75.png",
     "https://thealchemialab.com/product/dominio-del-fuego-eau-de-parfum-100-ml-perfume-masculino-especiado-intenso/"),
    ("Templo Oscuro", "Unisex", "Resino Amaderado",
     "https://thealchemialab.com/wp-content/uploads/2026/03/0919D8F2-49F6-479E-B7E4-E8EE98B87C2D.png",
     "https://thealchemialab.com/product/templo-oscuro-eau-de-parfum-100-ml-perfume-unisex-resino-amaderado/"),
    ("Copalli Ixtlan", "Unisex", "Incienso Ritual",
     "https://thealchemialab.com/wp-content/uploads/2026/03/304748E0-6AB5-444F-A22C-40E0369CB354.png",
     "https://thealchemialab.com/product/copalli-ixtlan-eau-de-parfum-100-ml-perfume-unisex-incienso-ritual/"),
    ("Tonalli Ambar", "Unisex", "Ambar Dulce",
     "https://thealchemialab.com/wp-content/uploads/2026/03/7E79945B-59C4-4241-8EAC-B1EB8B9C197C.png",
     "https://thealchemialab.com/product/tonallli-ambar-eau-de-parfum-100-ml-perfume-unisex-ambar-dulce/"),
    ("Velo Celestial", "Unisex", "Floral Suave",
     "https://thealchemialab.com/wp-content/uploads/2026/03/EE31F06B-38C4-4678-8E8B-32BE5779DA5D.png",
     "https://thealchemialab.com/product/velo-celestial-eau-de-parfum-100-ml-perfume-unisex-floral-suave/"),
    ("Luz del Desierto", "Unisex", "Amaderado Especiado",
     "https://thealchemialab.com/wp-content/uploads/2026/03/0A08F7BB-9974-40FE-B053-4645D9C5A9A8.png",
     "https://thealchemialab.com/product/luz-del-desierto-eau-de-parfum-100-ml-perfume-unisex-amaderado-especiado/"),
    ("Fuerza de Kukulkan", "Hombre", "Masculino",
     "https://thealchemialab.com/wp-content/uploads/2026/02/1E348E4F-D5EF-404A-A3AF-091E50054DCE.png",
     "https://thealchemialab.com/product/fuerza-de-kukulkan-eau-de-parfum-100-ml-perfume-masculino/"),
    ("Obsidian Eclipse", "Unisex", "LV Studio",
     "https://thealchemialab.com/wp-content/uploads/2025/11/img_7577.png",
     "https://thealchemialab.com/product/obsidian-eclipse-eau-de-parfum-100ml-lv-perfume-studio/"),
    ("Alma de Tenochtitlan", "Unisex", "Mexica",
     "https://thealchemialab.com/wp-content/uploads/2025/09/img_7526.png",
     "https://thealchemialab.com/product/alma-de-tenochtitlan-eau-de-parfum-100ml/"),
    ("Rosa del Viento", "Unisex", "Floral",
     "https://thealchemialab.com/wp-content/uploads/2025/12/A7D47AD3-8E4E-492F-992C-84F1E3BE243F.png",
     "https://thealchemialab.com/product/rosa-del-viento-eau-de-parfum/"),
    ("Quinto Aliento", "Unisex", "Mistico",
     "https://thealchemialab.com/wp-content/uploads/2026/03/F785E8C9-E40F-41CA-B276-4B077E144FA7.png",
     "https://thealchemialab.com/product/quinto-aliento-eau-de-parfum-100-ml-perfume-unisex-mistico/"),
]

# EAN-13 generator: usamos prefijo 290 (uso interno/marca propia restringida) +
# 9 dígitos derivados del nombre del producto (hash), + check digit calculado
def ean13(seed: str) -> str:
    """Genera EAN-13 valido con prefijo 290 (uso interno) y check digit correcto."""
    h = hashlib.md5(seed.encode()).hexdigest()
    nine = "".join(c for c in h if c.isdigit())[:9].ljust(9, "0")
    body = "290" + nine  # 12 digitos
    # Check digit EAN-13
    s = sum(int(d) * (3 if i % 2 else 1) for i, d in enumerate(body))
    check = (10 - (s % 10)) % 10
    return body + str(check)

PRICE_MELI = 799

# Build Workbook
wb = Workbook()
ws = wb.active
ws.title = "Alchemia Lab Catalog"

headers = ["#", "SKU", "Nombre Producto", "Marca", "EAN-13", "Precio MELI",
           "Volumen", "Tipo", "Genero", "Familia Olfativa",
           "Foto principal", "URL Producto", "Titulo MELI Sugerido"]

# Header row
header_fill = PatternFill("solid", fgColor="2C3E50")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style="thin", color="BDC3C7"),
                right=Side(style="thin", color="BDC3C7"),
                top=Side(style="thin", color="BDC3C7"),
                bottom=Side(style="thin", color="BDC3C7"))

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border

# Data rows
for i, (name, gender, family, img, url) in enumerate(PRODUCTS, 1):
    sku = f"TAL-{name.replace(' ','-').upper()[:25]}-100"
    ean = ean13(name)
    titulo = f"{name} Eau de Parfum 100ml The Alchemia Lab Perfume {gender}"[:60]
    row = [
        i, sku, name, "The Alchemia Lab", ean, PRICE_MELI,
        "100 ml", "Eau de Parfum", gender, family,
        img, url, titulo
    ]
    for col, val in enumerate(row, 1):
        c = ws.cell(row=i+1, column=col, value=val)
        c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=True)
        if col == 5:  # EAN
            c.font = Font(name="Consolas", size=10, bold=True)
        elif col == 6:  # precio
            c.font = Font(name="Arial", size=10, bold=True, color="27AE60")
            c.number_format = "$#,##0"
        elif col in (11, 12):  # URLs
            c.font = Font(color="3498DB", underline="single")
            c.hyperlink = val

# Column widths
widths = {1:5, 2:24, 3:24, 4:18, 5:16, 6:11, 7:9, 8:14, 9:9, 10:24, 11:50, 12:50, 13:55}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 38
for r in range(2, len(PRODUCTS)+2):
    ws.row_dimensions[r].height = 32

# Freeze header
ws.freeze_panes = "A2"

# Sheet 2: instrucciones
ws2 = wb.create_sheet("Notas")
notas = [
    "CATALOGO THE ALCHEMIA LAB — 22 perfumes para MELI",
    "",
    f"Total productos: {len(PRODUCTS)}",
    f"Precio MELI sugerido: ${PRICE_MELI} MXN cada uno",
    "Volumen: 100 ml todos",
    "Tipo: Eau de Parfum (EDP)",
    "",
    "EAN-13 generados con prefijo 290 (uso interno restringido GS1) + check digit valido.",
    "Estos EAN son codigos internos para inventario/MELI sin codigo universal real.",
    "Para MELI, indicar 'Sin codigo universal' o usar este EAN como referencia interna.",
    "",
    "FOTOS: Cada producto tiene 1 foto principal disponible desde la web.",
    "MELI requiere minimo 3 fotos por publicacion. Necesitas fotos adicionales:",
    "  - frente del frasco (la que tenemos)",
    "  - parte trasera con etiqueta",
    "  - vista superior del frasco",
    "  - foto en contexto/ambiente",
    "",
    "Distribuicion por genero:",
    "  - Unisex: 18 productos",
    "  - Hombre: 3 productos (Guerrero Sol, Dominio del Fuego, Fuerza de Kukulkan)",
    "  - Mujer: 1 producto (Flor de la Noche)",
]
for i, line in enumerate(notas, 1):
    c = ws2.cell(row=i, column=1, value=line)
    if i == 1:
        c.font = Font(size=14, bold=True, color="2C3E50")
ws2.column_dimensions["A"].width = 90

out = os.environ.get("OUT_PATH", "alchemia_lab_catalog.xlsx")
wb.save(out)
print(f"✅ XLSX: {out}")

# Imprimir resumen para terminal
print("\n=== EAN generados ===")
for i, (name, *_) in enumerate(PRODUCTS, 1):
    print(f"  {i:2}. {name:<28} EAN-13: {ean13(name)}")
