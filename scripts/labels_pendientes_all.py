"""ETIQUETAS PENDIENTES — TODAS LAS CUENTAS — formato 4x6.
Filtro estricto: status=ready_to_ship AND substatus IN (ready_to_print, printed).
Ventana de 60 días. Single PDF combinado + XLSX con breakdown por cuenta.
"""
import os, io, time, requests
from datetime import datetime, timedelta, timezone
from collections import defaultdict, Counter
from pypdf import PdfReader, PdfWriter, PageObject, Transformation
from reportlab.pdfgen import canvas
from reportlab.lib.colors import Color
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pdf2image import convert_from_bytes
from PIL import ImageOps

APP_ID="5211907102822632"
APP_SECRET=os.environ["MELI_APP_SECRET"]
ACCS={
    "Juan":     os.environ.get("MELI_REFRESH_TOKEN_JUAN") or os.environ.get("MELI_REFRESH_TOKEN"),
    "Raymundo": os.environ.get("MELI_REFRESH_TOKEN_RAYMUNDO"),
    "Wilbert":  os.environ.get("MELI_REFRESH_TOKEN_WILBERT"),
    "Claribel": os.environ.get("MELI_REFRESH_TOKEN_CLARIBEL"),
    "Asva":     os.environ.get("MELI_REFRESH_TOKEN_ASVA"),
    "Mildred":  os.environ.get("MELI_REFRESH_TOKEN_MILDRED"),
    "Dilcie":   os.environ.get("MELI_REFRESH_TOKEN_DILCIE"),
    "Bren":     os.environ.get("MELI_REFRESH_TOKEN_BREN"),
    "Yc_New":   os.environ.get("MELI_REFRESH_TOKEN_YC_NEW"),
}
TZ = timezone(timedelta(hours=-6))
PAGE_W = 4*72
PAGE_H = 6*72

# SOLO estas combinaciones (status, substatus) se incluyen
ALLOWED = {("ready_to_ship","ready_to_print"), ("ready_to_ship","printed")}


def tok(rt):
    if not rt: return None
    r=requests.post("https://api.mercadolibre.com/oauth/token",data={
        "grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":rt}).json()
    return r.get("access_token")


def _parse_color(text):
    """Devuelve color normalizado de un string (título o variation value)."""
    if not text: return None
    tl=" "+text.lower()+" "
    col_map=[("camuflaj","Camuflaje"),("camo","Camuflaje"),("azul marino","Azul Marino"),
             ("aqua","Aqua"),("celeste","Celeste"),("negr","Negro"),(" black","Negro"),
             ("roj","Rojo"),(" red","Rojo"),("rosa","Rosa"),("pink","Rosa"),
             ("morad","Morado"),("violeta","Morado"),("purple","Morado"),
             (" azul","Azul"),(" blue","Azul"),("blanco","Blanco"),("white","Blanco"),
             ("verde","Verde"),("green","Verde"),("amarillo","Amarillo"),("yellow","Amarillo"),
             ("naranja","Naranja"),("orange","Naranja"),("gris","Gris"),("gray","Gris"),
             ("plateado","Plata"),("silver","Plata"),("dorado","Dorado"),("gold","Dorado")]
    for k,v in col_map:
        if k in tl: return v
    return None


def get_variant_color(item_obj, H):
    """Obtiene COLOR del variant elegido. Fuentes en orden:
    1) item.variation_attributes  (lo manda MELI con la orden cuando hay variantes)
    2) GET /items/{id}/variations/{variation_id} → attribute_combinations
    """
    attrs = item_obj.get("variation_attributes") or []
    for a in attrs:
        if a.get("id") == "COLOR" or "color" in (a.get("name","") or "").lower():
            v = a.get("value_name") or ""
            c = _parse_color(v)
            if c: return c
    iid = item_obj.get("id"); vid = item_obj.get("variation_id")
    if iid and vid:
        try:
            r = requests.get(f"https://api.mercadolibre.com/items/{iid}/variations/{vid}",
                             headers=H, timeout=8)
            if r.status_code == 200:
                for ac in (r.json().get("attribute_combinations") or []):
                    if ac.get("id")=="COLOR" or "color" in (ac.get("name","") or "").lower():
                        return _parse_color(ac.get("value_name"))
        except: pass
    return None


def clean_title(title, item_obj=None, H=None):
    t=(title or "").strip()
    for w in ["Bocina ","bocina ","Parlante ","parlante ","Altavoz ","altavoz ","Speaker ","speaker ",
              "JBL ","jbl ","Jbl ","Sony ","SONY ","Bose ","BOSE "]:
        t=t.replace(w,"")
    tl=t.lower()
    if "go 4" in tl or "go4" in tl: model="Go 4"
    elif "go 3" in tl or "go3" in tl: model="Go 3"
    elif "clip 5" in tl or "clip5" in tl: model="Clip 5"
    elif "charge 6" in tl or "charge6" in tl: model="Charge 6"
    elif "flip 7" in tl or "flip7" in tl: model="Flip 7"
    elif "grip" in tl: model="Grip"
    elif "xb100" in tl: model="Sony XB100"
    elif "soundlink" in tl: model="Bose SoundLink"
    else: model=t[:30]
    # PRIORIDAD: variant > título
    color = None
    if item_obj is not None:
        color = get_variant_color(item_obj, H)
    if not color:
        color = _parse_color(t)
    return f"{model} {color}" if color else model


_COND_CACHE={}
def get_condition(item_obj, H):
    """Devuelve 'used' / 'new' / None. Usa cache + fallback a /items/{id}."""
    cond = item_obj.get("condition")
    if cond: return cond
    iid = item_obj.get("id")
    if not iid: return None
    if iid in _COND_CACHE: return _COND_CACHE[iid]
    try:
        r = requests.get(f"https://api.mercadolibre.com/items/{iid}",
                         headers=H, timeout=8,
                         params={"attributes":"condition"})
        if r.status_code == 200:
            cond = (r.json() or {}).get("condition")
            _COND_CACHE[iid] = cond
            return cond
    except Exception:
        pass
    _COND_CACHE[iid] = None
    return None


def detect_content_bbox(pdf_bytes, page_idx=0):
    try:
        imgs = convert_from_bytes(pdf_bytes, dpi=72,
                                  first_page=page_idx+1, last_page=page_idx+1)
        if not imgs: return None
        img = imgs[0].convert("L")
        bw = img.point(lambda p: 0 if p < 245 else 255, mode="L")
        inv = ImageOps.invert(bw)
        bbox = inv.getbbox()
        if not bbox: return None
        x0, y0_top, x1, y1_top = bbox
        img_w, img_h = img.size
        pdf_y0 = img_h - y1_top
        pdf_y1 = img_h - y0_top
        m = 2
        return (max(0, x0-m), max(0, pdf_y0-m),
                min(img_w, x1-x0+2*m), min(img_h, pdf_y1-pdf_y0+2*m))
    except Exception:
        return None


def render_header(s, header_h):
    has_used = bool(s.get("has_used"))
    usado_strip = 14 if has_used else 0
    total_h = header_h + usado_strip
    buf=io.BytesIO()
    c=canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    cx = PAGE_W/2.0
    # Banda amarilla para datos
    c.setFillColor(Color(1, 0.96, 0.74))
    c.rect(0, PAGE_H-total_h, PAGE_W, header_h, fill=1, stroke=0)
    # Strip rojo USADO arriba del todo (si aplica)
    if has_used:
        c.setFillColorRGB(0.85, 0.13, 0.13)
        c.rect(0, PAGE_H-usado_strip, PAGE_W, usado_strip, fill=1, stroke=0)
        c.setFillColorRGB(1,1,1)
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(cx, PAGE_H-11, "*** PRODUCTO USADO ***")
    # Texto en banda amarilla
    yellow_top = PAGE_H - usado_strip
    c.setFillColorRGB(0,0,0)
    c.setFont("Helvetica-Bold", 7.5)
    c.drawCentredString(cx, yellow_top-11, f"[{s['account'].upper()}] {s['buyer'][:30]} | Ship:{s['sid']}")
    big = s["comp_lines"][:4]; n=len(big); lh=16
    block_top=yellow_top-18; block_bot=PAGE_H-total_h+4
    block_h=block_top-block_bot; text_h=n*lh
    first_y = block_top - (block_h - text_h)/2.0 - 12
    c.setFont("Helvetica-Bold", 14)
    y=first_y
    for line in big:
        c.drawCentredString(cx, y, line[:30]); y-=lh
    c.showPage(); c.save()
    buf.seek(0)
    return PdfReader(buf).pages[0]


# === FASE 1: shipments ===
NOW = datetime.now(timezone.utc)
START = NOW - timedelta(days=60)
print(f"=== Ventana orders: {START.date()} a {NOW.date()} ===\n")
shipments=[]
acc_seen=Counter()  # cuántas (status,substatus) por cuenta

for acc, rt in ACCS.items():
    if not rt:
        print(f"--- {acc}: SIN TOKEN, skip"); continue
    print(f"--- {acc} ---")
    at = tok(rt)
    if not at:
        print(f"  no pude renovar token"); continue
    H = {"Authorization": f"Bearer {at}"}
    me = requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=15).json()
    uid = me.get("id")
    if not uid:
        print(f"  no uid"); continue
    orders=[]; offset=0
    while True:
        r=requests.get("https://api.mercadolibre.com/orders/search",headers=H,timeout=20,
            params={"seller":uid,"order.status":"paid",
                    "order.date_created.from":START.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "order.date_created.to":NOW.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                    "limit":50,"offset":offset}).json()
        res=r.get("results",[])
        if not res: break
        orders.extend(res); offset+=len(res)
        if offset>=r.get("paging",{}).get("total",0): break
    obs={}
    for o in orders:
        sid=(o.get("shipping") or {}).get("id")
        if sid: obs[sid]=o
    matches=0
    for sid, ord_o in obs.items():
        try:
            sh=requests.get(f"https://api.mercadolibre.com/shipments/{sid}",headers=H,timeout=10).json()
            st=sh.get("status"); sub=sh.get("substatus")
            acc_seen[(acc, st, sub or "(none)")] += 1
            if (st, sub) not in ALLOWED: continue
            items=ord_o.get("order_items",[])
            comp_lines=[]
            has_used=False
            for it in items:
                io_obj = it.get("item") or {}
                title = clean_title(io_obj.get("title",""), io_obj, H)
                qty = it.get("quantity",1)
                cond = get_condition(io_obj, H)
                if cond == "used":
                    has_used = True
                    comp_lines.append(f"USADO {title} x{qty}")
                else:
                    comp_lines.append(f"{title} x{qty}")
            buyer=(ord_o.get("buyer") or {}).get("nickname","?")
            deadline=None
            try:
                sla=requests.get(f"https://api.mercadolibre.com/shipments/{sid}/sla",headers=H,timeout=8).json()
                ed=sla.get("expected_date")
                if ed: deadline=datetime.fromisoformat(ed.replace("Z","+00:00")).astimezone(TZ)
            except: pass
            shipments.append({
                "sid":sid,"account":acc,"token":at,
                "comp_lines":comp_lines,"buyer":buyer,
                "status":st,"substatus":sub,"has_used":has_used,
                "deadline": deadline.strftime("%a %d %H:%M") if deadline else "s/d",
                "tracking": sh.get("tracking_number","")
            })
            matches += 1
            time.sleep(0.04)
        except Exception as e:
            print(f"  err shipment {sid}: {str(e)[:80]}")
    print(f"  matches: {matches}")

print(f"\n=== Total shipments seleccionados: {len(shipments)} ===")
shipments.sort(key=lambda s: (s["account"], s["substatus"], "/".join(s["comp_lines"])))


# === FASE 2: PDF combinado ===
print("\n=== Generando PDF 4x6 combinado ===")
writer = PdfWriter()
fail = []
ok_per_acc = Counter()
ok_per_subs = Counter()
for s in shipments:
    H = {"Authorization": f"Bearer {s['token']}"}
    try:
        r=requests.get("https://api.mercadolibre.com/shipment_labels",
                      headers=H, params={"shipment_ids":s["sid"],"response_type":"pdf"},
                      timeout=30)
        if r.status_code!=200 or not r.headers.get("content-type","").lower().startswith("application/pdf"):
            fail.append({"sid":s["sid"],"acc":s["account"],"sub":s["substatus"],"reason":r.text[:200]})
            print(f"  err {s['account']}/{s['sid']} sub={s['substatus']}: HTTP {r.status_code} {r.text[:140]}")
            continue
        raw = r.content
        lbl_pdf=PdfReader(io.BytesIO(raw))
        for pidx, label_page in enumerate(lbl_pdf.pages):
            box = label_page.cropbox if label_page.cropbox else label_page.mediabox
            lbl_x0=float(box.left); lbl_y0=float(box.bottom)
            lbl_w=float(box.width); lbl_h=float(box.height)
            cb = detect_content_bbox(raw, pidx)
            if cb:
                cx0,cy0,cw,ch = cb
                lbl_x0=float(box.left)+float(cx0)
                lbl_y0=float(box.bottom)+float(cy0)
                lbl_w=float(cw); lbl_h=float(ch)
            n_lines=min(len(s["comp_lines"]),4)
            header_h = 22 + n_lines*16
            label_area_h = PAGE_H - header_h
            new_page = PageObject.create_blank_page(width=PAGE_W, height=PAGE_H)
            sx = PAGE_W/lbl_w; sy = label_area_h/lbl_h
            op = (Transformation()
                  .translate(-lbl_x0, -lbl_y0)
                  .scale(sx, sy))
            new_page.merge_transformed_page(label_page, op)
            new_page.merge_page(render_header(s, header_h))
            writer.add_page(new_page)
            ok_per_acc[s["account"]] += 1
            ok_per_subs[(s["account"], s["substatus"])] += 1
    except Exception as e:
        fail.append({"sid":s["sid"],"acc":s["account"],"sub":s.get("substatus",""),"reason":f"{type(e).__name__}: {str(e)[:200]}"})
        print(f"  err {s['account']}/{s['sid']}: {type(e).__name__}: {str(e)[:140]}")
    time.sleep(0.08)

pdf_out = "ETIQUETAS_PENDIENTES_TODAS.pdf"
with open(pdf_out,"wb") as f: writer.write(f)
print(f"\n✅ PDF combinado: {pdf_out} ({len(writer.pages)} págs) | Fallidas: {len(fail)}")


# === FASE 3: XLSX ===
wb=Workbook(); ws=wb.active; ws.title="Pendientes todas"
hf=PatternFill("solid", fgColor="2C3E50")
hF=Font(bold=True, color="FFFFFF", size=11)
center=Alignment(horizontal="center", vertical="center", wrap_text=True)
border=Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
headers=["#","Cuenta","Substatus","Cond","Shipment ID","Comprador","PRODUCTO","Tracking","Deadline"]
for col,h in enumerate(headers,1):
    c=ws.cell(row=1,column=col,value=h)
    c.fill=hf; c.font=hF; c.alignment=center; c.border=border
acc_colors={"Juan":"D5E8D4","Raymundo":"DAE8FC","Wilbert":"FAD7A0","Claribel":"FFE6CC",
            "Asva":"E1D5E7","Mildred":"FFF2CC","Dilcie":"F5CBA7","Bren":"D4E6F1","Yc_New":"D5DBDB"}
used_fill = PatternFill("solid", fgColor="FF6B6B")
for i,s in enumerate(shipments,1):
    fill=PatternFill("solid", fgColor=acc_colors.get(s["account"],"FFFFFF"))
    cond_label = "USADO" if s.get("has_used") else "Nuevo"
    row=[i, s["account"], s["substatus"], cond_label, s["sid"], s["buyer"], "\n".join(s["comp_lines"]), s["tracking"], s["deadline"]]
    for col,val in enumerate(row,1):
        c=ws.cell(row=i+1, column=col, value=val)
        c.fill=fill; c.border=border
        c.alignment=Alignment(vertical="center", wrap_text=True)
        if col==7: c.font=Font(bold=True, size=10)
        if col==4 and val=="USADO":
            c.fill=used_fill; c.font=Font(bold=True, color="FFFFFF")
widths={1:5,2:11,3:14,4:8,5:14,6:24,7:50,8:24,9:14}
for col,w in widths.items(): ws.column_dimensions[chr(64+col)].width=w
for r in range(2, len(shipments)+2): ws.row_dimensions[r].height=28
ws.freeze_panes="A2"

# Hoja 2: breakdown por cuenta
ws2=wb.create_sheet("Por cuenta")
ws2.append(["Cuenta","ready_to_print","printed","TOTAL","USADOS"])
for c in ["A1","B1","C1","D1","E1"]: ws2[c].fill=hf; ws2[c].font=hF
totals = defaultdict(lambda: Counter())
usados = Counter()
for s in shipments:
    totals[s["account"]][s["substatus"]] += 1
    if s.get("has_used"): usados[s["account"]] += 1
for acc in sorted(totals.keys()):
    rtp = totals[acc]["ready_to_print"]
    pr  = totals[acc]["printed"]
    ws2.append([acc, rtp, pr, rtp+pr, usados[acc]])
ws2.append([])
ws2.append(["TOTAL", sum(t["ready_to_print"] for t in totals.values()),
            sum(t["printed"] for t in totals.values()),
            sum(t["ready_to_print"]+t["printed"] for t in totals.values()),
            sum(usados.values())])
for col in "ABCDE": ws2.column_dimensions[col].width=18

xlsx_out = "DESPACHO_PENDIENTES_TODAS.xlsx"
wb.save(xlsx_out)
print(f"✅ XLSX: {xlsx_out}")

print(f"\n=== BREAKDOWN POR CUENTA (PDFs OK) ===")
for acc in sorted(ok_per_acc.keys()):
    rtp = ok_per_subs[(acc,"ready_to_print")]
    pr  = ok_per_subs[(acc,"printed")]
    print(f"  {acc:10}  ready_to_print={rtp:4}  printed={pr:4}  TOTAL={ok_per_acc[acc]:4}")
print(f"  {'TOTAL':10}  {sum(ok_per_subs[(a,'ready_to_print')] for a in ok_per_acc):4}  {sum(ok_per_subs[(a,'printed')] for a in ok_per_acc):4}        {sum(ok_per_acc.values()):4}")

if fail:
    print(f"\n=== FALLIDAS: {len(fail)} ===")
    by_acc = Counter(f["acc"] for f in fail)
    for acc,n in by_acc.most_common(): print(f"  {acc}: {n}")
