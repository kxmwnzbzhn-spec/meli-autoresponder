"""
Etiquetas pendientes de entregar (ready_to_ship, todos substatus excepto picked_up).
- Anota cada PDF con composicion REAL (color desde variation_attributes/item.attributes).
- Texto pequeno con wrapping para que entre completo.
- Ventana 7 dias.
"""
import os, requests, re
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

APP_ID = os.environ["MELI_APP_ID"]; APP_SECRET = os.environ["MELI_APP_SECRET"]
TG_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN","")
TG_CHAT = os.environ.get("TELEGRAM_CHAT_ID","")

ACCOUNTS = [
    ("JUAN","MELI_REFRESH_TOKEN"),
    ("CLARIBEL","MELI_REFRESH_TOKEN_CLARIBEL"),
    ("ASVA","MELI_REFRESH_TOKEN_ASVA"),
    ("RAYMUNDO","MELI_REFRESH_TOKEN_RAYMUNDO"),
    ("DILCIE","MELI_REFRESH_TOKEN_DILCIE"),
    ("MILDRED","MELI_REFRESH_TOKEN_MILDRED"),
    ("YC_NEW","MELI_REFRESH_TOKEN_YC_NEW"),
    ("BREN","MELI_REFRESH_TOKEN_BREN"),
    ("WILBERT","MELI_REFRESH_TOKEN_WILBERT"),
]

INCLUDE_STATUSES = {"ready_to_ship"}
EXCLUDE_SUBSTATUS = {"picked_up"}

# Cache items para no pegar a la API muchas veces
ITEM_CACHE = {}

def get_item_full(item_id, headers):
    if item_id in ITEM_CACHE:
        return ITEM_CACHE[item_id]
    try:
        r = requests.get(f"https://api.mercadolibre.com/items/{item_id}", headers=headers, timeout=10).json()
        ITEM_CACHE[item_id] = r
        return r
    except:
        return {}

def extract_color(order_item, item_full):
    """Busca el color en variation_attributes -> attributes -> title (en ese orden)."""
    # 1. variation_attributes del order_item
    for va in order_item.get("item",{}).get("variation_attributes",[]) or []:
        if va.get("id","").upper() in ("COLOR","COLOR_NAME","MAIN_COLOR"):
            v = va.get("value_name") or va.get("value_id") or ""
            if v: return v
    # 2. variations del item completo si hay variation_id
    var_id = (order_item.get("item") or {}).get("variation_id")
    if var_id and item_full.get("variations"):
        for v in item_full["variations"]:
            if v.get("id") == var_id:
                for a in v.get("attribute_combinations",[]) or []:
                    if a.get("id","").upper() in ("COLOR","COLOR_NAME","MAIN_COLOR"):
                        nv = a.get("value_name") or ""
                        if nv: return nv
    # 3. attributes del item completo
    for a in item_full.get("attributes",[]) or []:
        if a.get("id","").upper() in ("COLOR","COLOR_NAME","MAIN_COLOR"):
            v = a.get("value_name") or ""
            if v: return v
    # 4. fallback: buscar en title
    t = (order_item.get("item",{}).get("title","") or "").lower()
    for c in ["azul marino","azul","celeste","aqua","rojo","roja","negro","negra","blanco","blanca","rosa","camuflaje","camo","morado","morada","violeta","verde","amarillo","gris","plateado","dorado","beige","cafe","marron"]:
        if c in t:
            return c.title()
    return ""

def extract_model_short(order_item, item_full):
    """Modelo+tamano corto para mostrar. Para perfumes: marca+nombre+ml. Para bocinas: marca+modelo."""
    t = (order_item.get("item",{}).get("title","") or "")
    tl = t.lower()
    # Bocinas conocidas
    if "go 4" in tl or "go4" in tl: return "JBL Go 4"
    if "go 3" in tl or "go3" in tl: return "JBL Go 3"
    if "go essential" in tl: return "JBL Go Essential"
    if "flip 7" in tl or "flip7" in tl: return "JBL Flip 7"
    if "flip 6" in tl: return "JBL Flip 6"
    if "charge 6" in tl: return "JBL Charge 6"
    if "charge 5" in tl: return "JBL Charge 5"
    if "grip" in tl and "jbl" in tl: return "JBL Grip"
    if "clip 5" in tl: return "JBL Clip 5"
    if "clip 4" in tl: return "JBL Clip 4"
    if "xb100" in tl: return "Sony XB100"
    if "buds" in tl and "redmi" in tl: return "Redmi Buds"
    # Perfumes: extraer marca + nombre principal + ml si esta en titulo
    if "perfume" in tl or "edp" in tl or "edt" in tl or " ml " in tl or "fragancia" in tl:
        # Sacar primeras palabras significativas
        # Eliminar prefijos comunes
        clean = re.sub(r'^(perfume\s+(original\s+)?|original\s+|fragancia\s+)', '', t, flags=re.IGNORECASE)
        # Tomar primeras 5-7 palabras + ml si lo encuentra
        ml_match = re.search(r'(\d+\s*ml)', t, re.IGNORECASE)
        words = clean.split()[:6]
        result = " ".join(words)
        if ml_match and ml_match.group(1).lower() not in result.lower():
            result += f" {ml_match.group(1)}"
        return result[:55]
    # Default: primeras 6 palabras
    return " ".join(t.split()[:6])[:55]

cdmx = datetime.now(timezone.utc) - timedelta(hours=6)
window_start = (cdmx - timedelta(days=7)).replace(hour=0,minute=0,second=0,microsecond=0)
date_from = window_start.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
print(f"Rango: {window_start.strftime('%Y-%m-%d')} CDMX -> ahora (7 dias)")

now_cdmx = cdmx.strftime("%Y-%m-%d_%H%M")
OUTDIR = f"labels_pending_{now_cdmx}"
os.makedirs(OUTDIR, exist_ok=True)

shipments = {}
status_counts = defaultdict(int)

for label, env_var in ACCOUNTS:
    RT = os.environ.get(env_var, "")
    if not RT:
        print(f"[{label}] sin token - skip"); continue
    try:
        r = requests.post("https://api.mercadolibre.com/oauth/token",
            data={"grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":RT},
            timeout=15).json()
        H = {"Authorization":f"Bearer {r['access_token']}"}
        me = requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=10).json()
        USER_ID = me["id"]
    except Exception as e:
        print(f"[{label}] auth err: {e}"); continue
    print(f"[{label}] {me.get('nickname','')}")
    offset = 0; account_count = 0
    while True:
        rr = requests.get(f"https://api.mercadolibre.com/orders/search?seller={USER_ID}&order.date_created.from={date_from}&limit=50&offset={offset}",headers=H,timeout=20).json()
        results = rr.get("results",[])
        if not results: break
        for o in results:
            if o.get("status") not in ("paid","shipped","partially_paid","cancelled"):
                continue
            sh = o.get("shipping",{}) or {}
            sh_id = sh.get("id")
            if not sh_id: continue
            try:
                sd = requests.get(f"https://api.mercadolibre.com/shipments/{sh_id}",headers=H,timeout=10).json()
                ship_status = sd.get("status","")
                substatus = sd.get("substatus","")
            except: continue
            status_counts[f"{ship_status}/{substatus or '-'}"] += 1
            if ship_status not in INCLUDE_STATUSES: continue
            if substatus in EXCLUDE_SUBSTATUS: continue
            buyer = (o.get("buyer") or {}).get("nickname","")
            items_in_order = []
            for oi in o.get("order_items",[]):
                title = (oi.get("item") or {}).get("title","")
                qty = oi.get("quantity",0)
                item_id = (oi.get("item") or {}).get("id","")
                full = get_item_full(item_id, H) if item_id else {}
                color = extract_color(oi, full)
                model = extract_model_short(oi, full)
                items_in_order.append({"model":model,"color":color,"qty":qty,"title":title})
            key = (label, sh_id)
            if key in shipments:
                shipments[key]["items"].extend(items_in_order)
                shipments[key]["orders"].append(o.get("id"))
            else:
                shipments[key] = {
                    "account": label, "shipment_id": sh_id,
                    "orders": [o.get("id")], "items": items_in_order,
                    "buyer": buyer, "ship_status": ship_status,
                    "substatus": substatus, "_token": H["Authorization"],
                }
            account_count += 1
        offset += 50
        if offset >= rr.get("paging",{}).get("total",0): break
    print(f"  -> {account_count} envios pendientes")

print("\n=== Status (top 15) ===")
for s, c in sorted(status_counts.items(), key=lambda x: -x[1])[:15]:
    mark = "OK" if s.split("/")[0] in INCLUDE_STATUSES and s.split("/")[1] not in EXCLUDE_SUBSTATUS else "no"
    print(f"  {mark} {s}: {c}")

def composition_signature(items):
    consol = defaultdict(int)
    for it in items:
        color = it["color"] or "S/Color"
        k = f"{it['model']}_{color}"
        k = re.sub(r'[^A-Za-z0-9_]+','_', k).strip("_")
        consol[k] += it["qty"]
    parts = sorted(f"{k}_x{v}" for k,v in consol.items())
    return "+".join(parts)

def composition_lines(items):
    """Lista de strings: 'JBL Go 4 Azul x1', uno por modelo+color para anotar."""
    consol = defaultdict(int)
    for it in items:
        color = it["color"] or "S/Color"
        k = f"{it['model']} {color}"
        consol[k] += it["qty"]
    return sorted(f"{k} x{v}" for k,v in consol.items())

groups = defaultdict(list)
mixed_shipments = []
for sh in shipments.values():
    sig = composition_signature(sh["items"])
    sh["composition"] = sig
    sh["lines"] = composition_lines(sh["items"])
    is_mixed = len({(it["model"],it["color"] or "S/Color") for it in sh["items"]}) > 1
    if is_mixed:
        sh["composition"] = "MIXTO__" + sig
        mixed_shipments.append(sh)
    groups[sh["composition"]].append(sh)

total_shipments = len(shipments)
total_units = sum(sum(i["qty"] for i in s["items"]) for s in shipments.values())
print(f"\nINCLUIDOS: {total_shipments} envios / {total_units} unidades / {len(groups)} grupos ({len(mixed_shipments)} MIXTOS)\n")

def annotate_pdf(pdf_bytes, shipment_info):
    """Sobrepone caja amarilla arriba con composicion (multiple lineas si hace falta)."""
    from pypdf import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas

    lines = shipment_info.get("lines", [])
    account = shipment_info.get("account","?")
    buyer = shipment_info.get("buyer","")
    sub = shipment_info.get("substatus","")
    ship_id = shipment_info.get("shipment_id","")

    base = PdfReader(BytesIO(pdf_bytes))
    writer = PdfWriter()
    for page in base.pages:
        media = page.mediabox
        w = float(media.width); h = float(media.height)
        # Calcular alto necesario
        n_lines = max(1, len(lines))
        box_h = 28 + (n_lines * 17) + 14  # padding + lineas + meta line
        overlay_buf = BytesIO()
        c = canvas.Canvas(overlay_buf, pagesize=(w, h))
        # Caja amarilla
        c.setFillColorRGB(1, 1, 0.7)
        c.rect(0, h-box_h, w, box_h, fill=1, stroke=0)
        c.setStrokeColorRGB(0,0,0)
        c.setLineWidth(1.5)
        c.rect(0, h-box_h, w, box_h, fill=0, stroke=1)
        # Meta line arriba (account, buyer, ship)
        c.setFillColorRGB(0,0,0)
        c.setFont("Helvetica-Bold", 10)
        meta = f"[{account}] {buyer} | Ship:{ship_id} | {sub}"
        c.drawString(8, h-15, meta[:110])
        # Lineas de composicion
        c.setFont("Helvetica-Bold", 13)
        y = h - 35
        for line in lines:
            # Wrap si > 80 chars
            if len(line) <= 80:
                c.drawString(8, y, line)
                y -= 17
            else:
                # split en 2
                c.drawString(8, y, line[:80])
                y -= 17
                if y > h - box_h + 5:
                    c.setFont("Helvetica-Bold", 11)
                    c.drawString(8, y, line[80:160])
                    c.setFont("Helvetica-Bold", 13)
                    y -= 17
        c.save()
        overlay_buf.seek(0)
        overlay = PdfReader(overlay_buf)
        page.merge_page(overlay.pages[0])
        writer.add_page(page)
    out = BytesIO()
    writer.write(out)
    return out.getvalue()

labels_index = []
for sig, ships in sorted(groups.items()):
    if not ships: continue
    safe_key = re.sub(r'[^A-Za-z0-9_+]+','_', sig).strip("_")[:80]
    is_mixed = sig.startswith("MIXTO__")
    icon = "MIXTO" if is_mixed else "PKG"
    print(f"{icon} {sig} ({len(ships)} envios)")

    annotated_pdfs = []
    for sh in ships:
        try:
            r = requests.get(f"https://api.mercadolibre.com/shipment_labels?shipment_ids={sh['shipment_id']}&response_type=pdf",
                headers={"Authorization": sh["_token"]}, timeout=60)
            if r.status_code == 200 and r.headers.get("content-type","").startswith("application/pdf"):
                try:
                    ann = annotate_pdf(r.content, sh)
                    annotated_pdfs.append(ann)
                except Exception as e:
                    print(f"  annot err {sh['shipment_id']}: {e}")
                    annotated_pdfs.append(r.content)
            else:
                print(f"  ERR shipment {sh['shipment_id']}: HTTP {r.status_code}")
        except Exception as e:
            print(f"  ERR shipment {sh['shipment_id']}: {e}")

    if annotated_pdfs:
        out_pdf = f"{OUTDIR}/{safe_key}.pdf"
        try:
            from pypdf import PdfWriter, PdfReader
            writer = PdfWriter()
            for pdf_bytes in annotated_pdfs:
                reader = PdfReader(BytesIO(pdf_bytes))
                for p in reader.pages: writer.add_page(p)
            with open(out_pdf, "wb") as f: writer.write(f)
            print(f"  OK {len(annotated_pdfs)} envios -> {out_pdf}")
        except Exception as e:
            print(f"  fallback: {e}")
        labels_index.append({
            "sig": sig, "file": f"{safe_key}.pdf", "is_mixed": is_mixed,
            "envios": len(ships), "unidades": sum(sum(i["qty"] for i in s["items"]) for s in ships),
            "ships": ships,
        })

# Manifest XLSX
wb = Workbook(); wb.remove(wb.active)
ws = wb.create_sheet("Resumen")
ws["A1"] = f"Pendientes - {now_cdmx} (ventana 7d)"
ws["A1"].font = Font(bold=True, size=16, color="1F4E78"); ws.merge_cells("A1:F1")
HEADER = ["Composicion","Envios","Unidades","Tipo","Archivo","Cuentas"]
for i, h in enumerate(HEADER, 1):
    c = ws.cell(row=3, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E78")
RED = PatternFill("solid", fgColor="FFC7CE")
r = 4
for grp in sorted(labels_index, key=lambda x: -x["envios"]):
    cuentas = sorted({s["account"] for s in grp["ships"]})
    pretty = grp["sig"].replace("MIXTO__","").replace("_x"," x").replace("+"," + ").replace("_"," ")
    ws.cell(row=r, column=1, value=pretty)
    ws.cell(row=r, column=2, value=grp["envios"])
    ws.cell(row=r, column=3, value=grp["unidades"])
    tcell = ws.cell(row=r, column=4, value="MIXTO" if grp["is_mixed"] else "Solo")
    if grp["is_mixed"]: tcell.fill = RED
    ws.cell(row=r, column=5, value=grp["file"])
    ws.cell(row=r, column=6, value=", ".join(cuentas))
    r += 1

ws2 = wb.create_sheet("Detalle envios")
HEADER2 = ["Cuenta","Order ID","Shipment","Composicion","# items","Comprador","Status","Substatus","Detalle items"]
for i, h in enumerate(HEADER2, 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E78")
r = 2
for grp in labels_index:
    for s in grp["ships"]:
        items_str = "; ".join(f"{i['model']} {i['color'] or 'S/Color'} x{i['qty']}" for i in s["items"])
        is_mixed = len({(i["model"],i["color"] or "S/Color") for i in s["items"]}) > 1
        ws2.cell(row=r, column=1, value=s["account"])
        ws2.cell(row=r, column=2, value=", ".join(str(o) for o in s["orders"]))
        ws2.cell(row=r, column=3, value=str(s["shipment_id"]))
        ws2.cell(row=r, column=4, value=grp["sig"].replace("MIXTO__",""))
        ws2.cell(row=r, column=5, value=len(s["items"]))
        ws2.cell(row=r, column=6, value=s["buyer"])
        ws2.cell(row=r, column=7, value=s["ship_status"])
        ws2.cell(row=r, column=8, value=s.get("substatus",""))
        ws2.cell(row=r, column=9, value=items_str)
        if is_mixed:
            for c_idx in range(1, 10): ws2.cell(row=r, column=c_idx).fill = RED
        r += 1

for ws_x in [ws, ws2]:
    for ci in range(1, ws_x.max_column+1):
        max_len = 10
        for ri in range(1, ws_x.max_row+1):
            try:
                v = ws_x.cell(row=ri, column=ci).value
                if v: max_len = max(max_len, len(str(v)))
            except: pass
        ws_x.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 80)

wb.save(f"{OUTDIR}/manifest.xlsx")
print(f"\n{len(labels_index)} grupos en {OUTDIR}/")

if TG_TOKEN and TG_CHAT and total_shipments > 0:
    text = f"Etiquetas - {now_cdmx}\n{total_shipments} envios / {total_units} unid / {len(labels_index)} grupos"
    if mixed_shipments:
        text += f"\n{len(mixed_shipments)} envios MIXTOS"
    text += "\n\nTop:\n"
    for grp in sorted(labels_index, key=lambda x: -x["envios"])[:8]:
        pretty = grp["sig"].replace("MIXTO__","MIXTO ").replace("_x"," x").replace("+"," + ").replace("_"," ")
        text += f"- {pretty}: {grp['envios']}\n"
    requests.post(f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage",
        json={"chat_id":TG_CHAT,"text":text})
