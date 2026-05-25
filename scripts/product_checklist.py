"""CHECK LIST: nombre original MELI (título + color real del variant) vs nombre corto de etiqueta.
Recorre listings activos de cada cuenta. XLSX con columna de alerta cuando el color no coincide.
"""
import os, requests, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Cuentas con su app (vieja=5211907102822632, nueva=2008666770714005)
OLD_ID="5211907102822632"; OLD_SECRET=os.environ.get("MELI_APP_SECRET")
NEW_ID=os.environ.get("MELI_APP_ID_NEW","2008666770714005"); NEW_SECRET=os.environ.get("MELI_APP_SECRET_NEW")
ACCOUNTS=[
    ("Wilbert", os.environ.get("MELI_REFRESH_TOKEN_WILBERT"), OLD_ID, OLD_SECRET),
    ("Asva",    os.environ.get("MELI_REFRESH_TOKEN_ASVA"),    OLD_ID, OLD_SECRET),
    ("Yiriam",  os.environ.get("MELI_REFRESH_TOKEN_YC_NEW"),  NEW_ID, NEW_SECRET),
]

def tok(rt, cid, csec):
    if not rt or not csec: return None
    r=requests.post("https://api.mercadolibre.com/oauth/token",data={
        "grant_type":"refresh_token","client_id":cid,"client_secret":csec,"refresh_token":rt}).json()
    return r.get("access_token")

def color_map(text):
    if not text: return None
    tl=" "+text.lower()+" "
    cm=[("camuflaj","Camuflaje"),("camo","Camuflaje"),("azul marino","Azul Marino"),
        ("aqua","Aqua"),("celeste","Celeste"),("negr","Negro"),(" black","Negro"),
        ("roj","Rojo"),(" red","Rojo"),("rosa","Rosa"),("pink","Rosa"),
        ("morad","Morado"),("violeta","Morado"),("purple","Morado"),
        (" azul","Azul"),(" blue","Azul"),("blanco","Blanco"),("white","Blanco"),
        ("verde","Verde"),("green","Verde"),("amarillo","Amarillo"),("yellow","Amarillo"),
        ("naranja","Naranja"),("orange","Naranja"),("gris","Gris"),("gray","Gris"),("grey","Gris"),
        ("plateado","Plata"),("silver","Plata"),("dorado","Dorado"),("gold","Dorado")]
    for k,v in cm:
        if k in tl: return v
    return None

def my_color(value_name):
    """Lo que YO pondría en la etiqueta a partir del value_name del variant."""
    c = color_map(value_name)
    if c: return c
    return (value_name or "").strip().title() or None

def get_model(title):
    t=(title or "")
    tl=t.lower()
    if "go 4" in tl or "go4" in tl: return "Go4"
    if "go 3" in tl or "go3" in tl: return "Go3"
    if "clip 5" in tl or "clip5" in tl: return "Clip5"
    if "charge 6" in tl or "charge6" in tl: return "Charge6"
    if "flip 7" in tl or "flip7" in tl: return "Flip7"
    if "grip" in tl: return "Grip"
    if "xb100" in tl: return "XB100"
    if "soundlink" in tl: return "SoundLink"
    if "modelo padrão" in tl or "modelo padrao" in tl: return "JBL Impermeable"
    for w in ["Bocina ","bocina ","JBL ","jbl ","Sony ","Bose "]:
        t=t.replace(w,"")
    return t.strip()[:24]

rows=[]
for acc, rt, cid, csec in ACCOUNTS:
    at=tok(rt,cid,csec)
    if not at:
        print(f"--- {acc}: token fail"); continue
    H={"Authorization":f"Bearer {at}"}
    me=requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=15).json()
    uid=me.get("id")
    print(f"--- {acc} uid={uid} ---")
    ids=[]; offset=0
    while True:
        r=requests.get(f"https://api.mercadolibre.com/users/{uid}/items/search",
            headers=H, params={"status":"active","limit":100,"offset":offset}, timeout=20).json()
        chunk=r.get("results",[])
        if not chunk: break
        ids.extend(chunk); offset+=len(chunk)
        if offset>=r.get("paging",{}).get("total",0): break
    print(f"  active items: {len(ids)}")
    for bs in range(0,len(ids),20):
        batch=ids[bs:bs+20]
        try:
            mg=requests.get(f"https://api.mercadolibre.com/items?ids={','.join(batch)}",headers=H,timeout=20).json()
        except: continue
        for entry in mg:
            if entry.get("code")!=200: continue
            it=entry.get("body") or {}
            iid=it.get("id"); title=it.get("title","")
            model=get_model(title)
            variations=it.get("variations") or []
            if not variations:
                meli_color = color_map(title) or "(sin variante)"
                etiqueta = f"{model} {my_color(title) or ''}".strip()
                rows.append([acc, iid, title, meli_color, etiqueta, ""])
            else:
                for v in variations:
                    vc=""
                    for ac in (v.get("attribute_combinations") or []):
                        if ac.get("id")=="COLOR" or "color" in (ac.get("name","") or "").lower():
                            vc=ac.get("value_name") or ""
                            break
                    etiqueta=f"{model} {my_color(vc) or ''}".strip()
                    # flag si lo que pondría difiere notablemente del value_name MELI
                    alerta = ""
                    if vc and my_color(vc) and vc.strip().lower() != (my_color(vc) or '').lower() and color_map(vc) is None:
                        alerta = "REVISAR"
                    rows.append([acc, f"{iid}/{v.get('id')}", title, vc or "(s/color)", etiqueta, alerta])
        time.sleep(0.1)

print(f"\nTotal filas: {len(rows)}")

# XLSX
wb=Workbook(); ws=wb.active; ws.title="Checklist productos"
hf=PatternFill("solid",fgColor="2C3E50"); hF=Font(bold=True,color="FFFFFF",size=11)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
border=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
heads=["Cuenta","Listing/Variante","NOMBRE ORIGINAL MELI","Color real (MELI)","NOMBRE EN ETIQUETA","Alerta"]
for c,h in enumerate(heads,1):
    cell=ws.cell(row=1,column=c,value=h); cell.fill=hf; cell.font=hF; cell.alignment=center; cell.border=border
rows.sort(key=lambda r:(r[0], r[4]))
red=PatternFill("solid",fgColor="E74C3C")
for i,r in enumerate(rows,2):
    for c,val in enumerate(r,1):
        cell=ws.cell(row=i,column=c,value=val); cell.border=border
        cell.alignment=Alignment(vertical="center",wrap_text=True)
        if c==5: cell.font=Font(bold=True)
        if c==6 and val=="REVISAR":
            for cc in range(1,7): ws.cell(row=i,column=cc).fill=red; ws.cell(row=i,column=cc).font=Font(bold=True,color="FFFFFF")
widths={1:10,2:24,3:55,4:18,5:26,6:10}
for c,w in widths.items(): ws.column_dimensions[chr(64+c)].width=w
ws.freeze_panes="A2"
for rr in range(2,len(rows)+2): ws.row_dimensions[rr].height=26
wb.save("CHECKLIST_PRODUCTOS.xlsx")
print("✅ CHECKLIST_PRODUCTOS.xlsx")
