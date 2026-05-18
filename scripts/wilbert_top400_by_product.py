"""WILBERT mañana + a partir del 14 may = etiquetas futuras (no las de hoy).
Filtro: status=ready_to_ship, substatus IN (printed, ready_to_print), deadline > hoy 11 AM CDMX.
Header: USADO en banda roja si aplica. 4x6 in.
"""
import os, io, time, requests
from datetime import datetime, timedelta, timezone
from collections import Counter
from pypdf import PdfReader, PdfWriter, PageObject, Transformation
from reportlab.pdfgen import canvas
from reportlab.lib.colors import Color
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pdf2image import convert_from_bytes
from PIL import ImageOps

APP_ID="5211907102822632"
APP_SECRET=os.environ["MELI_APP_SECRET"]
RT = os.environ["MELI_REFRESH_TOKEN_WILBERT"]
TZ = timezone(timedelta(hours=-6))
PAGE_W=4*72; PAGE_H=6*72

ALLOWED_SUBS = {"printed", "ready_to_print"}
TOP_N = 400
# Listings que el usuario quiere marcar manualmente como USADO (override condition)
USED_LISTINGS = {"MLM2911205487", "MLM5295749840", "MLM2911241939"}


def tok(rt):
    r=requests.post("https://api.mercadolibre.com/oauth/token",data={
        "grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":rt}).json()
    return r.get("access_token")


def _parse_color(text):
    if not text: return None
    tl=" "+text.lower()+" "
    col_map=[("camuflaj","Camuflaje"),("camo","Camuflaje"),("azul marino","Azul Marino"),
             ("aqua","Aqua"),("celeste","Celeste"),("negr","Negro"),(" black","Negro"),
             ("roj","Rojo"),(" red","Rojo"),("rosa","Rosa"),("pink","Rosa"),
             ("morad","Morado"),("violeta","Morado"),("purple","Morado"),
             (" azul","Azul"),(" blue","Azul"),("blanco","Blanco"),("white","Blanco"),
             ("verde","Verde"),("green","Verde"),("amarillo","Amarillo"),("yellow","Amarillo"),
             ("naranja","Naranja"),("orange","Naranja"),("gris","Gris"),("gray","Gris"),
             ("plateado","Plata"),("silver","Plata"),("dorado","Dorado"),("gold","Dorado")]
    for k,v in col_map:
        if k in tl: return v
    return None


def get_variant_color(item_obj, H):
    attrs = item_obj.get("variation_attributes") or []
    for a in attrs:
        if a.get("id")=="COLOR" or "color" in (a.get("name","") or "").lower():
            c = _parse_color(a.get("value_name") or "")
            if c: return c
    iid = item_obj.get("id"); vid = item_obj.get("variation_id")
    if iid and vid:
        try:
            r = requests.get(f"https://api.mercadolibre.com/items/{iid}/variations/{vid}",
                             headers=H, timeout=8)
            if r.status_code == 200:
                for ac in (r.json().get("attribute_combinations") or []):
                    if ac.get("id")=="COLOR" or "color" in (ac.get("name","") or "").lower():
                        return _parse_color(ac.get("value_name"))
        except: pass
    return None


def clean_title(title, item_obj=None, H=None):
    t=(title or "").strip()
    for w in ["Bocina ","bocina ","Parlante ","parlante ","Altavoz ","altavoz ","Speaker ","speaker ",
              "JBL ","jbl ","Jbl ","Sony ","SONY ","Bose ","BOSE "]:
        t=t.replace(w,"")
    tl=t.lower()
    if "go 4" in tl or "go4" in tl: model="Go 4"
    elif "go 3" in tl or "go3" in tl: model="Go 3"
    elif "clip 5" in tl or "clip5" in tl: model="Clip 5"
    elif "charge 6" in tl or "charge6" in tl: model="Charge 6"
    elif "flip 7" in tl or "flip7" in tl: model="Flip 7"
    elif "grip" in tl: model="Grip"
    elif "xb100" in tl: model="Sony XB100"
    elif "soundlink" in tl: model="Bose SoundLink"
    else: model=t[:30]
    color = None
    if item_obj is not None:
        color = get_variant_color(item_obj, H)
    if not color:
        color = _parse_color(t)
    return f"{model} {color}" if color else model


_COND_CACHE={}
def get_condition(item_obj, H):
    cond = item_obj.get("condition")
    if cond: return cond
    iid = item_obj.get("id")
    if not iid: return None
    if iid in _COND_CACHE: return _COND_CACHE[iid]
    try:
        r = requests.get(f"https://api.mercadolibre.com/items/{iid}",
                         headers=H, timeout=8, params={"attributes":"condition"})
        if r.status_code == 200:
            cond = (r.json() or {}).get("condition")
            _COND_CACHE[iid]=cond; return cond
    except: pass
    _COND_CACHE[iid]=None; return None


def detect_content_bbox(pdf_bytes, page_idx=0):
    try:
        imgs = convert_from_bytes(pdf_bytes, dpi=72, first_page=page_idx+1, last_page=page_idx+1)
        if not imgs: return None
        img = imgs[0].convert("L")
        bw = img.point(lambda p: 0 if p < 245 else 255, mode="L")
        inv = ImageOps.invert(bw)
        bbox = inv.getbbox()
        if not bbox: return None
        x0, y0_top, x1, y1_top = bbox
        img_w, img_h = img.size
        pdf_y0 = img_h - y1_top
        pdf_y1 = img_h - y0_top
        m = 2
        return (max(0, x0-m), max(0, pdf_y0-m),
                min(img_w, x1-x0+2*m), min(img_h, pdf_y1-pdf_y0+2*m))
    except: return None


def render_header(s, header_h):
    has_used = bool(s.get("has_used"))
    usado_strip = 14 if has_used else 0
    total_h = header_h + usado_strip
    buf=io.BytesIO()
    c=canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    cx = PAGE_W/2.0
    c.setFillColor(Color(1, 0.96, 0.74))
    c.rect(0, PAGE_H-total_h, PAGE_W, header_h, fill=1, stroke=0)
    if has_used:
        c.setFillColorRGB(0.85, 0.13, 0.13)
        c.rect(0, PAGE_H-usado_strip, PAGE_W, usado_strip, fill=1, stroke=0)
        c.setFillColorRGB(1,1,1)
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(cx, PAGE_H-11, "*** PRODUCTO USADO ***")
    yellow_top = PAGE_H - usado_strip
    c.setFillColorRGB(0,0,0)
    c.setFont("Helvetica-Bold", 7.5)
    tag = s.get("substatus","")[:8]
    c.drawCentredString(cx, yellow_top-11, f"[WILBERT/{tag}] {s['buyer'][:22]} | Ship:{s['sid']}")
    big = s["comp_lines"][:4]; n=len(big); lh=16
    block_top=yellow_top-18; block_bot=PAGE_H-total_h+4
    block_h=block_top-block_bot; text_h=n*lh
    first_y = block_top - (block_h - text_h)/2.0 - 12
    c.setFont("Helvetica-Bold", 14)
    y=first_y
    for line in big:
        c.drawCentredString(cx, y, line[:30]); y-=lh
    c.showPage(); c.save()
    buf.seek(0)
    return PdfReader(buf).pages[0]


# === FASE 1: shipments Wilbert ===
NOW = datetime.now(timezone.utc)
START = NOW - timedelta(days=60)
print(f"=== Wilbert demoradas+listas: ventana {START.date()} a {NOW.date()} ===")
at = tok(RT)
if not at: raise SystemExit("No pude obtener access_token de Wilbert")
H = {"Authorization": f"Bearer {at}"}
me = requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=15).json()
uid = me.get("id")
print(f"uid={uid}")
orders=[]; offset=0
while True:
    r=requests.get("https://api.mercadolibre.com/orders/search",headers=H,timeout=20,
        params={"seller":uid,"order.status":"paid",
                "order.date_created.from":START.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                "order.date_created.to":NOW.strftime("%Y-%m-%dT%H:%M:%S.000Z"),
                "limit":50,"offset":offset}).json()
    res=r.get("results",[])
    if not res: break
    orders.extend(res); offset+=len(res)
    if offset>=r.get("paging",{}).get("total",0): break
obs={}
for o in orders:
    sid=(o.get("shipping") or {}).get("id")
    if sid: obs[sid]=o
print(f"shipping ids: {len(obs)}")

shipments=[]
sub_count=Counter()
mode_count=Counter()  # (shipping_mode, logistic_type)
tags_count=Counter()
for sid, ord_o in obs.items():
    try:
        sh=requests.get(f"https://api.mercadolibre.com/shipments/{sid}",headers=H,timeout=10).json()
        st=sh.get("status"); sub=sh.get("substatus")
        sub_count[(st,sub or "(none)")] += 1
        if st != "ready_to_ship" or sub not in ALLOWED_SUBS: continue
        mode_count[(sh.get("shipping_mode") or "?", sh.get("logistic_type") or "?")] += 1
        for t in (sh.get("tags") or []): tags_count[t] += 1
        # *** HANDLING LIMIT (lo que usa el panel MELI) ***
        deadline = None
        # Path 1: lead_time.estimated_handling_limit.date
        lt = sh.get("lead_time") or {}
        ehl = lt.get("estimated_handling_limit") or {}
        if isinstance(ehl, dict):
            ed = ehl.get("date")
        else:
            ed = ehl  # string
        if ed:
            try: deadline = datetime.fromisoformat(ed.replace("Z","+00:00")).astimezone(TZ)
            except: pass
        # Path 2: date_handling.estimated_handling_limit
        if not deadline:
            dh = sh.get("date_handling") or {}
            ehl2 = dh.get("estimated_handling_limit")
            if isinstance(ehl2, dict):
                ed2 = ehl2.get("date")
            else:
                ed2 = ehl2
            if ed2:
                try: deadline = datetime.fromisoformat(ed2.replace("Z","+00:00")).astimezone(TZ)
                except: pass
        if not deadline:
            # Fallback
            hist=sh.get("status_history") or {}
            dh=hist.get("date_handling")
            if dh: deadline=(datetime.fromisoformat(dh.replace("Z","+00:00"))+timedelta(hours=48)).astimezone(TZ)
        # SIN filtro de fecha: incluimos TODO printed (incluso atrasadas de ayer).
        # El sort por handling_limit ASC pone las más urgentes primero.
        if not deadline: continue
        items=ord_o.get("order_items",[])
        comp_lines=[]; has_used=False
        for it in items:
            io_obj = it.get("item") or {}
            title = clean_title(io_obj.get("title",""), io_obj, H)
            qty = it.get("quantity",1)
            iid = io_obj.get("id") or ""
            cond = get_condition(io_obj, H)
            # Override por listing_id explícito
            if iid in USED_LISTINGS or cond == "used":
                has_used=True
                comp_lines.append(f"USADO {title} x{qty}")
            else:
                comp_lines.append(f"{title} x{qty}")
        buyer=(ord_o.get("buyer") or {}).get("nickname","?")
        shipments.append({"sid":sid,"buyer":buyer,"comp_lines":comp_lines,
                          "has_used":has_used,
                          "substatus":sub,
                          "deadline":deadline.strftime("%a %d %H:%M") if deadline else "s/d",
                          "_dl_iso": deadline.isoformat() if deadline else "9999",
                          "tracking":sh.get("tracking_number","")})
        time.sleep(0.04)
    except Exception as e:
        print(f"  err {sid}: {str(e)[:80]}")

print(f"\n=== Distribución (shipping_mode, logistic_type) entre los printed ===")
for k,n in mode_count.most_common(): print(f"  {n:4} mode={k[0]!r:10} logistic={k[1]!r}")
print(f"\n=== Tags más comunes entre los printed ===")
for t,n in tags_count.most_common(20): print(f"  {n:4} {t}")
print(f"\n=== Distribución TODOS status/substatus de Wilbert ===")
for (st,sub),n in sub_count.most_common(): print(f"  {n:4} {st}/{sub}")
seleccionados = Counter(s["substatus"] for s in shipments)
print(f"\n*** Seleccionadas (status=ready_to_ship + substatus en {ALLOWED_SUBS}): {len(shipments)} ***")
for sub,n in seleccionados.most_common(): print(f"   {sub}: {n}")
# Sort por handling_limit ASC (atrasadas primero) → toma TOP_N
def _dl_key(s): return s.get("_dl_iso","9999")
shipments.sort(key=lambda s: (_dl_key(s), s["sid"]))
print(f"\n*** ANTES de cortar: {len(shipments)} | Tomando TOP {TOP_N} ***")
shipments = shipments[:TOP_N]
print(f"*** Después de cortar: {len(shipments)} ***")
# Re-ordena POR PRODUCTO (agrupado), dentro de cada producto por handling_limit
shipments.sort(key=lambda s: ("/".join(s["comp_lines"]), _dl_key(s), s["sid"]))


# === FASE 2: PDF ===
print("\n=== Generando PDF 4x6 ===")
writer = PdfWriter()
fail = []
for s in shipments:
    try:
        r=requests.get("https://api.mercadolibre.com/shipment_labels",
                      headers=H, params={"shipment_ids":s["sid"],"response_type":"pdf"},
                      timeout=30)
        if r.status_code!=200 or not r.headers.get("content-type","").lower().startswith("application/pdf"):
            fail.append({"sid":s["sid"],"reason":r.text[:200]}); continue
        raw=r.content
        lbl_pdf=PdfReader(io.BytesIO(raw))
        for pidx,label_page in enumerate(lbl_pdf.pages):
            box = label_page.cropbox if label_page.cropbox else label_page.mediabox
            lbl_x0=float(box.left); lbl_y0=float(box.bottom)
            lbl_w=float(box.width); lbl_h=float(box.height)
            cb = detect_content_bbox(raw, pidx)
            if cb:
                cx0,cy0,cw,ch = cb
                lbl_x0=float(box.left)+float(cx0)
                lbl_y0=float(box.bottom)+float(cy0)
                lbl_w=float(cw); lbl_h=float(ch)
            n_lines=min(len(s["comp_lines"]),4)
            header_h = 22 + n_lines*16
            label_area_h = PAGE_H - header_h
            new_page = PageObject.create_blank_page(width=PAGE_W, height=PAGE_H)
            sx = PAGE_W/lbl_w; sy = label_area_h/lbl_h
            op = (Transformation().translate(-lbl_x0,-lbl_y0).scale(sx,sy))
            new_page.merge_transformed_page(label_page, op)
            new_page.merge_page(render_header(s, header_h))
            writer.add_page(new_page)
    except Exception as e:
        fail.append({"sid":s["sid"],"reason":f"{type(e).__name__}: {str(e)[:200]}"})
    time.sleep(0.08)

with open("ETIQUETAS_WILBERT_TOP400_BY_PRODUCT.pdf","wb") as f: writer.write(f)
print(f"\n✅ PDF: ETIQUETAS_WILBERT_TOP400_BY_PRODUCT.pdf ({len(writer.pages)} págs) | Fallidas: {len(fail)}")

# === PICKING LIST XLSX ===
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb = Workbook(); ws = wb.active; ws.title = "Picking List"
hf = PatternFill("solid", fgColor="2C3E50")
hF = Font(bold=True, color="FFFFFF", size=12)
red = PatternFill("solid", fgColor="E74C3C")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
headers = ["#", "Producto", "Cantidad", "Cond"]
for col,h in enumerate(headers,1):
    c=ws.cell(row=1, column=col, value=h)
    c.fill=hf; c.font=hF; c.alignment=center; c.border=border

# Cuenta productos (considerando qty por shipment)
from collections import Counter as _Counter
prod_qty = _Counter()
prod_used = {}
for s in shipments:
    for line in s["comp_lines"]:
        # line format: "USADO Go 4 Negro x1" o "Go 4 Negro x1"
        is_used = line.startswith("USADO ")
        # extrae cantidad del sufijo " xN"
        try:
            qty = int(line.rsplit("x",1)[-1].strip())
        except: qty = 1
        # nombre sin "x{n}"
        name = line.rsplit("x",1)[0].strip()
        if is_used and not name.startswith("USADO"):
            name = "USADO " + name
        prod_qty[name] += qty
        if is_used: prod_used[name] = True

# Ordena: USADOS primero, luego por cantidad desc
sorted_items = sorted(prod_qty.items(), key=lambda x: (0 if prod_used.get(x[0]) else 1, -x[1]))
for i,(prod,qty) in enumerate(sorted_items,1):
    is_used = prod_used.get(prod, False)
    fill = red if is_used else PatternFill("solid", fgColor="FFFFFF")
    cond = "USADO" if is_used else "Nuevo"
    row=[i, prod, qty, cond]
    for col,val in enumerate(row,1):
        c=ws.cell(row=i+1, column=col, value=val)
        c.border=border; c.alignment=Alignment(vertical="center", wrap_text=True)
        if col==2: c.font=Font(bold=True, size=11)
        if col==3: c.font=Font(bold=True, size=12); c.alignment=center
        if is_used:
            c.fill=red; c.font=Font(bold=True, color="FFFFFF", size=11)

# Total
total_row = len(sorted_items)+2
c=ws.cell(row=total_row, column=2, value="TOTAL"); c.font=Font(bold=True, size=12); c.fill=hf; c.font=hF; c.alignment=Alignment(horizontal="right")
c=ws.cell(row=total_row, column=3, value=sum(prod_qty.values())); c.font=Font(bold=True, size=12); c.fill=hf; c.font=hF; c.alignment=center

widths={1:5, 2:45, 3:12, 4:10}
for col,w in widths.items(): ws.column_dimensions[chr(64+col)].width=w
for r in range(2, len(sorted_items)+2): ws.row_dimensions[r].height=22
ws.freeze_panes="A2"
wb.save("PICKING_LIST_WILBERT_400.xlsx")
print(f"✅ Picking list: PICKING_LIST_WILBERT_400.xlsx ({len(sorted_items)} productos, {sum(prod_qty.values())} unidades)")
print("\n=== Distribución por día deadline ===")
day_count = Counter(s["deadline"][:6] for s in shipments)
for d,n in sorted(day_count.items()): print(f"  {n:4} {d}")
print("\n=== Distribución por substatus ===")
sub_in = Counter(s["substatus"] for s in shipments)
for s,n in sub_in.most_common(): print(f"  {n:4} {s}")
used_n = sum(1 for s in shipments if s["has_used"])
print(f"\nUSADOS: {used_n}")
for f_ in fail: print(f"   FALLIDA sid={f_['sid']} → {f_['reason'][:140]}")

# Resumen productos
print("\n=== Top productos ===")
prods = Counter("/".join(s["comp_lines"]) for s in shipments)
for p,n in sorted(prods.items(), key=lambda x:-x[1])[:15]: print(f"  {n:3} {p}")
