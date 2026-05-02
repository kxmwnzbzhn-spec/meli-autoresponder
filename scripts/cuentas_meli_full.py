"""Pull live de todas las cuentas MELI con refresh token activo."""
import os, requests, json
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ACCOUNTS = [
    ("JUAN", "MELI_REFRESH_TOKEN", "Juan Moreno Pedraza", "ventas.meli001@gmail.com"),
    ("CLARIBEL", "MELI_REFRESH_TOKEN_CLARIBEL", "Claribel Perez", "ventas.meli011@gmail.com"),
    ("ASVA", "MELI_REFRESH_TOKEN_ASVA", "Asva Electronics", "asvaelectronics@gmail.com"),
    ("RAYMUNDO", "MELI_REFRESH_TOKEN_RAYMUNDO", "Raymundo Santa Cruz", "enviamesantacruz1@gmail.com"),
    ("DILCIE", "MELI_REFRESH_TOKEN_DILCIE", "Dilcie Guadalupe Medina Chale", "ventas.meli014@gmail.com"),
    ("MILDRED", "MELI_REFRESH_TOKEN_MILDRED", "Mildred Tec", "ventas.meli015@gmail.com"),
    ("YC_NEW", "MELI_REFRESH_TOKEN_YC_NEW", "Yriam Cordoba", "ventas.meli016@gmail.com"),
    ("BREN", "MELI_REFRESH_TOKEN_BREN", "Bren Castillo", "ventas.meli013@gmail.com"),
    ("WILBERT", "MELI_REFRESH_TOKEN_WILBERT", "Wilbert De Jesus Torres", "ventas.meli017@gmail.com"),
    ("ASGARI", "MELI_REFRESH_TOKEN_ASGARI", "Asgari De Jesus", "(pendiente OAuth)"),
    ("ANGEL", "MELI_REFRESH_TOKEN_ANGEL", "Angel Nahuat", "(pendiente OAuth)"),
]

now_utc = datetime.now(timezone.utc)
since_24h = (now_utc - timedelta(days=1)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
since_30d = (now_utc - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%S.000Z")

results = []
for label, env, real_name, email in ACCOUNTS:
    rt = os.environ.get(env, "")
    row = {
        "label": label, "real_name": real_name, "email": email, "env": env,
        "nickname": "", "user_id": "", "status": "❌ sin token",
        "items_active": "—", "items_paused": "—", "items_closed": "—",
        "ventas_24h": "—", "ventas_30d": "—",
        "reputacion": "—", "claims_open": "—",
    }
    if not rt:
        results.append(row); print(f"[{label}] sin token"); continue
    try:
        r = requests.post("https://api.mercadolibre.com/oauth/token", data={
            "grant_type":"refresh_token","client_id":os.environ["MELI_APP_ID"],
            "client_secret":os.environ["MELI_APP_SECRET"],"refresh_token":rt
        }, timeout=15).json()
        tok = r.get("access_token")
        if not tok:
            row["status"] = f"❌ oauth fail: {r.get('message','')[:50]}"
            results.append(row); print(f"[{label}] oauth fail"); continue
        H = {"Authorization": f"Bearer {tok}"}
        me = requests.get("https://api.mercadolibre.com/users/me", headers=H, timeout=10).json()
        row["nickname"] = me.get("nickname","")
        row["user_id"] = me.get("id","")
        row["status"] = "✅ activa"
        # Reputación
        sr = me.get("seller_reputation", {}) or {}
        lvl = sr.get("level_id","")
        pwr = sr.get("power_seller_status","")
        row["reputacion"] = f"{lvl}/{pwr}" if lvl or pwr else "s/d"
        # Items por estado
        for st, key in [("active","items_active"),("paused","items_paused"),("closed","items_closed")]:
            d = requests.get(f"https://api.mercadolibre.com/users/{me['id']}/items/search?status={st}&limit=1", headers=H, timeout=10).json()
            row[key] = d.get("paging",{}).get("total", 0)
        # Ventas 24h y 30d
        for since, key in [(since_24h,"ventas_24h"),(since_30d,"ventas_30d")]:
            tot_orders = 0; tot_revenue = 0; off = 0
            while True:
                j = requests.get(f"https://api.mercadolibre.com/orders/search?seller={me['id']}&order.date_created.from={since}&limit=50&offset={off}", headers=H, timeout=15).json()
                res = j.get("results", [])
                if not res: break
                for o in res:
                    if o.get("status") in ("paid","shipped","delivered"):
                        tot_orders += 1
                        tot_revenue += float(o.get("total_amount",0) or 0)
                if len(res) < 50: break
                off += 50
                if off > 2000: break
            row[key] = f"{tot_orders} ord / ${tot_revenue:,.0f}"
        # Claims abiertos
        try:
            c = requests.get(f"https://api.mercadolibre.com/post-purchase/v1/claims/search?stage=claim&status=opened&limit=50", headers=H, timeout=10).json()
            row["claims_open"] = len(c.get("data", []) or [])
        except: row["claims_open"] = "?"
        results.append(row)
        print(f"[{label}] OK nick={row['nickname']} active={row['items_active']} 24h={row['ventas_24h']}")
    except Exception as e:
        row["status"] = f"❌ err: {str(e)[:50]}"
        results.append(row); print(f"[{label}] err: {e}")

# Build XLSX
wb = Workbook()
ws = wb.active
ws.title = "Cuentas MELI"
header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
header_fill = PatternFill('solid', start_color='1F4E78')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
arial = Font(name='Arial', size=10)
arial_b = Font(name='Arial', size=10, bold=True)

ws['A1'] = f'CUENTAS MELI - Snapshot {now_utc.strftime("%Y-%m-%d %H:%M UTC")}'
ws['A1'].font = Font(bold=True, size=14, name='Arial')
ws.merge_cells('A1:N1')

headers = ["Cuenta","Real Name","Email","Nickname MELI","User ID","Status",
           "Reputación","Items Activos","Pausados","Cerrados","Claims Abiertos",
           "Ventas 24h","Ventas 30d","Secret GH"]
for col, h in enumerate(headers, 1):
    c = ws.cell(3, col, h)
    c.font = header_font; c.fill = header_fill
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border

for r, row in enumerate(results, 4):
    cells = [row["label"], row["real_name"], row["email"], row["nickname"],
             str(row["user_id"]), row["status"], row["reputacion"],
             row["items_active"], row["items_paused"], row["items_closed"],
             row["claims_open"], row["ventas_24h"], row["ventas_30d"], row["env"]]
    for col, v in enumerate(cells, 1):
        cell = ws.cell(r, col, v)
        cell.font = arial_b if col == 1 else arial
        cell.border = border
        if "❌" in str(row["status"]) and col >= 6:
            cell.fill = PatternFill('solid', start_color='FFE5E5')

widths = [11,28,28,22,12,18,16,11,11,11,15,22,22,32]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[3].height = 22
ws.freeze_panes = 'C4'

wb.save("/tmp/cuentas_meli_live.xlsx")
print(f"\n✅ Saved: /tmp/cuentas_meli_live.xlsx ({len(results)} cuentas)")
