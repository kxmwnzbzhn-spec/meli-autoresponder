#!/usr/bin/env python3
"""Auditar manifest.xlsx en cada carpeta fecha. Listar columnas y primeras filas
para entender el contenido y detectar shipments duplicados entre dias."""
import os, io, sys, re
from collections import defaultdict
from google.oauth2.credentials import Credentials as UC
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

cid = os.environ["GOOGLE_OAUTH_CLIENT_ID"]
csec = os.environ["GOOGLE_OAUTH_CLIENT_SECRET"]
rt = os.environ["GOOGLE_OAUTH_REFRESH_TOKEN"]
creds = UC(token=None, refresh_token=rt,
    token_uri="https://oauth2.googleapis.com/token",
    client_id=cid, client_secret=csec,
    scopes=["https://www.googleapis.com/auth/drive"])
svc = build("drive","v3",credentials=creds,cache_discovery=False)


def find_folder(name, parent=None):
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent: q += f" and '{parent}' in parents"
    r = svc.files().list(q=q, fields="files(id,name)").execute()
    return (r.get("files",[]) or [{}])[0].get("id")


def list_all(parent_id):
    out, page = [], None
    q = f"'{parent_id}' in parents and trashed=false"
    while True:
        r = svc.files().list(q=q, fields="nextPageToken,files(id,name,mimeType,modifiedTime)",
                             pageSize=1000, pageToken=page).execute()
        out.extend(r.get("files",[]))
        page = r.get("nextPageToken")
        if not page: break
    return out


def download_file(fid):
    req = svc.files().get_media(fileId=fid)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


import openpyxl
parent_id = find_folder("ETIQUETAS")
print(f"ETIQUETAS={parent_id}\n")

all_items = list_all(parent_id)
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
date_folders = [i for i in all_items if i["mimeType"]=="application/vnd.google-apps.folder" and DATE_RE.match(i["name"])]
date_folders.sort(key=lambda x:x["name"])

ship_seen = defaultdict(list)  # shipment_id → [(date, comp_pdf_name)]
mod_seen = defaultdict(list)   # composition_str → [date]

for df in date_folders:
    print(f"\n=== {df['name']} ===")
    files = list_all(df["id"])
    pdfs = [f for f in files if f["name"].lower().endswith(".pdf") and not f["name"].startswith("TODAS")]
    manifest = next((f for f in files if f["name"]=="manifest.xlsx"), None)

    if manifest:
        try:
            data = download_file(manifest["id"])
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                rows = list(ws.iter_rows(values_only=True))
                if not rows: continue
                hdr = rows[0]
                print(f"  manifest sheet '{sn}': {len(rows)-1} filas, cols: {hdr}")
                # Buscar columna shipment_id
                ship_col = None
                for i, h in enumerate(hdr or ()):
                    if h and isinstance(h,str) and ("shipment" in h.lower() or "envio" in h.lower() or "shipping" in h.lower()):
                        ship_col = i; break
                if ship_col is None and len(hdr or ()) > 0:
                    # primera columna podria tener IDs hex
                    ship_col = 0
                # Recolectar ids
                count = 0
                for r in rows[1:30]:
                    val = r[ship_col] if r and ship_col < len(r) else None
                    if val is not None:
                        ship_seen[str(val)].append(df["name"])
                        count += 1
                if count:
                    print(f"    ej. shipment_ids: {[str(r[ship_col]) for r in rows[1:4] if r]}")
        except Exception as e:
            print(f"  manifest err: {e}")

    # composition pdfs (filename matching)
    for p in pdfs:
        nm = p["name"].rsplit(".pdf",1)[0]
        mod_seen[nm].append(df["name"])

print(f"\n{'='*60}\n=== DUPLICADOS shipments ===")
ship_dups = {sid:dates for sid,dates in ship_seen.items() if len(set(dates))>1}
print(f"Shipments en >1 fecha: {len(ship_dups)}")
for sid, dates in list(ship_dups.items())[:20]:
    print(f"  {sid} → {dates}")

print(f"\n=== DUPLICADOS composition (filename) ===")
mod_dups = {n:dates for n,dates in mod_seen.items() if len(set(dates))>1}
print(f"Composiciones en >1 fecha: {len(mod_dups)}")
for n, dates in list(mod_dups.items())[:30]:
    print(f"  {n} → {dates}")
