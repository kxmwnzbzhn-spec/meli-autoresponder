#!/usr/bin/env python3
"""Cierre mensual contabilidad MELI: por modelo + devoluciones por motivo + reembolsos."""
import os, requests, json, sys, re
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

OUTPUT = sys.argv[1] if len(sys.argv) > 1 else "contabilidad_meli.xlsx"
MONTH = os.environ.get("ACCOUNTING_MONTH","")  # YYYY-MM CDMX
if not MONTH:
    cdmx = datetime.now(timezone.utc) - timedelta(hours=6)
    last_day_prev = cdmx.replace(day=1) - timedelta(days=1)
    MONTH = last_day_prev.strftime("%Y-%m")
print(f"Cierre mensual: {MONTH}")

APP_ID = os.environ["MELI_APP_ID"]; APP_SECRET = os.environ["MELI_APP_SECRET"]

ACCOUNTS = [
    ("JUAN", "MELI_REFRESH_TOKEN"),
    ("CLARIBEL", "MELI_REFRESH_TOKEN_CLARIBEL"),
    ("ASVA", "MELI_REFRESH_TOKEN_ASVA"),
    ("RAYMUNDO", "MELI_REFRESH_TOKEN_RAYMUNDO"),
    ("DILCIE", "MELI_REFRESH_TOKEN_DILCIE"),
    ("MILDRED", "MELI_REFRESH_TOKEN_MILDRED"),
]

# Date range: full month
year, month = MONTH.split("-")
year = int(year); month = int(month)
start_cdmx = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone(timedelta(hours=-6)))
if month == 12:
    end_cdmx = datetime(year+1, 1, 1, tzinfo=timezone(timedelta(hours=-6)))
else:
    end_cdmx = datetime(year, month+1, 1, tzinfo=timezone(timedelta(hours=-6)))
date_from = start_cdmx.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
date_to = end_cdmx.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")

REASON_LABELS = {
    "PDD9943": "No es original",
    "PDD9944": "Producto defectuoso/dañado",
    "PNR9501": "No llegó el producto",
    "PDD8975": "Producto incompleto",
    "PDD9945": "Producto distinto al publicado",
    "PDD8974": "Diferente al pedido",
    "PDD8973": "No le gustó",
    "PDR9526": "No recibió el producto",
    "PDD9939": "Diferente al publicado",
}

def categorize_product(title):
    """Devuelve (categoria, modelo). Categorias: Bocina, Perfume, Audífonos, Otro"""
    tl = title.lower()
    # Bocinas JBL por modelo
    if "go 4" in tl or "go4" in tl: return ("Bocina", "JBL Go 4")
    if "go 3" in tl or "go3" in tl: return ("Bocina", "JBL Go 3")
    if "go essential" in tl: return ("Bocina", "JBL Go Essential")
    if "flip 7" in tl or "flip7" in tl: return ("Bocina", "JBL Flip 7")
    if "flip 6" in tl: return ("Bocina", "JBL Flip 6")
    if "charge 6" in tl or "charge6" in tl: return ("Bocina", "JBL Charge 6")
    if "charge 5" in tl: return ("Bocina", "JBL Charge 5")
    if "grip" in tl and "jbl" in tl: return ("Bocina", "JBL Grip")
    if "clip 5" in tl: return ("Bocina", "JBL Clip 5")
    if "srs-xb100" in tl or "xb100" in tl: return ("Bocina", "Sony XB100")
    if "bocina" in tl or "parlante" in tl or "altavoz" in tl or "speaker" in tl: return ("Bocina", "Otra bocina")
    # Perfumes
    if "armaf" in tl: 
        # Detect Armaf model
        if "club de nuit" in tl: return ("Perfume", "Armaf Club De Nuit")
        if "odyssey" in tl: return ("Perfume", "Armaf Odyssey")
        if "delight" in tl: return ("Perfume", "Armaf Delight")
        if "niche" in tl: return ("Perfume", "Armaf Niche")
        return ("Perfume", "Armaf otros")
    if "lattafa" in tl: return ("Perfume", "Lattafa")
    if "mugler" in tl or "angel nova" in tl: return ("Perfume", "Mugler/Angel")
    if "perfume" in tl or "edp" in tl or "edt" in tl: return ("Perfume", "Otro perfume")
    # Audífonos
    if "redmi buds" in tl or "buds" in tl or "auriculares" in tl: return ("Audífonos", "Auriculares")
    return ("Otro", title[:30])

# ====== Collect data ======
print(f"  Range: {date_from} → {date_to}")
account_totals = {}
by_model = defaultdict(lambda: {"qty":0, "revenue":0, "category":"", "model":"", "accounts":set()})
returns_by_reason = defaultdict(lambda: {"count":0, "amount":0, "products":[]})
total_returns_amount = 0
total_returns_count = 0

for label, env_var in ACCOUNTS:
    RT = os.environ.get(env_var,"")
    if not RT: continue
    try:
        r = requests.post("https://api.mercadolibre.com/oauth/token",data={
            "grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":RT
        }).json()
        H = {"Authorization":f"Bearer {r['access_token']}"}
        me = requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=10).json()
        USER_ID = me["id"]
    except: continue
    
    a_orders = a_qty = 0; a_bruto = a_fees = a_ship = 0
    a_ret_cnt = a_ret_amt = 0
    
    offset = 0
    while True:
        rr = requests.get(
            f"https://api.mercadolibre.com/orders/search?seller={USER_ID}&order.date_created.from={date_from}&order.date_created.to={date_to}&limit=50&offset={offset}",
            headers=H, timeout=20
        ).json()
        results = rr.get("results",[])
        if not results: break
        for o in results:
            st = o.get("status","")
            order_id = o.get("id")
            amt = o.get("total_amount",0) or 0
            
            if st == "cancelled":
                a_ret_cnt += 1; a_ret_amt += amt
                # Look up associated claim if any
                items_o = o.get("order_items",[])
                prod = items_o[0].get("item",{}).get("title","")[:60] if items_o else ""
                # Try get claim reason from /post-purchase/v1/claims/search?resource_id=order_id
                reason = "cancelled (sin claim)"
                try:
                    cs = requests.get(f"https://api.mercadolibre.com/post-purchase/v1/claims/search?resource_id={order_id}",headers=H,timeout=10).json()
                    cdata = cs.get("data") or []
                    if cdata:
                        rid = cdata[0].get("reason_id","")
                        reason = REASON_LABELS.get(rid, rid or "cancelled")
                except: pass
                returns_by_reason[reason]["count"] += 1
                returns_by_reason[reason]["amount"] += amt
                returns_by_reason[reason]["products"].append((prod, label, order_id, amt))
                continue
            
            if st not in ("paid","shipped","delivered"): continue
            a_orders += 1; a_bruto += amt
            
            try:
                od = requests.get(f"https://api.mercadolibre.com/orders/{order_id}",headers=H,timeout=10).json()
                for pay in od.get("payments",[]):
                    if pay.get("status") == "approved":
                        a_fees += (pay.get("marketplace_fee",0) or 0)
                sh_id = od.get("shipping",{}).get("id")
                if sh_id:
                    sd = requests.get(f"https://api.mercadolibre.com/shipments/{sh_id}",headers=H,timeout=10).json()
                    so = sd.get("shipping_option",{}) or {}
                    a_ship += max(0, (so.get("list_cost",0) or 0) - (so.get("cost",0) or 0))
            except: pass
            
            for oi in o.get("order_items",[]):
                qty = oi.get("quantity",0)
                a_qty += qty
                title = oi.get("item",{}).get("title","")
                cat, mdl = categorize_product(title)
                rev = qty * (oi.get("unit_price",0) or 0)
                by_model[mdl]["qty"] += qty
                by_model[mdl]["revenue"] += rev
                by_model[mdl]["category"] = cat
                by_model[mdl]["model"] = mdl
                by_model[mdl]["accounts"].add(label)
        offset += 50
        if offset >= rr.get("paging",{}).get("total",0): break
    
    a_iva = a_bruto - (a_bruto/1.16) if a_bruto > 0 else 0
    a_net = a_bruto - a_fees - a_ship
    a_net_after_iva = a_net - a_iva
    account_totals[label] = {
        "nickname": me.get("nickname",""),
        "orders": a_orders, "qty": a_qty, "bruto": a_bruto,
        "fees": a_fees, "ship": a_ship, "iva": a_iva,
        "ret_count": a_ret_cnt, "ret_amount": a_ret_amt,
        "net": a_net, "net_after_iva": a_net_after_iva
    }
    total_returns_amount += a_ret_amt; total_returns_count += a_ret_cnt
    print(f"  {label}: {a_orders} ord / {a_qty}u / ${a_bruto:.0f}")

# ====== Build / load XLSX ======
if os.path.exists(OUTPUT):
    wb = load_workbook(OUTPUT)
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
NUM_FMT = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
INT_FMT = '#,##0;[Red](#,##0);"-"'
PCT_FMT = '0.0%;[Red](0.0%);"-"'
BORDER = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

def sh(c):
    c.fill=HDR_FILL; c.font=HDR_FONT
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=BORDER

# Sheet "Mensual YYYY-MM"
sname = f"Mensual {MONTH}"
if sname in wb.sheetnames: del wb[sname]
ms = wb.create_sheet(sname)

# === SECTION 1: KPIs Mensuales ===
ms.cell(row=1,column=1,value=f"📊 CIERRE MENSUAL — {MONTH}").font = Font(bold=True, size=14, color="1F4E78")
ms.merge_cells("A1:F1")

t_o = sum(d["orders"] for d in account_totals.values())
t_q = sum(d["qty"] for d in account_totals.values())
t_b = sum(d["bruto"] for d in account_totals.values())
t_f = sum(d["fees"] for d in account_totals.values())
t_s = sum(d["ship"] for d in account_totals.values())
t_iva = sum(d["iva"] for d in account_totals.values())
t_net = t_b - t_f - t_s
t_net_iva = t_net - t_iva

kpis = [
    ("Órdenes Totales", t_o, INT_FMT),
    ("Unidades Vendidas", t_q, INT_FMT),
    ("Bruto", t_b, NUM_FMT),
    ("Comisión MELI", t_f, NUM_FMT),
    ("Envío Seller", t_s, NUM_FMT),
    ("IVA Remite", t_iva, NUM_FMT),
    ("Devoluciones #", total_returns_count, INT_FMT),
    ("Devoluciones $", total_returns_amount, NUM_FMT),
    ("NET (sin IVA)", t_net, NUM_FMT),
    ("NET (post IVA)", t_net_iva, NUM_FMT),
    ("Margen %", t_net/t_b if t_b else 0, PCT_FMT),
]
ms.cell(row=3,column=1,value="KPI"); sh(ms.cell(row=3,column=1))
ms.cell(row=3,column=2,value="Valor"); sh(ms.cell(row=3,column=2))
for i,(k,v,fmt) in enumerate(kpis,4):
    ms.cell(row=i,column=1,value=k).font = Font(bold=True)
    c = ms.cell(row=i,column=2,value=v); c.number_format = fmt
    c.alignment = Alignment(horizontal="right")

# === SECTION 2: Por Cuenta ===
start = len(kpis)+5
ms.cell(row=start,column=1,value="📊 POR CUENTA").font = Font(bold=True, size=12, color="1F4E78")
hdr_acc = ["Cuenta","Nickname","Órdenes","Unidades","Bruto","Comisión","Envío","IVA","Devol $","NET","Margen"]
for col,h in enumerate(hdr_acc,1): sh(ms.cell(row=start+1,column=col,value=h))
r = start+2
for label,d in sorted(account_totals.items(), key=lambda x: -x[1]["bruto"]):
    if d["orders"]==0 and d["ret_count"]==0: continue
    ms.cell(row=r,column=1,value=label).font = Font(bold=True)
    ms.cell(row=r,column=2,value=d["nickname"])
    ms.cell(row=r,column=3,value=d["orders"]).number_format = INT_FMT
    ms.cell(row=r,column=4,value=d["qty"]).number_format = INT_FMT
    ms.cell(row=r,column=5,value=d["bruto"]).number_format = NUM_FMT
    ms.cell(row=r,column=6,value=d["fees"]).number_format = NUM_FMT
    ms.cell(row=r,column=7,value=d["ship"]).number_format = NUM_FMT
    ms.cell(row=r,column=8,value=d["iva"]).number_format = NUM_FMT
    ms.cell(row=r,column=9,value=d["ret_amount"]).number_format = NUM_FMT
    ms.cell(row=r,column=10,value=f"=E{r}-F{r}-G{r}").number_format = NUM_FMT
    ms.cell(row=r,column=11,value=f"=IFERROR(J{r}/E{r},0)").number_format = PCT_FMT
    r += 1

# === SECTION 3: Por Modelo (Bocinas + Perfumes + Otros) ===
start_m = r+2
ms.cell(row=start_m,column=1,value="🎵 BOCINAS POR MODELO").font = Font(bold=True, size=12, color="1F4E78")
hdr_m = ["Modelo","Categoría","Unidades","Revenue","% Total"]
for col,h in enumerate(hdr_m,1): sh(ms.cell(row=start_m+1,column=col,value=h))
r = start_m+2
boc_models = {k:v for k,v in by_model.items() if v["category"]=="Bocina"}
for mdl,info in sorted(boc_models.items(), key=lambda x: -x[1]["revenue"]):
    ms.cell(row=r,column=1,value=mdl)
    ms.cell(row=r,column=2,value=info["category"])
    ms.cell(row=r,column=3,value=info["qty"]).number_format = INT_FMT
    ms.cell(row=r,column=4,value=info["revenue"]).number_format = NUM_FMT
    ms.cell(row=r,column=5,value=f"=IFERROR(D{r}/{t_b},0)").number_format = PCT_FMT
    r += 1

start_p = r+2
ms.cell(row=start_p,column=1,value="🌸 PERFUMES POR MARCA").font = Font(bold=True, size=12, color="1F4E78")
for col,h in enumerate(hdr_m,1): sh(ms.cell(row=start_p+1,column=col,value=h))
r = start_p+2
perf_models = {k:v for k,v in by_model.items() if v["category"]=="Perfume"}
for mdl,info in sorted(perf_models.items(), key=lambda x: -x[1]["revenue"]):
    ms.cell(row=r,column=1,value=mdl)
    ms.cell(row=r,column=2,value=info["category"])
    ms.cell(row=r,column=3,value=info["qty"]).number_format = INT_FMT
    ms.cell(row=r,column=4,value=info["revenue"]).number_format = NUM_FMT
    ms.cell(row=r,column=5,value=f"=IFERROR(D{r}/{t_b},0)").number_format = PCT_FMT
    r += 1

# Otros
start_o = r+2
ms.cell(row=start_o,column=1,value="🎧 OTROS PRODUCTOS").font = Font(bold=True, size=12, color="1F4E78")
for col,h in enumerate(hdr_m,1): sh(ms.cell(row=start_o+1,column=col,value=h))
r = start_o+2
other_models = {k:v for k,v in by_model.items() if v["category"] not in ("Bocina","Perfume")}
for mdl,info in sorted(other_models.items(), key=lambda x: -x[1]["revenue"]):
    ms.cell(row=r,column=1,value=mdl)
    ms.cell(row=r,column=2,value=info["category"])
    ms.cell(row=r,column=3,value=info["qty"]).number_format = INT_FMT
    ms.cell(row=r,column=4,value=info["revenue"]).number_format = NUM_FMT
    ms.cell(row=r,column=5,value=f"=IFERROR(D{r}/{t_b},0)").number_format = PCT_FMT
    r += 1

# === SECTION 4: Devoluciones por Motivo ===
start_d = r+2
ms.cell(row=start_d,column=1,value="↩️ DEVOLUCIONES POR MOTIVO").font = Font(bold=True, size=12, color="C00000")
hdr_d = ["Motivo","# Devoluciones","Monto Reembolsado","% del Total Bruto"]
for col,h in enumerate(hdr_d,1): sh(ms.cell(row=start_d+1,column=col,value=h))
r = start_d+2
for reason,info in sorted(returns_by_reason.items(), key=lambda x: -x[1]["amount"]):
    ms.cell(row=r,column=1,value=reason)
    ms.cell(row=r,column=2,value=info["count"]).number_format = INT_FMT
    ms.cell(row=r,column=3,value=info["amount"]).number_format = NUM_FMT
    ms.cell(row=r,column=4,value=f"=IFERROR(C{r}/{t_b},0)").number_format = PCT_FMT
    r += 1
# Total devoluciones
ms.cell(row=r,column=1,value="TOTAL").font = Font(bold=True)
ms.cell(row=r,column=2,value=total_returns_count).number_format = INT_FMT
ms.cell(row=r,column=3,value=total_returns_amount).number_format = NUM_FMT

# Column widths
for col,w in enumerate([22, 18, 14, 14, 14, 14, 14, 14, 14, 14, 12], 1):
    ms.column_dimensions[get_column_letter(col)].width = w

# === Charts ===
# Chart 1: Revenue por modelo (Bar)
if boc_models:
    bc = BarChart()
    bc.type = "bar"
    bc.title = f"Bocinas por modelo — {MONTH}"
    bc.width = 16; bc.height = 10
    rng_d = Reference(ms, min_col=4, min_row=start_m+1, max_row=start_m+1+len(boc_models))
    cat_d = Reference(ms, min_col=1, min_row=start_m+2, max_row=start_m+1+len(boc_models))
    bc.add_data(rng_d, titles_from_data=True)
    bc.set_categories(cat_d)
    ms.add_chart(bc, "M3")

# Chart 2: Pie devoluciones por motivo
if returns_by_reason:
    pc = PieChart()
    pc.title = f"Devoluciones por motivo — {MONTH}"
    pc.width = 14; pc.height = 10
    rng_d = Reference(ms, min_col=3, min_row=start_d+1, max_row=start_d+1+len(returns_by_reason))
    cat_d = Reference(ms, min_col=1, min_row=start_d+2, max_row=start_d+1+len(returns_by_reason))
    pc.add_data(rng_d, titles_from_data=True)
    pc.set_categories(cat_d)
    pc.dataLabels = DataLabelList(showPercent=True)
    ms.add_chart(pc, "M25")

wb.save(OUTPUT)
print(f"\n✅ Cierre mensual {MONTH} guardado en {OUTPUT}")
print(f"   Sheet: '{sname}'")
print(f"   Bruto: ${t_b:,.0f}")
print(f"   NET (post IVA): ${t_net_iva:,.0f}")
print(f"   Devoluciones: {total_returns_count} órdenes / ${total_returns_amount:,.0f}")
print(f"   Modelos vendidos: {len(by_model)}")
