"""
Excel cuentas_meli.xlsx — todas las cuentas activas con datos en vivo.
"""
import os, requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta

APP_ID = os.environ["MELI_APP_ID"]; APP_SECRET = os.environ["MELI_APP_SECRET"]

ACCOUNTS = [
    ("JUAN",     "MELI_REFRESH_TOKEN",          "Juan Pedraza",       "MercadoLider — JBL/Sony", "70u/dia"),
    ("CLARIBEL", "MELI_REFRESH_TOKEN_CLARIBEL", "Claribel",           "Perfumes + bocinas",       "70/80/90/100/150 ramp"),
    ("ASVA",     "MELI_REFRESH_TOKEN_ASVA",     "ASVA Electronics",   "Cuenta principal autoridad","sin cap"),
    ("RAYMUNDO", "MELI_REFRESH_TOKEN_RAYMUNDO", "Raymundo",           "Catalog WINNER bocinas",   "sin cap (winner)"),
    ("DILCIE",   "MELI_REFRESH_TOKEN_DILCIE",   "Dilcie",             "Cuenta nueva",             "70u/dia"),
    ("MILDRED",  "MELI_REFRESH_TOKEN_MILDRED",  "Mildred",            "Bocinas + perfumes",       "50u/dia inicio"),
    ("YC_NEW",   "MELI_REFRESH_TOKEN_YC_NEW",   "Cuenta nueva (YC)",  "Pendiente configurar",     "sin cap"),
    ("BREN",     "MELI_REFRESH_TOKEN_BREN",     "Bren Castillo",      "Cuenta nueva",             "50u/dia"),
]

now_cdmx = (datetime.now(timezone.utc) - timedelta(hours=6)).strftime("%Y-%m-%d %H:%M CDMX")
midnight_cdmx_utc = (datetime.now(timezone.utc) - timedelta(hours=6)).replace(hour=0,minute=0,second=0,microsecond=0).astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")

rows = []
for label, env_var, owner, role, cap in ACCOUNTS:
    RT = os.environ.get(env_var, "")
    row = {
        "label": label,
        "owner": owner,
        "secret": env_var,
        "role": role,
        "cap": cap,
        "user_id": "—",
        "nickname": "—",
        "email": "—",
        "reputation": "—",
        "items_total": "—",
        "items_active": "—",
        "items_paused": "—",
        "ventas_hoy": "—",
        "bruto_hoy": "—",
        "status": "—",
    }
    if not RT:
        row["status"] = "❌ Sin token"
        rows.append(row); continue
    
    try:
        r = requests.post("https://api.mercadolibre.com/oauth/token",
            data={"grant_type":"refresh_token","client_id":APP_ID,"client_secret":APP_SECRET,"refresh_token":RT},
            timeout=15).json()
        H = {"Authorization":f"Bearer {r['access_token']}"}
        me = requests.get("https://api.mercadolibre.com/users/me",headers=H,timeout=10).json()
        row["user_id"] = me.get("id","—")
        row["nickname"] = me.get("nickname","")
        row["email"] = me.get("email","") or "—"
        rep = (me.get("seller_reputation") or {}).get("level_id") or "nueva"
        row["reputation"] = rep
        
        # Item counts
        ic = requests.get(f"https://api.mercadolibre.com/users/{me['id']}/items/search?status=active&limit=1",headers=H,timeout=10).json()
        row["items_active"] = ic.get("paging",{}).get("total",0)
        ic2 = requests.get(f"https://api.mercadolibre.com/users/{me['id']}/items/search?status=paused&limit=1",headers=H,timeout=10).json()
        row["items_paused"] = ic2.get("paging",{}).get("total",0)
        row["items_total"] = row["items_active"] + row["items_paused"]
        
        # Ventas hoy CDMX
        sales = 0; bruto = 0
        offset = 0
        while True:
            rr = requests.get(f"https://api.mercadolibre.com/orders/search?seller={me['id']}&order.date_created.from={midnight_cdmx_utc}&limit=50&offset={offset}",headers=H,timeout=20).json()
            res = rr.get("results",[])
            if not res: break
            for o in res:
                if o.get("status") in ("paid","shipped","delivered"):
                    bruto += o.get("total_amount",0) or 0
                    for oi in o.get("order_items",[]): sales += oi.get("quantity",0)
            offset += 50
            if offset >= rr.get("paging",{}).get("total",0): break
        row["ventas_hoy"] = sales
        row["bruto_hoy"] = bruto
        row["status"] = "✅ Activa"
    except Exception as e:
        row["status"] = f"❌ {str(e)[:30]}"
    rows.append(row)

# Build XLSX
wb = Workbook(); wb.remove(wb.active)
ws = wb.create_sheet("Cuentas MELI")

ws["A1"] = f"📋 CUENTAS MELI — {now_cdmx}"
ws["A1"].font = Font(bold=True, size=16, color="1F4E78"); ws.merge_cells("A1:N1")

HEADER = ["#","Nombre","Cuenta MELI","User ID","Owner","Email","Reputación","Items act","Items paus","Items tot","Ventas hoy","Bruto hoy","Cap diario","Status","Rol"]
for col_idx, h in enumerate(HEADER, 1):
    c = ws.cell(row=3, column=col_idx, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E78")
    c.alignment = Alignment(horizontal="center", vertical="center")

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")

r = 4
for i, row in enumerate(rows, 1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=row["label"])
    ws.cell(row=r, column=3, value=row["nickname"])
    ws.cell(row=r, column=4, value=str(row["user_id"]))
    ws.cell(row=r, column=5, value=row["owner"])
    ws.cell(row=r, column=6, value=row["email"])
    ws.cell(row=r, column=7, value=row["reputation"])
    ws.cell(row=r, column=8, value=row["items_active"])
    ws.cell(row=r, column=9, value=row["items_paused"])
    ws.cell(row=r, column=10, value=row["items_total"])
    ws.cell(row=r, column=11, value=row["ventas_hoy"])
    cell_b = ws.cell(row=r, column=12, value=row["bruto_hoy"])
    if isinstance(row["bruto_hoy"], (int,float)):
        cell_b.number_format = '"$"#,##0.00'
    ws.cell(row=r, column=13, value=row["cap"])
    cell_s = ws.cell(row=r, column=14, value=row["status"])
    if "✅" in str(row["status"]): cell_s.fill = GREEN
    elif "❌" in str(row["status"]): cell_s.fill = RED
    else: cell_s.fill = YELLOW
    ws.cell(row=r, column=15, value=row["role"])
    r += 1

# Totales
ws.cell(row=r, column=2, value="TOTAL").font = Font(bold=True)
ws.cell(row=r, column=8, value=sum((x["items_active"] for x in rows if isinstance(x["items_active"],int)), 0)).font = Font(bold=True)
ws.cell(row=r, column=9, value=sum((x["items_paused"] for x in rows if isinstance(x["items_paused"],int)), 0)).font = Font(bold=True)
ws.cell(row=r, column=10, value=sum((x["items_total"] for x in rows if isinstance(x["items_total"],int)), 0)).font = Font(bold=True)
ws.cell(row=r, column=11, value=sum((x["ventas_hoy"] for x in rows if isinstance(x["ventas_hoy"],int)), 0)).font = Font(bold=True)
total_bruto = sum((x["bruto_hoy"] for x in rows if isinstance(x["bruto_hoy"],(int,float))), 0)
b = ws.cell(row=r, column=12, value=total_bruto)
b.font = Font(bold=True); b.number_format = '"$"#,##0.00'

widths = [4, 12, 24, 14, 22, 26, 16, 10, 10, 10, 12, 14, 22, 14, 35]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Hoja 2: Configuración
ws2 = wb.create_sheet("Configuración App")
ws2["A1"] = "⚙️ Configuración App MELI"
ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
config = [
    ("App ID", "5211907102822632"),
    ("Client Secret", "GitHub Secret MELI_APP_SECRET"),
    ("Redirect URI", "https://oauth.pstmn.io/v1/callback"),
    ("", ""),
    ("Link autorización nueva cuenta", ""),
    ("URL", "https://auth.mercadolibre.com.mx/authorization?response_type=code&client_id=5211907102822632&redirect_uri=https://oauth.pstmn.io/v1/callback"),
    ("", ""),
    ("Cron contabilidad", "5 6 * * * UTC = 00:05 CDMX (todos los días)"),
    ("Cron throttle", "*/5 min — pausa al llegar a cap"),
    ("Cron reactivate", "0 15 UTC = 09:00 CDMX"),
    ("Cron catalog war", "*/5 min — Raymundo siempre gana"),
]
r = 3
for k, v in config:
    ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
    ws2.cell(row=r, column=2, value=v)
    r += 1
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 90

wb.save("cuentas_meli.xlsx")
print(f"✅ cuentas_meli.xlsx creado con {len(rows)} cuentas")
for row in rows:
    print(f"  {row['label']:10} {row['nickname']:24} | items={row['items_total']} | ventas hoy={row['ventas_hoy']} | {row['status']}")
