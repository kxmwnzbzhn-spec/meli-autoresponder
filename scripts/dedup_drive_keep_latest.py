#!/usr/bin/env python3
"""DEDUP Drive ETIQUETAS: política simple — conservar solo la carpeta de fecha
mas RECIENTE (hoy). Las anteriores son obsoletas porque:
- shipments que ya se enviaron → label ya usado, no necesario
- shipments aun pendientes → estan en la carpeta nueva regenerados

Antes de borrar viejas, compara shipment_ids vs hoy y reporta cuantos
shipments del folder viejo NO estan en hoy (cuantos se perderian).
Si DRY_RUN=1 solo reporta. Si KEEP_DAYS=N conserva las ultimas N fechas.
"""
import os, io, sys, re
from datetime import datetime
from google.oauth2.credentials import Credentials as UC
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl

cid = os.environ["GOOGLE_OAUTH_CLIENT_ID"]
csec = os.environ["GOOGLE_OAUTH_CLIENT_SECRET"]
rt = os.environ["GOOGLE_OAUTH_REFRESH_TOKEN"]
DRY_RUN = os.environ.get("DRY_RUN","0") == "1"
KEEP_DAYS = int(os.environ.get("KEEP_DAYS","1"))

creds = UC(token=None, refresh_token=rt,
    token_uri="https://oauth2.googleapis.com/token",
    client_id=cid, client_secret=csec,
    scopes=["https://www.googleapis.com/auth/drive"])
svc = build("drive","v3",credentials=creds,cache_discovery=False)


def find_folder(name, parent=None):
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent: q += f" and '{parent}' in parents"
    r = svc.files().list(q=q, fields="files(id,name)").execute()
    return (r.get("files",[]) or [{}])[0].get("id")


def list_all(parent_id):
    out, page = [], None
    while True:
        r = svc.files().list(q=f"'{parent_id}' in parents and trashed=false",
                             fields="nextPageToken,files(id,name,mimeType)",
                             pageSize=1000, pageToken=page).execute()
        out.extend(r.get("files",[]))
        page = r.get("nextPageToken")
        if not page: break
    return out


def download(fid):
    req = svc.files().get_media(fileId=fid)
    buf = io.BytesIO()
    d = MediaIoBaseDownload(buf, req)
    while True:
        _, done = d.next_chunk()
        if done: break
    return buf.getvalue()


def get_ship_ids(folder_id):
    """Lee manifest.xlsx 'Detalle envios' y devuelve set de shipment_ids."""
    files = list_all(folder_id)
    mf = next((f for f in files if f["name"]=="manifest.xlsx"), None)
    if not mf: return set()
    try:
        data = download(mf["id"])
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        if "Detalle envios" not in wb.sheetnames: return set()
        ws = wb["Detalle envios"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return set()
        hdr = rows[0]
        ship_col = next((i for i,h in enumerate(hdr) if h=="Shipment"), 2)
        return {str(r[ship_col]) for r in rows[1:] if r and ship_col<len(r) and r[ship_col]}
    except Exception as e:
        print(f"err manifest {folder_id}: {e}")
        return set()


parent_id = find_folder("ETIQUETAS")
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
all_items = list_all(parent_id)
date_folders = sorted(
    [i for i in all_items if i["mimeType"]=="application/vnd.google-apps.folder" and DATE_RE.match(i["name"])],
    key=lambda x: x["name"]
)
print(f"Date folders: {len(date_folders)}")
for f in date_folders: print(f"  {f['name']}")

if len(date_folders) <= KEEP_DAYS:
    print(f"\nSolo {len(date_folders)} carpetas, KEEP_DAYS={KEEP_DAYS}. Nada que borrar.")
    sys.exit(0)

# Las que conservamos: ultimas KEEP_DAYS
keep_folders = date_folders[-KEEP_DAYS:]
delete_folders = date_folders[:-KEEP_DAYS]
print(f"\nKEEP ({len(keep_folders)}): {[f['name'] for f in keep_folders]}")
print(f"DELETE ({len(delete_folders)}): {[f['name'] for f in delete_folders]}")

# Get shipment_ids del set keep
keep_ships = set()
for kf in keep_folders:
    s = get_ship_ids(kf["id"])
    print(f"  {kf['name']}: {len(s)} shipments")
    keep_ships |= s

# Para cada delete folder, comparar
print(f"\n=== Análisis de pérdida ===")
for df in delete_folders:
    old_ships = get_ship_ids(df["id"])
    only_in_old = old_ships - keep_ships
    overlap = old_ships & keep_ships
    print(f"  {df['name']}: {len(old_ships)} shipments | {len(overlap)} en keep | {len(only_in_old)} solo aquí")
    if only_in_old:
        print(f"    ⚠️ Shipments en {df['name']} que NO están en keep (probablemente ya enviados):")
        for s in list(only_in_old)[:10]:
            print(f"       {s}")
        if len(only_in_old) > 10:
            print(f"       ...+{len(only_in_old)-10} más")

if DRY_RUN:
    print(f"\n=== DRY RUN: no se borra nada ===")
    sys.exit(0)

# Mover a Trash (svc.files().delete = permanent. Mejor update trashed=True)
print(f"\n=== EJECUTANDO BORRADO ===")
deleted = 0
errs = []
for df in delete_folders:
    try:
        svc.files().update(fileId=df["id"], body={"trashed":True}).execute()
        print(f"  🗑️ Trashed {df['name']}")
        deleted += 1
    except Exception as e:
        errs.append({"folder":df["name"],"err":str(e)[:100]})
        print(f"  ❌ {df['name']}: {e}")

print(f"\n=== RESUMEN ===")
print(f"Borradas: {deleted}, Errores: {len(errs)}")

TG = os.environ.get("TELEGRAM_BOT_TOKEN","")
TGCID = os.environ.get("TELEGRAM_CHAT_ID","")
if TG and TGCID:
    import requests
    msg = f"🧹 *Drive ETIQUETAS limpio*\n\n"
    msg += f"📁 Conservadas (últimas {KEEP_DAYS}): {', '.join(f['name'] for f in keep_folders)}\n"
    msg += f"🗑️ Borradas: *{deleted}*\n"
    if errs: msg += f"❌ Errores: {len(errs)}\n"
    msg += f"\nDe ahora en adelante el bot conservará solo la carpeta del día (regenera todo)."
    requests.post(f"https://api.telegram.org/bot{TG}/sendMessage", data={
        "chat_id":TGCID,"parse_mode":"Markdown","text":msg[:4000]}, timeout=20)
